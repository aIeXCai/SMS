from django.shortcuts import render, redirect, get_object_or_404
from django.urls import reverse
from django.views.decorators.http import require_POST, require_http_methods
from django.contrib import messages # 用於顯示操作成功或失敗訊息
from django.db import transaction # 用於確保批量操作的原子性
from django.utils import timezone
import openpyxl # 導入 openpyxl 庫
from datetime import datetime
from django.db.models import Q
from django.http import HttpResponse, JsonResponse
import json
import io

from .models import Student, Class, GRADE_LEVEL_CHOICES, CLASS_NAME_CHOICES, STATUS_CHOICES
from .forms import StudentForm, ExcelUploadForm, BatchUpdateStatusForm, BatchPromoteGradeForm

# 學生列表頁面 (PRD 3.1.1)
def student_list(request):
    students = Student.objects.all()

    # --- 篩選邏輯開始 ---
    # 1. 全局搜索 (學號和姓名)
    search_query = request.GET.get('q') # 從 URL 參數獲取搜索關鍵字，例如: /students/?q=张三
    if search_query:
        students = students.filter(
            Q(student_id__icontains=search_query) | # 學號包含關鍵字 (不區分大小寫)
            Q(name__icontains=search_query)        # 姓名包含關鍵字 (不區分大小寫)
        )

    # 2. 班級名稱篩選
    # class_name 變數會接收 '1班', '2班' 等值
    class_name_filter = request.GET.get('class_name_filter') # 注意這裡的參數名，與前端對應
    if class_name_filter:
        # 篩選出所有符合該班級名稱的 Class 對象 (例如所有年級的 '1班')
        matching_classes = Class.objects.filter(class_name=class_name_filter)
        # 然後用這些班級來篩選學生
        students = students.filter(current_class__in=matching_classes)
            
    # 3. 年級篩選
    grade_level = request.GET.get('grade') # 從 URL 參數獲取年級
    if grade_level:
        students = students.filter(current_class__grade_level=grade_level)

    # 4. 狀態篩選
    status = request.GET.get('status') # 從 URL 參數獲取狀態
    if status:
        students = students.filter(status=status)
    # --- 篩選邏輯結束 ---

    return render(request, 'students/student_list.html', {
        'students': students,
        'status_choices': STATUS_CHOICES,        
        'grade_level_choices': GRADE_LEVEL_CHOICES, 
        'class_name_choices': CLASS_NAME_CHOICES,   # 傳遞全局班級名稱選項
        'search_query': search_query if search_query else '', 
        'selected_class_name_filter': class_name_filter if class_name_filter else '', # 將選中的班級名稱傳回模板
        'selected_grade': grade_level if grade_level else '', 
        'selected_status': status if status else '',  
    })

# 新增學生頁面 (PRD 3.1.3)
@require_http_methods(["GET", "POST"])
def student_add(request):
    if request.method == 'POST':
        form = StudentForm(request.POST)
        if form.is_valid():
            form.save() # StudentForm 的 save 方法會處理 Class 的查找或創建
            messages.success(request, "学生新增成功！")
            return redirect('student_list')
        else:
            messages.error(request, "新增学生失败，请检查表单数据。")
    else:
        form = StudentForm()
    
    # 傳遞年級和班級選項給模板，以便前端可以動態加載（如果未來實現聯動）
    return render(request, 'students/student_form.html', {
        'form': form,
        'title': '新增学生',
        'grade_level_choices': GRADE_LEVEL_CHOICES, # 傳遞給模板
        'class_name_choices': CLASS_NAME_CHOICES,   # 傳遞給模板
    })

# 修改學生頁面 (PRD 3.1.4)
@require_http_methods(["GET", "POST"])
def student_edit(request, pk):
    student = get_object_or_404(Student, pk=pk)
    if request.method == 'POST':
        form = StudentForm(request.POST, instance=student)
        if form.is_valid():
            form.save() # StudentForm 的 save 方法會處理 Class 的查找或創建
            messages.success(request, "学生信息更新成功！")
            return redirect('student_list')
        else:
            messages.error(request, "更新学生信息失败，请检查表单数据。")
    else:
        form = StudentForm(instance=student)
    
    return render(request, 'students/student_form.html', {
        'form': form,
        'title': '编辑学生',
        'grade_level_choices': GRADE_LEVEL_CHOICES, # 傳遞給模板
        'class_name_choices': CLASS_NAME_CHOICES,   # 傳遞給模板
    })

# 刪除學生功能 (PRD 3.1.4)
def student_delete(request, pk):
    student = get_object_or_404(Student, pk=pk)
    if request.method == 'POST':
        student.delete()
        return redirect('student_list') # 刪除成功後跳轉回列表頁
    return render(request, 'students/student_confirm_delete.html', {'student': student})

# 學生狀態一鍵切換功能 (PRD 3.1.5)
@require_POST # 确保这个视图只接受 POST 请求，增强安全性
def student_update_status(request, pk):
    student = get_object_or_404(Student, pk=pk)
    
    # 從 POST 數據中獲取新的狀態
    new_status = request.POST.get('status') 
    
    # 檢查新狀態是否在 Student 模型定義的有效選項中
    valid_statuses = [choice[0] for choice in STATUS_CHOICES]
    if new_status and new_status in valid_statuses:
        student.status = new_status
        # 如果狀態是「畢業」，自動設定畢業日期 (根據 PRD 3.1.7 預留的邏輯，這裡可以做簡易處理)
        if new_status == '畢業' and not student.graduation_date:
            from django.utils import timezone
            student.graduation_date = timezone.now().date() # 設置為當前日期
        
        student.save()
        # 可以添加 Django messages 提示操作成功，這裡暫不展開
        # from django.contrib import messages
        # messages.success(request, f"{student.name} 的狀態已更新為 {new_status}。")
    
    return redirect('student_list') # 更新後導回學生列表頁面

# 批量導入學生 (從 Excel) - PRD 3.1.2 擴展
@require_http_methods(["GET", "POST"])
def student_batch_import(request):
    if request.method == 'POST':
        form = ExcelUploadForm(request.POST, request.FILES)
        if form.is_valid():
            excel_file = request.FILES['excel_file']
            
            # 检查文件类型，确保是 Excel 文件
            if not excel_file.name.endswith(('.xlsx', '.xls')):
                return JsonResponse({
                    'success': False,
                    'error': "文件格式不正确，请上传 .xlsx 或 .xls 文件。"
                })

            try:
                workbook = openpyxl.load_workbook(excel_file)
                sheet = workbook.active
                
                header = [cell.value for cell in sheet[1]] # 获取第一行的标题
                
                # 定义一个映射，将Excel中的列名映射到Student模型的字段名
                header_mapping = {
                    "学号 (必填)": "student_id",
                    "姓名 (必填)": "name",
                    "性别 (男/女)": "gender",
                    "出生日期 (YYYY-MM-DD)": "date_of_birth",
                    "年级 (初一/初二/初三/高一/高二/高三)": "grade_level",
                    "班级名称 (1班-20班)": "class_name",
                    "在校状态 (在读/转学/休学/复学/毕业)": "status",
                    "身份证号码": "id_card_number",
                    "学籍号": "student_enrollment_number",
                    "家庭地址": "home_address",
                    "监护人姓名": "guardian_name",
                    "监护人联系电话": "guardian_contact_phone",
                    "入学日期 (YYYY-MM-DD)": "entry_date",
                    "毕业日期 (YYYY-MM-DD, 毕业状态必填)": "graduation_date",
                }

                imported_count = 0
                failed_rows = []
                success_messages = []
                error_messages = []
                warning_messages = []

                # 逐行处理，不使用整体事务
                for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                    # 跳过空行
                    if not any(row):
                        continue
                        
                    try:
                        with transaction.atomic():  # 每行单独事务
                            row_data = dict(zip(header, row))
                            
                            student_data = {}
                            for excel_header, model_field in header_mapping.items():
                                if excel_header in row_data:
                                    student_data[model_field] = row_data[excel_header]

                            # 验证必填字段
                            if not student_data.get('student_id') or not student_data.get('name'):
                                raise ValueError("学号和姓名为必填字段")
                            
                            # 确保学号为字符串格式
                            if student_data.get('student_id'):
                                student_data['student_id'] = str(student_data['student_id']).strip()
                            
                            # 确保身份证号码和学籍号为字符串格式
                            if student_data.get('id_card_number'):
                                student_data['id_card_number'] = str(student_data['id_card_number']).strip()
                            if student_data.get('student_enrollment_number'):
                                student_data['student_enrollment_number'] = str(student_data['student_enrollment_number']).strip()

                            # 处理日期字段的格式转换
                            for date_field in ['date_of_birth', 'entry_date', 'graduation_date']:
                                if date_field in student_data and student_data[date_field]:
                                    try:
                                        if isinstance(student_data[date_field], datetime):
                                            # openpyxl 可能会直接读取为 datetime 对象
                                            student_data[date_field] = student_data[date_field].date()
                                        else:
                                            # 尝试从字符串解析日期
                                            student_data[date_field] = datetime.strptime(str(student_data[date_field]), '%Y-%m-%d').date()
                                    except ValueError:
                                        student_data[date_field] = None
                                        warning_messages.append(f"第 {row_idx} 行的 '{date_field}' 日期格式不正确，已设置为空。")

                            # 处理性别字段的容错解析
                            if 'gender' in student_data and student_data['gender']:
                                gender_value = str(student_data['gender']).strip()
                                gender_mapping = {
                                    '男': '男', 'M': '男', 'Male': '男', 'male': '男', '1': '男',
                                    '女': '女', 'F': '女', 'Female': '女', 'female': '女', '0': '女'
                                }
                                student_data['gender'] = gender_mapping.get(gender_value, gender_value)
                                
                                # 验证性别值是否有效
                                if student_data['gender'] not in ['男', '女']:
                                    warning_messages.append(f"第 {row_idx} 行的性别值 '{gender_value}' 无效，已设置为空")
                                    student_data['gender'] = None

                            # 从 student_data 中提取 grade_level 和 class_name，用于查找或创建 Class 对象
                            grade_level = student_data.pop('grade_level', None)
                            class_name = student_data.pop('class_name', None)
                            current_class_obj = None

                            if grade_level and class_name:
                                try:
                                    current_class_obj, created = Class.objects.get_or_create(
                                        grade_level=grade_level,
                                        class_name=class_name
                                    )
                                    if created:
                                        success_messages.append(f"自动创建班级：{grade_level}{class_name}")
                                except Exception as e:
                                    raise ValueError(f"班级创建/查找失败: {e}")

                            student_data['current_class'] = current_class_obj
                            student_data['grade_level'] = grade_level

                            # 尝试获取现有学生，如果学号重复，则更新
                            student_id = student_data.get('student_id')
                            student_obj, created = Student.objects.update_or_create(
                                student_id=student_id,
                                defaults=student_data
                            )
                            
                            if created:
                                success_messages.append(f"成功新增学生：{student_obj.name} ({student_obj.student_id})")
                            else:
                                success_messages.append(f"成功更新学生：{student_obj.name} ({student_obj.student_id})")
                            imported_count += 1
                            
                    except Exception as e:
                        failed_rows.append((row_idx, row_data, str(e)))
                        error_messages.append(f"第 {row_idx} 行学生导入失败: {e}")
                        continue  # 继续处理下一行

                # 返回JSON响应用于弹窗显示
                return JsonResponse({
                    'success': True,
                    'imported_count': imported_count,
                    'failed_count': len(failed_rows),
                    'success_messages': success_messages,
                    'error_messages': error_messages,
                    'warning_messages': warning_messages,
                    'failed_rows': failed_rows
                })

            except Exception as e:
                return JsonResponse({
                    'success': False,
                    'error': f"文件处理失败或格式不正确: {e}"
                })
        else:
            return JsonResponse({
                'success': False,
                'error': "请选择正确的 Excel 文件。"
            })
            
    else:
        form = ExcelUploadForm()
        return render(request, 'students/student_batch_import.html', {
            'form': form,
            'title': '批量导入学生',
            'download_template_url': reverse('download_student_import_template')
        })

# 批量刪除學生
@require_POST
def student_batch_delete(request):
    # 從 POST 請求中獲取選中學生的 ID 列表
    selected_student_ids = request.POST.getlist('selected_students')
    
    if not selected_student_ids:
        messages.warning(request, "没有选择任何学生进行删除。")
        return redirect('student_list')
    
    try:
        with transaction.atomic(): # 確保刪除操作的原子性
            # 獲取所有選中的學生對象
            students_to_delete = Student.objects.filter(pk__in=selected_student_ids)
            deleted_count, _ = students_to_delete.delete() # 执行删除
            messages.success(request, f"成功删除 {deleted_count} 名学生。")
    except Exception as e:
        messages.error(request, f"批量删除学生失败：{e}")
    
    return redirect('student_list')

# 批量修改學生狀態
@require_POST
def student_batch_update_status(request):
    selected_student_ids = request.POST.getlist('selected_students')
    form = BatchUpdateStatusForm(request.POST)

    if not selected_student_ids:
        messages.warning(request, "没有选择任何学生进行状态修改。")
        return redirect('student_list')
        
    if form.is_valid():
        new_status = form.cleaned_data['status']
        try:
            with transaction.atomic(): # 確保更新操作的原子性
                # 篩選出需要更新的學生
                students_to_update = Student.objects.filter(pk__in=selected_student_ids)
                
                # 如果新狀態是「畢業」，自動設定畢業日期
                if new_status == '畢業':
                    # 只更新狀態不是「畢業」的學生，並設定畢業日期
                    updated_count = students_to_update.exclude(status='畢業').update(status=new_status, graduation_date=timezone.now().date())
                    # 對於已經是畢業狀態的，只更新狀態（雖然這裡實際不會再更新）
                    students_to_update.filter(status='畢業').update(status=new_status)
                else:
                    updated_count = students_to_update.update(status=new_status)

                messages.success(request, f"成功更新 {updated_count} 名学生的狀態为 '{new_status}'。")
        except Exception as e:
            messages.error(request, f"批量更新學生狀態失敗：{e}")
    else:
        # 如果表單無效，通常是前端驗證被繞過，或有其他問題
        messages.error(request, "提交的狀態無效，請重試。")
    
    return redirect('student_list')

# 批量升年級 PRD 3.1.6
@require_http_methods(["GET", "POST"])
def student_batch_promote_grade(request):
    if request.method == 'POST':
        form = BatchPromoteGradeForm(request.POST)
        if form.is_valid():
            current_grade_level = form.cleaned_data['current_grade_level']
            target_grade_level = form.cleaned_data['target_grade_level']
            auto_create_classes = form.cleaned_data['auto_create_classes']
            
            # 獲取所有選中的學生 ID
            selected_student_ids = request.POST.getlist('selected_students')
            
            if not selected_student_ids:
                messages.warning(request, "没有选择任何学生进行升年级操作。")
                return redirect('student_list')
            
            # 邏輯檢查：目標年級不能是空，也不能與當前年級相同（如果當前年級被選中）
            if not target_grade_level:
                messages.error(request, "请选择目标年级。")
                return render(request, 'students/batch_promote_grade_form.html', {'form': form, 'title': '批量升年级'})

            if current_grade_level and current_grade_level == target_grade_level:
                messages.warning(request, "目标年级不能与当前年级相同。")
                return render(request, 'students/batch_promote_grade_form.html', {'form': form, 'title': '批量升年级'})
            
            updated_count = 0
            errors = []
            
            with transaction.atomic():
                students_to_promote = Student.objects.filter(pk__in=selected_student_ids)

                # 如果設定了篩選的「當前年級」，則只處理符合該年級的學生
                if current_grade_level:
                    students_to_promote = students_to_promote.filter(current_class__grade_level=current_grade_level)

                for student in students_to_promote:
                    try:
                        current_class = student.current_class
                        if current_class:
                            # 找出目標年級中，與當前班級名稱相同的班級
                            target_class_obj = Class.objects.filter(
                                class_name=current_class.class_name,
                                grade_level=target_grade_level
                            ).first()

                            if not target_class_obj and auto_create_classes:
                                # 如果目標班級不存在且允許自動創建，則創建新班級
                                target_class_obj = Class.objects.create(
                                    class_name=current_class.class_name,
                                    grade_level=target_grade_level
                                )
                                messages.info(request, f"已自动创建班级：{target_grade_level}{current_class.class_name}。")
                            elif not target_class_obj and not auto_create_classes:
                                errors.append(f"学生 {student.name} (学号: {student.student_id}) 所在的班级 '{current_class.class_name}' 在目标年级 '{target_grade_level}' 中不存在，且未启用自动创建。跳过升级。")
                                continue # 跳過此學生

                            if target_class_obj:
                                student.current_class = target_class_obj
                                student.save()
                                updated_count += 1
                        else:
                            errors.append(f"学生 {student.name} (学号: {student.student_id}) 当前没有班级，跳过升级。")
                    except Exception as e:
                        errors.append(f"学生 {student.name} (学号: {student.student_id}) 升年级失败：{e}")
            
            if updated_count > 0:
                messages.success(request, f"成功将 {updated_count} 名学生升入 {target_grade_level}。")
            
            if errors:
                for err in errors:
                    messages.warning(request, err)
                messages.error(request, "部分学生升年级失败，请检查警告信息。")
            
            return redirect('student_list') # 升級後導回學生列表頁面
        else:
            # 表單驗證失敗
            messages.error(request, "表单提交有误，请检查。")
            return render(request, 'students/batch_promote_grade_form.html', {'form': form, 'title': '批量升年级'})
    else:
        form = BatchPromoteGradeForm()
    return render(request, 'students/batch_promote_grade_form.html', {'form': form, 'title': '批量升年级'})

# 批量畢業 (PRD 3.1.7）
@require_POST # 確保這個 View 只接受 POST 請求
def student_batch_graduate(request):
    selected_student_ids = request.POST.getlist('selected_students') # 獲取選中的學生 ID 列表

    if not selected_student_ids:
        messages.warning(request, "没有选择任何学生进行批量毕业操作。")
        return redirect('student_list')
    
    updated_count = 0
    
    try:
        with transaction.atomic(): # 使用事務，確保操作的原子性
            # 獲取所有選中的學生對象
            students_to_graduate = Student.objects.filter(pk__in=selected_student_ids)
            
            # 將這些學生的狀態更新為 '畢業'，並設定畢業日期為當前日期
            # 我們只更新那些狀態不是 '畢業' 的學生，以避免不必要的資料庫寫入
            updated_count = students_to_graduate.exclude(status='畢業').update(
                status='畢業',
                graduation_date=timezone.now().date() # 設定為今天的日期
            )
            
            # 如果有學生已經是畢業狀態，但仍然被選中，這不會再次更新
            # 但我們可以向用戶確認他們確實被包含在操作中
            already_graduated_count = students_to_graduate.filter(status='畢業').count()

            messages.success(request, f"成功將 {updated_count} 名學生設置為『畢業』狀態。")
            if already_graduated_count > 0:
                messages.info(request, f"其中有 {already_graduated_count} 名學生已是『畢業』狀態，無需重複操作。")

    except Exception as e:
        messages.error(request, f"批量畢業操作失敗：{e}")
    
    return redirect('student_list') # 操作完成後導回學生列表頁面

# 下載學生導入模板函數
def download_student_import_template(request):
    # 創建一個新的工作簿
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "学生导入模板"

    # 定義 Excel 表頭 (對應 Student 模型的關鍵字段)
    # 這裡的順序和名稱可以根據實際導入邏輯調整
    headers = [
        "学号 (必填)", "姓名 (必填)", "性别 (男/女)", "出生日期 (YYYY-MM-DD)",
        "年级 (初一/初二/初三/高一/高二/高三)", "班级名称 (1班-20班)", "在校状态 (在读/转学/休学/复学/毕业)",
        "身份证号码", "学籍号", "家庭地址", "监护人姓名", "监护人联系电话",
        "入学日期 (YYYY-MM-DD)", "毕业日期 (YYYY-MM-DD, 毕业状态必填)"
    ]
    sheet.append(headers)

    # 針對有選項的欄位，添加提示信息或數據驗證 (更友好的提示)
    # 性別提示
    gender_validation_text = "请填写 '男'或'女'"
    sheet.cell(row=2, column=headers.index("性别 (男/女)") + 1).comment = openpyxl.comments.Comment(gender_validation_text, "System")

    # 年級提示
    grade_level_options = ', '.join([choice[0] for choice in GRADE_LEVEL_CHOICES])
    grade_validation_text = f"请填写以下任一年级: {grade_level_options}"
    sheet.cell(row=2, column=headers.index("年级 (初一/初二/初三/高一/高二/高三)") + 1).comment = openpyxl.comments.Comment(grade_validation_text, "System")

    # 班級名稱提示
    class_name_options = ', '.join([choice[0] for choice in CLASS_NAME_CHOICES])
    class_validation_text = f"请填写以下任一班级: {class_name_options}"
    sheet.cell(row=2, column=headers.index("班级名称 (1班-20班)") + 1).comment = openpyxl.comments.Comment(class_validation_text, "System")

    # 在校狀態提示
    status_options = ', '.join([choice[0] for choice in STATUS_CHOICES])
    status_validation_text = f"请填写以下任一状态: {status_options}"
    sheet.cell(row=2, column=headers.index("在校状态 (在读/转学/休学/复学/毕业)") + 1).comment = openpyxl.comments.Comment(status_validation_text, "System")

    # 日期格式提示
    date_format_text = "日期格式必须是 YYYY-MM-DD，例如 2006-01-23"
    sheet.cell(row=2, column=headers.index("出生日期 (YYYY-MM-DD)") + 1).comment = openpyxl.comments.Comment(date_format_text, "System")
    sheet.cell(row=2, column=headers.index("入学日期 (YYYY-MM-DD)") + 1).comment = openpyxl.comments.Comment(date_format_text, "System")
    sheet.cell(row=2, column=headers.index("毕业日期 (YYYY-MM-DD, 毕业状态必填)") + 1).comment = openpyxl.comments.Comment(date_format_text + " (毕业状态必填此项)", "System")

    # 添加示例数据行
    example_data = [
        "2024001",  # 学号 (必填) - 字符串格式
        "张三",     # 姓名 (必填)
        "男",       # 性别
        "2006-05-15",  # 出生日期
        "高一",     # 年级
        "1班",      # 班级名称
        "在读",     # 在校状态
        "110101200605151234",  # 身份证号码
        "G20240001",  # 学籍号
        "北京市朝阳区某某街道",  # 家庭地址
        "张父",     # 监护人姓名
        "13800138000",  # 监护人联系电话
        "2024-09-01",  # 入学日期
        ""          # 毕业日期 (非毕业状态可为空)
    ]
    sheet.append(example_data)
    
    # 为示例数据行添加注释说明这是示例
    sheet.cell(row=3, column=1).comment = openpyxl.comments.Comment("这是示例数据，请删除此行后填入真实数据", "System")

    # 將工作簿內容寫入內存緩衝區
    excel_file = io.BytesIO()
    workbook.save(excel_file)
    excel_file.seek(0) # 將文件指針移到開頭

    # 設置 HTTP 響應頭，觸發下載
    response = HttpResponse(
        excel_file.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="student_import_template.xlsx"'
    return response



















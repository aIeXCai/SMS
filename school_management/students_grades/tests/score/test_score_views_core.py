"""核心视图层测试：score 模块的关键视图（列表、添加、编辑、导入队列提交、学生搜索）。

这些测试侧重于视图行为、表单处理与 django-rq 的交互（enqueue），使用 Django 的 TestCase 提供测试数据库。
"""
from io import BytesIO
from unittest.mock import patch, Mock

import openpyxl
from datetime import date

from django.test import TestCase, Client
from django.urls import reverse
from django.core.files.uploadedfile import SimpleUploadedFile

from school_management.students_grades.models import Student, Exam, Score, ExamSubject, Class


def make_excel_bytes(headers, rows):
    """Create an in-memory xlsx and return a SimpleUploadedFile suitable for client.post."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(headers)
    for row in rows:
        ws.append(list(row))
    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    return SimpleUploadedFile('scores.xlsx', bio.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


class ScoreViewsCoreTests(TestCase):
    """包含 score_list, score_add (POST), score_edit (POST), 导入时 enqueue 行为和 search_students_ajax 的测试。"""

    def setUp(self):
        self.client = Client()
        # 创建班级和学生
        self.class1 = Class.objects.create(grade_level='Grade8', class_name='1班')
        self.stu1 = Student.objects.create(student_id='S001', name='张三', grade_level='Grade8', current_class=self.class1)
        self.stu2 = Student.objects.create(student_id='S002', name='李四', grade_level='Grade8', current_class=self.class1)

        # 创建考试与科目
        self.exam = Exam.objects.create(name='期中考试', academic_year='2024-2025', grade_level='Grade8', date=date(2025, 6, 1))
        ExamSubject.objects.create(exam=self.exam, subject_code='语文', subject_name='语文', max_score=150)
        ExamSubject.objects.create(exam=self.exam, subject_code='数学', subject_name='数学', max_score=150)

        self.url_import = reverse('students_grades:score_batch_import_ajax')
        self.url_list = reverse('students_grades:score_list')
        self.url_search = reverse('students_grades:search_students_ajax')
        self.url_add = reverse('students_grades:score_add')

    def test_score_list_filters_and_pagination(self):
        """测试 score_list 的基本显示与按学号过滤行为以及 per_page 参数的处理。"""
        # 先创建一些成绩用于列表
        Score.objects.create(student=self.stu1, exam=self.exam, subject='语文', score_value=90)
        Score.objects.create(student=self.stu2, exam=self.exam, subject='语文', score_value=80)

        # 不带过滤
        resp = self.client.get(self.url_list)
        self.assertEqual(resp.status_code, 200)
        self.assertIn('aggregated_scores', resp.context)

        # 按学号过滤
        resp = self.client.get(self.url_list, {'student_id_filter': 'S001'})
        self.assertEqual(resp.status_code, 200)
        page = resp.context['aggregated_scores']
        # 应该只包含 S001 的记录
        rows = list(page.object_list)
        self.assertTrue(any(getattr(r.student_obj, 'student_id', '') == 'S001' for r in rows))

    def test_score_add_post_creates_scores_and_enqueues(self):
        """测试通过 ScoreAddForm 提交多科成绩后创建 Score 并尝试 enqueue 排名任务。

        这里我们直接调用 POST 并 patch django_rq.get_queue 模拟队列。
        """
        # 准备 POST 数据：student_id 要使用 hidden 字段 student_id（测试中的 student 对象 id）
        post_data = {
            'student': self.stu1.name,
            'student_id': str(self.stu1.id),
            'exam': str(self.exam.pk),
            'score_语文': '95.5',
            'score_数学': '88',
        }

        with patch('school_management.students_grades.tasks.update_all_rankings_async') as mock_task:
            mock_job = Mock(id='job-1')
            mock_task.delay.return_value = mock_job

            resp = self.client.post(self.url_add, post_data, follow=True)
            # POST 成功通常会 redirect
            self.assertEqual(resp.status_code, 200)
            # 确认成绩创建
            self.assertTrue(Score.objects.filter(student=self.stu1, exam=self.exam, subject='语文').exists())
            mock_task.delay.assert_called()

    def test_score_edit_post_updates_and_enqueues(self):
        """测试 score_edit 的 POST 更新分数并提交异步任务的分支。"""
        s = Score.objects.create(student=self.stu1, exam=self.exam, subject='语文', score_value=70)

        url_edit = reverse('students_grades:score_edit', args=[s.pk])
        post_data = {
            'student': str(self.stu1.pk),
            'exam': str(self.exam.pk),
            'subject': '语文',
            'score_value': '99.5'
        }

        with patch('school_management.students_grades.tasks.update_all_rankings_async') as mock_task:
            mock_job = Mock(id='job-2')
            mock_task.delay.return_value = mock_job

            resp = self.client.post(url_edit, post_data, follow=True)
            self.assertEqual(resp.status_code, 200)
            s.refresh_from_db()
            self.assertEqual(float(s.score_value), 99.5)
            mock_task.delay.assert_called()

    def test_import_enqueue_called_on_success(self):
        """上传成功后应调用 django-rq 的 enqueue 并在响应中表明任务已提交。"""
        headers = ['学号', '学生姓名', '语文']
        rows = [
            ('S001', '张三', 100),
        ]
        bio = make_excel_bytes(headers, rows)

        with patch('django_rq.get_queue') as mock_get_queue:
            mock_queue = Mock()
            mock_queue.connection.ping.return_value = True
            mock_queue.enqueue.return_value = Mock(id='job-3')
            mock_get_queue.return_value = mock_queue

            resp = self.client.post(self.url_import, {'exam': self.exam.pk, 'excel_file': bio}, format='multipart')
            data = resp.json()
            self.assertTrue(data['success'])
            self.assertEqual(data.get('ranking_update_status'), 'async_submitted')
            mock_queue.enqueue.assert_called()

    def test_search_students_ajax_returns_results(self):
        """测试 search_students_ajax 能够根据 query 返回学生列表（包含学号和姓名）。"""
        resp = self.client.get(self.url_search, {'q': '张三'})
        self.assertEqual(resp.status_code, 200)
        data = resp.json()
        self.assertIn('students', data)
        self.assertTrue(any(s['student_id'] == 'S001' for s in data['students']))


class ScoreBatchEditandOpsTests(TestCase):
    """来自独立测试文件的 score_batch_edit 覆盖：成功、验证失败回滚、队列不可用分支。"""
    """包含 score_batch_export, score_batch_delete_filtered, score_batch_export_selected, score_batch_delete_selected 的测试。"""
    def setUp(self):
        self.client = Client()
        self.class1 = Class.objects.create(grade_level='Grade8', class_name='1班')
        self.stu = Student.objects.create(student_id='S100', name='测试生', grade_level='Grade8', current_class=self.class1)
        self.exam = Exam.objects.create(name='期末', academic_year='2024-2025', grade_level='Grade8', date=date(2025,6,1))
        # subjects
        ExamSubject.objects.create(exam=self.exam, subject_code='语文', subject_name='语文', max_score=150)
        ExamSubject.objects.create(exam=self.exam, subject_code='数学', subject_name='数学', max_score=150)

        self.url = reverse('students_grades:score_batch_edit')
        # batch operations
        self.url_export = reverse('students_grades:score_batch_export')
        self.url_delete_filtered = reverse('students_grades:score_batch_delete_filtered')
        self.url_export_selected = reverse('students_grades:score_batch_export_selected')
        self.url_delete_selected = reverse('students_grades:score_batch_delete_selected')

    def test_batch_edit_success_creates_and_updates_and_enqueues(self):
        """POST 包含两个科目的分数，应创建记录并调用异步任务提交。"""
        data = {
            'student': str(self.stu.pk),
            'exam': str(self.exam.pk),
            'score_语文': '88',
            'score_数学': '92'
        }

        with patch('school_management.students_grades.tasks.update_all_rankings_async') as mock_task:
            mock_task.delay.return_value = Mock(id='job-x')

            resp = self.client.post(f"{self.url}?student={self.stu.pk}&exam={self.exam.pk}", data, follow=True)
            self.assertEqual(resp.status_code, 200)

            # 两门成绩应被创建
            self.assertTrue(Score.objects.filter(student=self.stu, exam=self.exam, subject='语文').exists())
            self.assertTrue(Score.objects.filter(student=self.stu, exam=self.exam, subject='数学').exists())

            mock_task.delay.assert_called()

    def test_batch_edit_validation_error_rolls_back(self):
        """如果某科分数格式不对或超范围，应回滚并在响应中显示错误信息（不创建任何记录）。"""
        # 给一个非法分数（非数字）
        data = {
            'student': str(self.stu.pk),
            'exam': str(self.exam.pk),
            'score_语文': 'not-a-number',
            'score_数学': '50'
        }

        resp = self.client.post(f"{self.url}?student={self.stu.pk}&exam={self.exam.pk}", data)
        # 响应应返回页面（200）并包含错误消息
        self.assertEqual(resp.status_code, 200)
        # 验证没有任何成绩写入数据库
        self.assertFalse(Score.objects.filter(student=self.stu, exam=self.exam).exists())

    def test_batch_edit_queue_unavailable_still_saves(self):
        """当 django-rq/Redis 不可用时，应仍保存成绩，但记录 warning 并不抛出异常。"""
        data = {
            'student': str(self.stu.pk),
            'exam': str(self.exam.pk),
            'score_语文': '77',
        }

        # 模拟 update_all_rankings_async.delay 抛出异常
        with patch('school_management.students_grades.tasks.update_all_rankings_async') as mock_task:
            mock_task.delay.side_effect = Exception('redis down')

            resp = self.client.post(f"{self.url}?student={self.stu.pk}&exam={self.exam.pk}", data, follow=True)
            self.assertEqual(resp.status_code, 200)

            # 成绩应已保存
            self.assertTrue(Score.objects.filter(student=self.stu, exam=self.exam, subject='语文').exists())

    def test_score_batch_export_filters_returns_excel(self):
        """按筛选导出应返回 xlsx 文件，且内容包含符合筛选条件的行。"""
        # 创建另一个学生和成绩用于验证过滤
        other_class = Class.objects.create(grade_level='Grade8', class_name='2班')
        other_stu = Student.objects.create(student_id='S101', name='另一个', grade_level='Grade8', current_class=other_class)

        # 为两个学生分别创建成绩
        Score.objects.create(student=self.stu, exam=self.exam, subject='语文', score_value=88)
        Score.objects.create(student=other_stu, exam=self.exam, subject='语文', score_value=77)

        # 仅导出学号包含 S100 的记录
        resp = self.client.get(self.url_export, {'student_id_filter': 'S100', 'exam_filter': str(self.exam.pk)})
        self.assertEqual(resp.status_code, 200)
        self.assertEqual(resp['Content-Type'], 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

        # 响应应包含 xlsx 文件名（文件名可能被编码），至少应包含 .xlsx 或 filename 字段
        cd = resp.get('Content-Disposition', '')
        self.assertTrue('.xlsx' in cd.lower() or 'filename' in cd.lower() or '=?utf' in cd.lower())

        # 解析返回的 Excel 内容，确保第一条数据行是 S100
        from io import BytesIO as _BytesIO
        wb = openpyxl.load_workbook(_BytesIO(resp.content))
        sheet = wb.active
        # header 在第一行，数据从第二行开始
        data_rows = list(sheet.iter_rows(min_row=2, values_only=True))
        exported_student_ids = {row[0] for row in data_rows}
        # 由于请求时使用了 student_id_filter=S100，因此导出应只包含 self.stu
        self.assertIn(self.stu.student_id, exported_student_ids)
        self.assertNotIn(other_stu.student_id, exported_student_ids)

    def test_score_batch_delete_filtered_deletes_and_enqueues(self):
        """按筛选删除应删除对应条目并为受影响的考试提交排名任务；当无匹配时不提交任务。"""
        # 创建两条成绩用于删除
        Score.objects.create(student=self.stu, exam=self.exam, subject='语文', score_value=60)
        Score.objects.create(student=self.stu, exam=self.exam, subject='数学', score_value=70)

        with patch('school_management.students_grades.tasks.update_all_rankings_async') as mock_task:
            mock_task.delay.return_value = Mock(id='job-del')

            resp = self.client.post(self.url_delete_filtered, {'student_id_filter': 'S100', 'exam_filter': str(self.exam.pk)}, follow=True)
            # 跳转成功
            self.assertEqual(resp.status_code, 200)

            # 所有该学生/考试的成绩应被删除
            self.assertFalse(Score.objects.filter(student=self.stu, exam=self.exam).exists())

            # update_all_rankings_async.delay 应被调用一次（针对受影响的 exam id）
            self.assertTrue(mock_task.delay.called)

        # 当没有匹配记录时，不应提交任务
        # 确保数据库为空
        Score.objects.all().delete()
        with patch('school_management.students_grades.tasks.update_all_rankings_async') as mock_task2:
            resp2 = self.client.post(self.url_delete_filtered, {'student_id_filter': 'NONEXIST', 'exam_filter': str(self.exam.pk)}, follow=True)
            self.assertEqual(resp2.status_code, 200)
            # 没有记录被删除，任务不应被提交
            self.assertFalse(mock_task2.delay.called)

    def test_score_batch_export_selected_and_delete_selected(self):
        """导出选中项应接受 selected_records 并返回 xlsx；删除选中项应删除并 enqueue 排名任务。"""
        # 准备成绩
        Score.objects.create(student=self.stu, exam=self.exam, subject='语文', score_value=99)
        Score.objects.create(student=self.stu, exam=self.exam, subject='数学', score_value=95)

        # 前端使用 student.pk 作为选中记录的 student 部分
        selected_value = f"{self.stu.pk}_{self.exam.pk}"

        # 导出选中项
        resp = self.client.post(self.url_export_selected, {'selected_records': [selected_value]}, follow=True)
        # 导出返回 xlsx 文件
        self.assertEqual(resp.status_code, 200)
        self.assertEqual(resp['Content-Type'], 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

        # 解析 Excel，确认学号存在
        from io import BytesIO as _BytesIO
        wb = openpyxl.load_workbook(_BytesIO(resp.content))
        sheet = wb.active
        first_data_row = list(list(sheet.iter_rows(min_row=2, max_row=2, values_only=True))[0])
        self.assertEqual(first_data_row[0], self.stu.student_id)

        # 删除选中项并确保 enqueue 被调用
        with patch('school_management.students_grades.tasks.update_all_rankings_async') as mock_task3:
            mock_task3.delay.return_value = Mock(id='job-sel')
            resp_del = self.client.post(self.url_delete_selected, {'selected_records': [selected_value]}, follow=True)
            self.assertEqual(resp_del.status_code, 200)
            # 成绩应被删除
            self.assertFalse(Score.objects.filter(student=self.stu, exam=self.exam).exists())
            self.assertTrue(mock_task3.delay.called)


class ScoreQueryTests(TestCase):
    """测试成绩查询页面、结果与导出：参数验证、排序与导出文件内容。

    本测试类创建两个学生与一次考试，保证有可比的总分用于排序断言；同时验证导出的 Excel 包含关键列："总分" 与 "年级排名"。
    """

    def setUp(self):
        # 初始化客户端与测试数据：两个学生、一个考试与两个科目
        self.client = Client()
        self.class1 = Class.objects.create(grade_level='Grade8', class_name='1班')
        self.stu1 = Student.objects.create(student_id='Q001', name='查询甲', grade_level='Grade8', current_class=self.class1)
        self.stu2 = Student.objects.create(student_id='Q002', name='查询乙', grade_level='Grade8', current_class=self.class1)

        # 创建考试与科目
        self.exam = Exam.objects.create(name='查询考试', academic_year='2024-2025', grade_level='Grade8', date=date(2025, 5, 20))
        ExamSubject.objects.create(exam=self.exam, subject_code='语文', subject_name='语文', max_score=150)
        ExamSubject.objects.create(exam=self.exam, subject_code='数学', subject_name='数学', max_score=150)

        # 准备成绩数据：stu1 的总分高于 stu2，便于验证排序
        Score.objects.create(student=self.stu1, exam=self.exam, subject='语文', score_value=95)
        Score.objects.create(student=self.stu1, exam=self.exam, subject='数学', score_value=90)

        Score.objects.create(student=self.stu2, exam=self.exam, subject='语文', score_value=88)
        Score.objects.create(student=self.stu2, exam=self.exam, subject='数学', score_value=79)

        # 目标 URL
        self.url_query = reverse('students_grades:score_query')
        self.url_results = reverse('students_grades:score_query_results')
        self.url_export = reverse('students_grades:score_query_export')

    def test_score_query_page_renders(self):
        """score_query 页面应该能正常加载并包含查询表单（sanity check）。

        重点：返回状态码 200，并在上下文中包含名为 'form' 的表单对象。
        """
        resp = self.client.get(self.url_query)
        self.assertEqual(resp.status_code, 200)
        # 页面应包含查询表单，供用户输入筛选条件
        self.assertIn('form', resp.context)

    def test_score_query_results_invalid_params(self):
        """当传入无效查询参数（例如不存在的 exam id）时，视图应返回页面但不展示结果。

        断言：HTTP 200，且上下文 'has_results' 为 False（表示表单校验未通过或未返回结果）。
        """
        resp = self.client.get(self.url_results, {'exam': '999999'})
        self.assertEqual(resp.status_code, 200)
        # 无效表单时视图会将 has_results 标记为 False
        self.assertIn('has_results', resp.context)
        self.assertFalse(resp.context['has_results'])

    def test_score_query_results_sort_by_total_score(self):
        """验证按总分排序（subject_sort=total_score，降序）时，分数高的学生排在前面。

        该测试通过检查 results 上下文中的第一页顺序来确认排序生效。
        """
        resp = self.client.get(self.url_results, {'exam': str(self.exam.pk), 'subject_sort': 'total_score', 'sort_order': 'desc'})
        self.assertEqual(resp.status_code, 200)
        self.assertIn('results', resp.context)
        page = resp.context['results']
        items = list(page.object_list)
        # 至少包含两个结果用于比较
        self.assertGreaterEqual(len(items), 2)

        # items 中的元素是字典，包含 'student' 与 'total_score' 字段
        first = items[0]
        second = items[1]

        # 断言第一个记录的总分不小于第二个（降序），并且第一个是 stu1
        self.assertTrue(first['total_score'] >= second['total_score'])
        self.assertEqual(first['student'].student_id, self.stu1.student_id)

    def test_score_query_export_returns_excel_with_totals_and_ranks(self):
        """调用导出接口应返回一个包含查询结果的 xlsx 文件。

        验证点：
        - 返回 Content-Type 为 xlsx
        - Content-Disposition 中包含文件名或 .xlsx
        - 表头最后两列为 '总分' 和 '年级排名'
        - 至少有一行数据包含学号
        """
        resp = self.client.get(self.url_export, {'exam': str(self.exam.pk)})
        self.assertEqual(resp.status_code, 200)
        self.assertEqual(resp['Content-Type'], 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

        # Content-Disposition 可能会因为编码被变化，允许包含多种形式
        cd = resp.get('Content-Disposition', '')
        self.assertTrue('.xlsx' in cd.lower() or 'filename' in cd.lower() or '=?utf' in cd.lower())

        # 解析返回的 Excel，检查表头与数据行
        from io import BytesIO as _BytesIO
        wb = openpyxl.load_workbook(_BytesIO(resp.content))
        sheet = wb.active
        header = list(sheet.iter_rows(min_row=1, max_row=1, values_only=True))[0]

        # 最后两列应为 总分 与 年级排名（保证导出包含这两列）
        self.assertGreaterEqual(len(header), 2)
        self.assertEqual(header[-2], '总分')
        self.assertEqual(header[-1], '年级排名')

        # 检查导出包含两位学生的学号（不依赖行顺序）
        data_rows = list(sheet.iter_rows(min_row=2, values_only=True))
        exported_student_ids = {row[0] for row in data_rows}
        self.assertTrue({self.stu1.student_id, self.stu2.student_id}.issubset(exported_student_ids))

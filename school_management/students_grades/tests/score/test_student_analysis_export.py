from datetime import date
from io import BytesIO

import openpyxl

from .test_base import BaseTestCase
from school_management.students_grades.models import Class, Exam, ExamSubject, Score, Student


class StudentAnalysisExportApiTests(BaseTestCase):
    def setUp(self):
        super().setUp()
        self.cls = Class.objects.create(grade_level="Grade8", class_name="3班")
        self.student = Student.objects.create(
            student_id="EXP001",
            name="导出生",
            cohort="2024级",
            grade_level="Grade8",
            current_class=self.cls,
        )

        self.exam1 = Exam.objects.create(name="期中", academic_year="2025-2026", grade_level="Grade8", date=date(2025, 11, 20))
        self.exam2 = Exam.objects.create(name="期末", academic_year="2025-2026", grade_level="Grade8", date=date(2026, 1, 20))

        for exam in [self.exam1, self.exam2]:
            ExamSubject.objects.create(exam=exam, subject_code="语文", subject_name="语文", max_score=150)
            ExamSubject.objects.create(exam=exam, subject_code="数学", subject_name="数学", max_score=150)

        Score.objects.create(
            student=self.student,
            exam=self.exam1,
            subject="语文",
            score_value=95,
            grade_rank_in_subject=8,
            class_rank_in_subject=3,
            total_score_rank_in_grade=10,
            total_score_rank_in_class=4,
        )
        Score.objects.create(
            student=self.student,
            exam=self.exam1,
            subject="数学",
            score_value=98,
            grade_rank_in_subject=6,
            class_rank_in_subject=2,
            total_score_rank_in_grade=10,
            total_score_rank_in_class=4,
        )

        Score.objects.create(
            student=self.student,
            exam=self.exam2,
            subject="语文",
            score_value=108,
            grade_rank_in_subject=6,
            class_rank_in_subject=2,
            total_score_rank_in_grade=7,
            total_score_rank_in_class=3,
        )
        Score.objects.create(
            student=self.student,
            exam=self.exam2,
            subject="数学",
            score_value=112,
            grade_rank_in_subject=4,
            class_rank_in_subject=1,
            total_score_rank_in_grade=7,
            total_score_rank_in_class=3,
        )

    def _export(self, student_id):
        return self.client.get(
            "/api/scores/student-analysis-report-export",
            {"student_id": str(student_id)},
        )

    def test_export_success_returns_excel_with_four_sheets(self):
        resp = self._export(self.student.id)
        self.assertEqual(resp.status_code, 200)
        self.assertEqual(
            resp["Content-Type"],
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        self.assertTrue(resp.get("Content-Disposition"))

        wb = openpyxl.load_workbook(BytesIO(resp.content))
        self.assertEqual(wb.sheetnames, ["总览", "总分趋势", "科目明细", "科目趋势"])
        self.assertGreaterEqual(len(wb["总分趋势"]._charts), 2)
        self.assertGreaterEqual(len(wb["科目趋势"]._charts), 1)

    def test_export_missing_student_id_returns_400(self):
        resp = self.client.get("/api/scores/student-analysis-report-export")
        self.assertEqual(resp.status_code, 400)
        body = resp.json()
        self.assertIn("error", body)

    def test_export_student_not_found_returns_404(self):
        resp = self._export(999999)
        self.assertEqual(resp.status_code, 404)
        body = resp.json()
        self.assertEqual(body.get("error"), "学生不存在")

    def test_export_no_exam_data_returns_400(self):
        empty_student = Student.objects.create(
            student_id="EXP-NO-DATA",
            name="无数据学生",
            cohort="2024级",
            grade_level="Grade8",
            current_class=self.cls,
        )
        resp = self._export(empty_student.id)
        self.assertEqual(resp.status_code, 400)
        body = resp.json()
        self.assertEqual(body.get("error"), "该学生暂无可导出分析数据")

    def test_rank_change_formula_and_subject_trend_header(self):
        resp = self._export(self.student.id)
        self.assertEqual(resp.status_code, 200)
        wb = openpyxl.load_workbook(BytesIO(resp.content), data_only=True)

        total_trend = wb["总分趋势"]
        # 第2条考试记录（第3行）：上次10 - 本次7 = 3
        self.assertEqual(total_trend["G3"].value, 3)
        self.assertEqual(total_trend["I3"].value, 3)

        subject_trend = wb["科目趋势"]
        headers = [cell.value for cell in subject_trend[1]]
        self.assertIn("语文级排变化", headers)
        # 第2条考试记录：语文上次8 - 本次6 = 2
        self.assertEqual(subject_trend["F3"].value, 2)

    def test_chart_downgrade_notice_when_only_one_exam(self):
        cls = Class.objects.create(grade_level="Grade8", class_name="8班")
        student = Student.objects.create(
            student_id="EXP-ONE",
            name="单场学生",
            cohort="2024级",
            grade_level="Grade8",
            current_class=cls,
        )
        exam = Exam.objects.create(name="单场考试", academic_year="2025-2026", grade_level="Grade8", date=date(2025, 10, 1))
        ExamSubject.objects.create(exam=exam, subject_code="语文", subject_name="语文", max_score=150)
        Score.objects.create(
            student=student,
            exam=exam,
            subject="语文",
            score_value=103,
            grade_rank_in_subject=9,
            class_rank_in_subject=3,
            total_score_rank_in_grade=9,
            total_score_rank_in_class=3,
        )

        resp = self._export(student.id)
        self.assertEqual(resp.status_code, 200)
        wb = openpyxl.load_workbook(BytesIO(resp.content), data_only=True)
        total_trend = wb["总分趋势"]

        self.assertEqual(len(total_trend._charts), 0)
        self.assertEqual(total_trend["K2"].value, "数据不足，未生成趋势图")

    def test_subject_trend_chart_helper_does_not_overwrite_politics_rank(self):
        for exam in [self.exam1, self.exam2]:
            ExamSubject.objects.create(exam=exam, subject_code="英语", subject_name="英语", max_score=150)
            ExamSubject.objects.create(exam=exam, subject_code="政治", subject_name="政治", max_score=100)

        Score.objects.create(
            student=self.student,
            exam=self.exam1,
            subject="英语",
            score_value=92,
            grade_rank_in_subject=9,
            class_rank_in_subject=4,
            total_score_rank_in_grade=10,
            total_score_rank_in_class=4,
        )
        Score.objects.create(
            student=self.student,
            exam=self.exam1,
            subject="政治",
            score_value=86,
            grade_rank_in_subject=15,
            class_rank_in_subject=5,
            total_score_rank_in_grade=10,
            total_score_rank_in_class=4,
        )
        Score.objects.create(
            student=self.student,
            exam=self.exam2,
            subject="英语",
            score_value=102,
            grade_rank_in_subject=7,
            class_rank_in_subject=3,
            total_score_rank_in_grade=7,
            total_score_rank_in_class=3,
        )
        Score.objects.create(
            student=self.student,
            exam=self.exam2,
            subject="政治",
            score_value=91,
            grade_rank_in_subject=12,
            class_rank_in_subject=4,
            total_score_rank_in_grade=7,
            total_score_rank_in_class=3,
        )

        resp = self._export(self.student.id)
        self.assertEqual(resp.status_code, 200)
        wb = openpyxl.load_workbook(BytesIO(resp.content), data_only=True)
        subject_trend = wb["科目趋势"]

        politics_rank_col = None
        for col_index in range(1, subject_trend.max_column + 1):
            if subject_trend.cell(row=1, column=col_index).value == "政治级排":
                politics_rank_col = col_index
                break

        self.assertIsNotNone(politics_rank_col)
        self.assertNotEqual(subject_trend.cell(row=2, column=politics_rank_col).value, "科目")
        self.assertEqual(subject_trend.cell(row=2, column=politics_rank_col).value, 15)


class StudentAnalysisExportConsistencyTests(BaseTestCase):
    def setUp(self):
        super().setUp()
        self.cls = Class.objects.create(grade_level="Grade8", class_name="2班")
        self.exam = Exam.objects.create(name="一致性考试", academic_year="2025-2026", grade_level="Grade8", date=date(2025, 12, 1))
        ExamSubject.objects.create(exam=self.exam, subject_code="语文", subject_name="语文", max_score=150)

        self.students = []
        scores = [101, 105, 109]
        ranks = [12, 9, 6]
        for index in range(3):
            student = Student.objects.create(
                student_id=f"CONS{index + 1}",
                name=f"样本{index + 1}",
                cohort="2024级",
                grade_level="Grade8",
                current_class=self.cls,
            )
            Score.objects.create(
                student=student,
                exam=self.exam,
                subject="语文",
                score_value=scores[index],
                grade_rank_in_subject=ranks[index],
                class_rank_in_subject=index + 1,
                total_score_rank_in_grade=ranks[index],
                total_score_rank_in_class=index + 1,
            )
            self.students.append(student)

    def test_three_students_export_matches_analysis_api_total_score(self):
        for student in self.students:
            with self.subTest(student=student.student_id):
                analysis_resp = self.client.get(
                    "/api/scores/student-analysis-data",
                    {"student_id": str(student.id)},
                )
                self.assertEqual(analysis_resp.status_code, 200)
                analysis_total = analysis_resp.json()["data"]["exams"][-1]["total_score"]

                export_resp = self.client.get(
                    "/api/scores/student-analysis-report-export",
                    {"student_id": str(student.id)},
                )
                self.assertEqual(export_resp.status_code, 200)

                wb = openpyxl.load_workbook(BytesIO(export_resp.content), data_only=True)
                overview_sheet = wb["总览"]

                found_latest_total = None
                for row in overview_sheet.iter_rows(min_row=1, max_row=overview_sheet.max_row, min_col=1, max_col=2):
                    if row[0].value == "最新考试总分":
                        found_latest_total = row[1].value
                        break

                self.assertIsNotNone(found_latest_total)
                self.assertEqual(float(found_latest_total), float(analysis_total))

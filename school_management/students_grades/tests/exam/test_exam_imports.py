"""
考试导入功能测试

- 覆盖考试批量导入、科目配置导入、异常数据处理等场景。
- 使用Django的TestCase，测试导入逻辑的正确性和健壮性。
"""

from django.test import TestCase
from django.core.exceptions import ValidationError
from school_management.students_grades.models.exam import Exam, ExamSubject
import io
import openpyxl

class ExamImportTests(TestCase):
    """考试导入相关测试"""

    def setUp(self):
        # 构造一个简单的Excel文件用于导入测试
        self.wb = openpyxl.Workbook()
        ws = self.wb.active
        ws.title = "考试导入"
        ws.append(["考试名称", "学年", "年级", "日期", "科目", "满分"])
        ws.append(["期中考试", "2025-2026", "初一", "2025-09-01", "语文", 120])
        ws.append(["期中考试", "2025-2026", "初一", "2025-09-01", "数学", 120])
        ws.append(["期末考试", "2025-2026", "初二", "2025-12-01", "英语", 120])
        self.excel_bytes = io.BytesIO()
        self.wb.save(self.excel_bytes)
        self.excel_bytes.seek(0)

    def test_exam_import_basic(self):
        """测试基本的考试和科目导入流程（模拟Excel导入）"""
        ws = self.wb.active
        exams_created = {}
        for row in ws.iter_rows(min_row=2, values_only=True):
            name, academic_year, grade_level, date, subject_code, max_score = row
            key = (name, academic_year, grade_level, date)
            if not name or not academic_year or not grade_level or not date or not subject_code or not max_score:
                continue  # 跳过不完整数据
            if key not in exams_created:
                exam = Exam.objects.create(name=name, academic_year=academic_year, grade_level=grade_level, date=date)
                exams_created[key] = exam
            else:
                exam = exams_created[key]
            ExamSubject.objects.create(exam=exam, subject_code=subject_code, max_score=max_score)
        # 校验考试和科目数量
        self.assertEqual(Exam.objects.count(), 2)
        exam1 = Exam.objects.get(name="期中考试", academic_year="2025-2026", grade_level="初一")
        self.assertEqual(exam1.exam_subjects.count(), 2)
        exam2 = Exam.objects.get(name="期末考试", academic_year="2025-2026", grade_level="初二")
        self.assertEqual(exam2.exam_subjects.count(), 1)

    def test_exam_import_with_invalid_data(self):
        """测试导入异常数据的处理，如缺少必填字段、重复考试等"""
        ws = self.wb.active
        ws.append(["", "2025-2026", "初一", "2025-09-01", "语文", 120])  # 缺考试名称
        ws.append(["期中考试", "2025-2026", "初一", "2025-09-01", "语文", 120])  # 重复考试科目
        exams_created = {}
        errors = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            name, academic_year, grade_level, date, subject_code, max_score = row
            if not name or not academic_year or not grade_level or not date or not subject_code or not max_score:
                errors.append("缺少必填字段")
                continue
            key = (name, academic_year, grade_level, date)
            if key not in exams_created:
                exam = Exam.objects.create(name=name, academic_year=academic_year, grade_level=grade_level, date=date)
                exams_created[key] = exam
            else:
                exam = exams_created[key]
            # 检查重复科目
            if exam.exam_subjects.filter(subject_code=subject_code).exists():
                errors.append(f"考试 {name} 已存在科目 {subject_code}")
                continue
            ExamSubject.objects.create(exam=exam, subject_code=subject_code, max_score=max_score)
        # 校验异常捕获
        self.assertIn("缺少必填字段", errors)
        self.assertIn("考试 期中考试 已存在科目 语文", errors)

    def test_exam_import_subject_config(self):
        """测试科目满分配置导入是否正确"""
        ws = self.wb.active
        ws.append(["期末考试", "2025-2026", "初二", "2025-12-01", "物理", 100])
        exams_created = {}
        for row in ws.iter_rows(min_row=2, values_only=True):
            name, academic_year, grade_level, date, subject_code, max_score = row
            key = (name, academic_year, grade_level, date)
            if not name or not academic_year or not grade_level or not date or not subject_code or not max_score:
                continue
            if key not in exams_created:
                exam = Exam.objects.create(name=name, academic_year=academic_year, grade_level=grade_level, date=date)
                exams_created[key] = exam
            else:
                exam = exams_created[key]
            ExamSubject.objects.create(exam=exam, subject_code=subject_code, max_score=max_score)
        exam = Exam.objects.get(name="期末考试", academic_year="2025-2026", grade_level="初二")
        subj = exam.exam_subjects.get(subject_code="物理")
        self.assertEqual(subj.max_score, 100)

    # 可继续补充边界、异常、批量等场景测试

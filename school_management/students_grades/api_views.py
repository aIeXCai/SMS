import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import datetime
import json
from collections import defaultdict
from django.http import HttpResponse
from django.db import transaction, models
from django.utils import timezone
from django.core.paginator import Paginator
from rest_framework.parsers import MultiPartParser, FormParser, JSONParser
from rest_framework import viewsets, permissions, status
from rest_framework.decorators import action
from rest_framework.response import Response
from django_filters.rest_framework import DjangoFilterBackend
from rest_framework import filters
from .models.student import (
    Student, Class,
    STATUS_CHOICES, GRADE_LEVEL_CHOICES, CLASS_NAME_CHOICES, COHORT_CHOICES,
)
from .models.score import Score, SUBJECT_CHOICES as SCORE_SUBJECT_CHOICES
from .models.exam import Exam, ExamSubject, ACADEMIC_YEAR_CHOICES, SUBJECT_DEFAULT_MAX_SCORES
from .serializers import StudentSerializer, ClassSerializer, ExamSerializer, ScoreSerializer
from .tasks import update_all_rankings_async
from .services.analysis_service import analyze_single_class, analyze_multiple_classes, analyze_grade
from .services import execute_target_student_rule
from school_management.users.permissions import IsAdminOrStaff, IsAdminOrGradeManagerOrStaff


class StudentViewSet(viewsets.ModelViewSet):
    """
    学生管理 ViewSet
    提供学生的 CRUD 操作

    这个 ViewSet 同时也提供一些辅助 API 以方便前端在单页应用中获取数据。
    """
    queryset = Student.objects.all().select_related('current_class')
    serializer_class = StudentSerializer
    permission_classes = [permissions.IsAuthenticated]
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    # cohort 存储 cohort 格式（如"初中2023级"），grade_level 保留用于展示
    filterset_fields = ['status', 'gender', 'current_class__cohort', 'current_class__class_name']
    search_fields = ['name', 'student_id', 'current_class__class_name']
    ordering_fields = ['student_id', 'name', 'entry_date']
    ordering = ['student_id']

    def get_permissions(self):
        """
        根据操作类型设置不同的权限
        """
        if self.action in [
            'create', 'update', 'partial_update', 'destroy',
            'batch_delete', 'batch_update_status', 'batch_promote', 'batch_import'
        ]:
            permission_classes = [permissions.IsAuthenticated, IsAdminOrStaff]
        else:
            # 查看操作所有登录用户都可以
            permission_classes = [permissions.IsAuthenticated]
        
        return [permission() for permission in permission_classes]

    def perform_destroy(self, instance):
        # 获取该学生参与的所有考试ID，用于后续排名更新
        affected_exam_ids = list(Score.objects.filter(student=instance).values_list('exam_id', flat=True).distinct())
        instance.delete()
        
        # 异步更新受影响考试的排名
        if affected_exam_ids:
            for exam_id in affected_exam_ids:
                try:
                    update_all_rankings_async.delay(exam_id)
                except Exception:
                    pass  # 静默处理任务提交错误

    def perform_update(self, serializer):
        # 检查是否变更为毕业状态，如果是则补充毕业日期
        instance = serializer.save()
        if instance.status == '毕业' and not instance.graduation_date:
            instance.graduation_date = timezone.now().date()
            instance.save(update_fields=['graduation_date'])

    @action(detail=False, methods=['post'], url_path='batch-delete')
    def batch_delete(self, request):
        """批量删除学生"""
        student_ids = request.data.get('student_ids', [])
        if not student_ids:
            return Response({'success': False, 'message': '没有选择任何学生。'}, status=status.HTTP_400_BAD_REQUEST)

        try:
            with transaction.atomic():
                students_to_delete = Student.objects.filter(pk__in=student_ids)
                affected_exam_ids = list(Score.objects.filter(student__in=students_to_delete).values_list('exam_id', flat=True).distinct())
                
                deleted_count, _ = students_to_delete.delete()
                
                if affected_exam_ids:
                    for exam_id in affected_exam_ids:
                        try:
                            update_all_rankings_async.delay(exam_id)
                        except Exception:
                            pass
                            
                return Response({
                    'success': True,
                    'message': f'成功删除 {deleted_count} 名学生。'
                })
        except Exception as e:
            return Response({'success': False, 'message': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    @action(detail=False, methods=['post'], url_path='batch-update-status')
    def batch_update_status(self, request):
        """批量修改学生状态（支持批量毕业）"""
        student_ids = request.data.get('student_ids', [])
        new_status = request.data.get('status', '')
        
        if not student_ids or not new_status:
            return Response({'success': False, 'message': '缺少必要参数。'}, status=status.HTTP_400_BAD_REQUEST)
            
        try:
            with transaction.atomic():
                students_to_update = Student.objects.filter(pk__in=student_ids)
                
                if new_status == '毕业':
                    # 只更新状态不是毕业的，设置当前的 graduation_date
                    updated_count = students_to_update.exclude(status='毕业').update(
                        status=new_status, 
                        graduation_date=timezone.now().date()
                    )
                    # 已经是毕业的无需改毕业时间，但也保障新状态一致性，不过这句其实无异于不变。
                    students_to_update.filter(status='毕业').update(status=new_status)
                else:
                    updated_count = students_to_update.update(status=new_status)
                    
                return Response({
                    'success': True, 
                    'message': f'成功更新 {updated_count} 名学生的状态为 "{new_status}"。'
                })
        except Exception as e:
            return Response({'success': False, 'message': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    @action(detail=False, methods=['get'], url_path='stats')
    def stats(self, request):
        """返回学生统计数据，用于前端展示总数/状态分布等。"""
        # 与原 student_list 视图保持一致：统计所有学生（不受过滤影响）
        all_students = Student.objects.all()
        total_students = all_students.count()
        active_students = all_students.filter(status='在读').count()
        graduated_students = all_students.filter(status='毕业').count()
        suspended_students = all_students.filter(status='休学').count()

        return Response({
            'total_students': total_students,
            'active_students': active_students,
            'graduated_students': graduated_students,
            'suspended_students': suspended_students,
            'status_choices': [c[0] for c in STATUS_CHOICES],
            'grade_level_choices': [c[0] for c in GRADE_LEVEL_CHOICES],  # deprecated
            'cohort_choices': [c[0] for c in COHORT_CHOICES],
            'class_name_choices': [c[0] for c in CLASS_NAME_CHOICES],
        })


    @action(detail=False, methods=['post'], url_path='batch-promote')
    def batch_promote(self, request):
        """
        批量升年级
        期望 payload:
        {
            "student_ids": [1, 2, 3],
            "target_grade_level": "二年级",
            "current_grade_level": "(可选)一年级",
            "auto_create_classes": true
        }
        """
        from django.db import transaction
        student_ids = request.data.get('student_ids', [])
        target_grade = request.data.get('target_grade_level')
        from_grade = request.data.get('current_grade_level')
        auto_create = request.data.get('auto_create_classes', False)

        if not student_ids:
            return Response({"success": False, "message": "未选择任何学生"}, status=400)
        if not target_grade:
            return Response({"success": False, "message": "请选择目标年级"}, status=400)
        if from_grade and from_grade == target_grade:
            return Response({"success": False, "message": "目标年级不能与当前年级相同"}, status=400)

        updated_count = 0
        errors = []

        with transaction.atomic():
            students = Student.objects.filter(pk__in=student_ids)
            if from_grade:
                students = students.filter(current_class__grade_level=from_grade)

            for student in students:
                try:
                    current_class = student.current_class
                    if current_class:
                        target_class_obj = Class.objects.filter(
                            class_name=current_class.class_name,
                            grade_level=target_grade
                        ).first()

                        if not target_class_obj and auto_create:
                            target_class_obj = Class.objects.create(
                                class_name=current_class.class_name,
                                grade_level=target_grade
                            )

                        if target_class_obj:
                            student.current_class = target_class_obj
                            student.grade_level = target_grade
                            student.save()
                            updated_count += 1
                        else:
                            errors.append(f"学生 {student.name} 所在的班级在目标年级中不存在，且未开启自动创建。")
                    else:
                        errors.append(f"学生 {student.name} 当前无班级，跳过。")
                except Exception as e:
                    errors.append(f"学生 {student.name} 失败: {str(e)}")

        return Response({
            "success": updated_count > 0,
            "message": f"成功将 {updated_count} 名学生升入 {target_grade}",
            "updated_count": updated_count,
            "errors": errors
        })

    @action(detail=False, methods=['get'], url_path='download-template')
    def download_template(self, request):
        """下载批量导入模板（新版：学段+届别格式）"""
        from .models.student import SECTION_CHOICES, COHORT_YEAR_CHOICES

        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "学生导入模板"

        # 样式定义
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill("solid", fgColor="00897B")
        header_alignment = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )

        headers = [
            "学号 (必填)", "姓名 (必填)", "性别 (男/女)", "出生日期 (YYYY-MM-DD)",
            "学段 (初中/高中)", "届别年份 (纯数字，如2026)", "年级 (初一/初二/初三/高一/高二/高三)",
            "班级名称 (1班-20班)", "在校状态 (在读/转学/休学/复学/毕业)", "身份证号码", "学籍号",
            "家庭地址", "监护人姓名", "监护人联系电话",
            "入学日期 (YYYY-MM-DD)", "毕业日期 (YYYY-MM-DD)"
        ]

        # 添加表头
        for col, header in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        # 添加示例数据
        example_row = [
            "S001", "张三", "男", "2010-05-15",
            "初中", "2026", "初一", "1班", "在读", "", "",
            "", "张父", "13800138000", "2023-09-01", ""
        ]
        for col, value in enumerate(example_row, 1):
            cell = sheet.cell(row=2, column=col, value=value)
            cell.alignment = header_alignment
            cell.border = thin_border

        # 设置列宽
        col_widths = [15, 10, 8, 18, 15, 18, 22, 15, 20, 15, 15, 20, 12, 15, 15, 15]
        for col, width in enumerate(col_widths, 1):
            sheet.column_dimensions[chr(64 + col) if col <= 26 else 'A' + chr(64 + col - 26)].width = width

        # 设置行高
        sheet.row_dimensions[1].height = 25

        # 添加提示信息
        section_options = ', '.join([f"{s[0]}({s[1]})" for s in SECTION_CHOICES])
        sheet.cell(row=2, column=5).comment = openpyxl.comments.Comment(
            f"请填写: {section_options}", "System")

        cohort_options = ', '.join([c[0] for c in COHORT_YEAR_CHOICES])
        sheet.cell(row=2, column=6).comment = openpyxl.comments.Comment(
            f"请填写纯数字，如: {cohort_options}", "System")

        grade_level_options = ', '.join([g[0] for g in GRADE_LEVEL_CHOICES])
        sheet.cell(row=2, column=7).comment = openpyxl.comments.Comment(
            f"请填写: {grade_level_options}", "System")

        class_name_options = ', '.join([choice[0] for choice in CLASS_NAME_CHOICES])
        sheet.cell(row=2, column=8).comment = openpyxl.comments.Comment(
            f"请填写: {class_name_options}", "System")

        status_options = ', '.join([s[0] for s in STATUS_CHOICES])
        sheet.cell(row=2, column=9).comment = openpyxl.comments.Comment(
            f"请填写: {status_options}", "System")

        date_format_text = "日期格式必须是 YYYY-MM-DD，例如 2006-01-23"
        sheet.cell(row=2, column=4).comment = openpyxl.comments.Comment(date_format_text, "System")
        sheet.cell(row=2, column=15).comment = openpyxl.comments.Comment(date_format_text, "System")
        sheet.cell(row=2, column=16).comment = openpyxl.comments.Comment(date_format_text, "System")

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="student_import_template.xlsx"'
        workbook.save(response)
        return response

    @action(detail=False, methods=['post'], url_path='batch-import', parser_classes=[MultiPartParser, FormParser])
    def batch_import(self, request):
        """批量导入学生"""
        if 'file' not in request.FILES:
            return Response({'success': False, 'message': "请选择正确的 Excel 文件。"}, status=400)
        
        excel_file = request.FILES['file']
        
        if not excel_file.name.endswith(('.xlsx', '.xls')):
            return Response({'success': False, 'message': "文件格式不正确，请上传 .xlsx 或 .xls 文件。"}, status=400)

        try:
            workbook = openpyxl.load_workbook(excel_file)
            sheet = workbook.active
            
            header = [cell.value for cell in sheet[1]]
            
            header_mapping = {
                "学号 (必填)": "student_id",
                "姓名 (必填)": "name",
                "性别 (男/女)": "gender",
                "出生日期 (YYYY-MM-DD)": "date_of_birth",
                "年级 (初一/初二/初三/高一/高二/高三)": "grade_level",
                "学段 (初中/高中)": "_section",
                "届别年份 (纯数字，如2026)": "_cohort_year",
                "班级名称 (1班-20班)": "class_name",
                "在校状态 (在读/转学/休学/复学/毕业)": "status",
                "身份证号码": "id_card_number",
                "学籍号": "student_enrollment_number",
                "家庭地址": "home_address",
                "监护人姓名": "guardian_name",
                "监护人联系电话": "guardian_contact_phone",
                "入学日期 (YYYY-MM-DD)": "entry_date",
                "毕业日期 (YYYY-MM-DD)": "graduation_date",
            }

            imported_count = 0
            failed_rows = []
            success_messages = []
            error_messages = []
            warning_messages = []

            # 检测Excel中是否有重复学号
            excel_student_ids = []
            duplicate_rows = []

            for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                if not any(row):
                    continue
                row_data = dict(zip(header, row))
                sid = row_data.get('学号 (必填)') or row_data.get('学号')
                if sid:
                    sid = str(sid).strip()
                    if sid in excel_student_ids:
                        duplicate_rows.append({'row': row_idx, 'student_id': sid})
                    excel_student_ids.append(sid)

            if duplicate_rows:
                warning_messages.append(f"检测到Excel中有重复学号：{', '.join([r['student_id'] for r in duplicate_rows])}")

            from django.db import transaction

            for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                if not any(row):
                    continue

                try:
                    with transaction.atomic():
                        row_data = dict(zip(header, row))
                        student_data = {}
                        for excel_header, model_field in header_mapping.items():
                            if excel_header in row_data:
                                student_data[model_field] = row_data[excel_header]

                        if not student_data.get('student_id') or not student_data.get('name'):
                            raise ValueError("学号和姓名为必填字段")

                        if student_data.get('student_id'):
                            student_data['student_id'] = str(student_data['student_id']).strip()

                        if student_data.get('id_card_number'):
                            student_data['id_card_number'] = str(student_data['id_card_number']).strip()
                        # 学籍号：空字符串也设为None，避免唯一约束冲突
                        enrollment = student_data.get('student_enrollment_number')
                        if enrollment and str(enrollment).strip():
                            student_data['student_enrollment_number'] = str(enrollment).strip()
                        else:
                            student_data['student_enrollment_number'] = None

                        for date_field in ['date_of_birth', 'entry_date', 'graduation_date']:
                            if date_field in student_data:
                                date_value = student_data[date_field]
                                if date_value is None or date_value == '' or str(date_value).strip() == '':
                                    student_data[date_field] = None
                                    continue
                                
                                try:
                                    if isinstance(date_value, datetime.datetime):
                                        student_data[date_field] = date_value.date()
                                    elif isinstance(date_value, datetime.date):
                                        pass
                                    else:
                                        date_str = str(date_value).strip()
                                        if date_str:
                                            for date_format in ['%Y-%m-%d', '%Y/%m/%d', '%Y.%m.%d']:
                                                try:
                                                    student_data[date_field] = datetime.datetime.strptime(date_str, date_format).date()
                                                    break
                                                except ValueError:
                                                    continue
                                            else:
                                                student_data[date_field] = None
                                                warning_messages.append(f"第 {row_idx} 行的 '{date_field}' 日期格式不正确 (值: '{date_value}')")
                                        else:
                                            student_data[date_field] = None
                                except Exception as e:
                                    student_data[date_field] = None
                                    warning_messages.append(f"第 {row_idx} 行的 '{date_field}' 日期处理出错: {e}")
                            else:
                                student_data[date_field] = None

                        if 'gender' in student_data and student_data['gender']:
                            gender_value = str(student_data['gender']).strip()
                            gender_mapping = {
                                '男': '男', 'M': '男', 'Male': '男', 'male': '男', '1': '男',
                                '女': '女', 'F': '女', 'Female': '女', 'female': '女', '0': '女'
                            }
                            student_data['gender'] = gender_mapping.get(gender_value, gender_value)
                            if student_data['gender'] not in ['男', '女']:
                                warning_messages.append(f"第 {row_idx} 行的性别值 '{gender_value}' 无效，已设置为空")
                                student_data['gender'] = None

                        # 处理学段 + 届别年份 → cohort
                        section = student_data.pop('_section', None)
                        cohort_year = student_data.pop('_cohort_year', None)
                        class_name = student_data.pop('class_name', None)
                        current_class_obj = None

                        # grade_level 直接从 Excel 读取（传统格式如"初一"）
                        # cohort = section + cohort_year + "级"（如"初中2024级"）
                        if section and cohort_year:
                            cohort_year_str = str(cohort_year).strip()
                            cohort_value = f"{section}{cohort_year_str}级"
                            student_data['cohort'] = cohort_value

                        if class_name:
                            # 使用 grade_level + class_name 作为查找键，确保正确匹配
                            try:
                                current_class_obj, created = Class.objects.get_or_create(
                                    grade_level=student_data.get('grade_level', ''),
                                    class_name=class_name,
                                    defaults={
                                        'grade_level': student_data.get('grade_level', ''),
                                        'cohort': student_data.get('cohort', '')
                                    }
                                )
                                # 如果已存在但 cohort 为空，自动补全 cohort
                                if not created and not current_class_obj.cohort and student_data.get('cohort'):
                                    current_class_obj.cohort = student_data.get('cohort')
                                    current_class_obj.save(update_fields=['cohort'])
                                    warning_messages.append(
                                        f"第 {row_idx} 行：班级 {student_data.get('grade_level', '')}{class_name} 的 cohort 已自动补全为 {student_data.get('cohort')}"
                                    )
                                if created:
                                    success_messages.append(f"自动创建班级：{student_data.get('grade_level', '')}{class_name}")
                            except Exception as e:
                                raise ValueError(f"班级创建/查找失败: {e}")

                        student_data['current_class'] = current_class_obj

                        student_id = student_data.get('student_id')
                        student_obj, created = Student.objects.update_or_create(
                            student_id=student_id,
                            defaults=student_data
                        )

                        if created:
                            success_messages.append(f"成功新增学生：{student_obj.name} ({student_obj.student_id})")
                        else:
                            success_messages.append(f"成功更新学生：{student_obj.name} ({student_obj.student_id})")
                        imported_count += 1
                        
                except Exception as e:
                    failed_rows.append({'row': row_idx, 'error': str(e)})
                    error_messages.append(f"第 {row_idx} 行学生导入失败: {e}")
                    continue

            return Response({
                'success': True,
                'imported_count': imported_count,
                'failed_count': len(failed_rows),
                'success_messages': success_messages,
                'error_messages': error_messages,
                'warning_messages': warning_messages,
                'failed_rows': failed_rows
            })

        except Exception as e:
            return Response({
                'success': False,
                'message': f"解析文件时出现严重错误: {str(e)}"
            }, status=500)


class ClassViewSet(viewsets.ReadOnlyModelViewSet):
    """
    班级管理 ViewSet
    只提供查看操作
    """
    queryset = Class.objects.all()
    serializer_class = ClassSerializer
    permission_classes = [permissions.IsAuthenticated]
    filter_backends = [DjangoFilterBackend, filters.OrderingFilter]
    filterset_fields = ['cohort']  # grade_level 保留用于展示，cohort 用于查询
    ordering = ['grade_level', 'class_name']


class ExamViewSet(viewsets.ModelViewSet):
    """
    考试管理 ViewSet
    """
    queryset = Exam.objects.all()
    serializer_class = ExamSerializer
    permission_classes = [permissions.IsAuthenticated]
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = ['academic_year', 'grade_level']
    search_fields = ['name', 'description']
    ordering_fields = ['date', 'name']
    ordering = ['-date']

    def get_permissions(self):
        if self.action in ['create', 'update', 'partial_update', 'destroy']:
            return [permissions.IsAuthenticated(), IsAdminOrGradeManagerOrStaff()]
        return [permissions.IsAuthenticated()]

    @action(detail=False, methods=['get'], url_path='options')
    def options(self, request):
        """返回学年、年级下拉选项"""
        return Response({
            'academic_years': [{'value': choice[0], 'label': choice[1]} for choice in ACADEMIC_YEAR_CHOICES],
            'grade_levels': [{'value': choice[0], 'label': choice[1]} for choice in COHORT_CHOICES],
        })

    @action(detail=False, methods=['get'], url_path='default-subjects')
    def default_subjects(self, request):
        """根据年级和学年返回默认科目和满分配置"""
        from .models.exam import Exam, SUBJECT_DEFAULT_MAX_SCORES, SUBJECT_CHOICES
        grade_level = request.query_params.get('grade_level', '')
        academic_year = request.query_params.get('academic_year', '')

        # 创建一个临时 Exam 对象来计算基础年级
        exam = Exam(grade_level=grade_level, academic_year=academic_year)
        base_grade_level = exam.get_grade_level_from_cohort()
        grade_config = SUBJECT_DEFAULT_MAX_SCORES.get(base_grade_level, {})
        # 按 SUBJECT_CHOICES 顺序返回
        subject_order = [code for code, _ in SUBJECT_CHOICES]
        subjects = [
            {'subject_code': code, 'max_score': grade_config[code]}
            for code in subject_order
            if code in grade_config
        ]
        all_subjects = [{'value': code, 'label': label} for code, label in SUBJECT_CHOICES]
        return Response({
            'subjects': subjects,
            'all_subjects': all_subjects,
        })


class ScoreViewSet(viewsets.ModelViewSet):
    """
    成绩管理 API
    - 列表：按(学生,考试)聚合，输出前端 `/scores` 页面所需数据结构
    - 批量：导出选中、删除选中
    - 导入：Excel 批量导入
    """
    queryset = Score.objects.all().select_related('student', 'student__current_class', 'exam', 'exam_subject')
    serializer_class = ScoreSerializer
    permission_classes = [permissions.IsAuthenticated]
    parser_classes = [JSONParser, MultiPartParser, FormParser]

    def get_permissions(self):
        if self.action == 'target_students_query':
            return [permissions.IsAuthenticated()]

        if self.action in [
            'create', 'update', 'partial_update', 'destroy',
            'manual_add', 'batch_edit_save',
            'batch_delete_selected', 'batch_delete_filtered', 'batch_import'
        ]:
            return [permissions.IsAuthenticated(), IsAdminOrGradeManagerOrStaff()]
        return [permissions.IsAuthenticated()]

    @action(detail=False, methods=['post'], url_path='target-students-query')
    def target_students_query(self, request):
        """按规则筛选目标生（第一期：单条件 + 时序量词）。"""
        try:
            result = execute_target_student_rule(request.data)

            page = self._parse_pagination_value(
                request.query_params.get('page', request.data.get('page', 1)),
                default_value=1,
                field_name='page',
                min_value=1,
                max_value=100000,
            )
            page_size = self._parse_pagination_value(
                request.query_params.get('page_size', request.data.get('page_size', 200)),
                default_value=200,
                field_name='page_size',
                min_value=1,
                max_value=1000,
            )

            students = result.get('students', [])
            total = len(students)
            num_pages = ((total - 1) // page_size + 1) if total else 1

            if page > num_pages and total > 0:
                raise ValueError('page 超出范围')

            start = (page - 1) * page_size
            end = start + page_size
            paged_students = students[start:end]

            result['students'] = paged_students
            result['pagination'] = {
                'page': page,
                'page_size': page_size,
                'total': total,
                'num_pages': num_pages,
                'has_next': page < num_pages,
                'has_previous': page > 1,
            }

            return Response({'success': True, 'data': result})
        except ValueError as e:
            return Response({'success': False, 'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)
        except Exception as e:
            return Response({'success': False, 'error': f'服务器错误: {str(e)}'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    @staticmethod
    def _parse_pagination_value(raw_value, default_value, field_name, min_value, max_value):
        if raw_value in [None, '']:
            return default_value

        try:
            value = int(raw_value)
        except (TypeError, ValueError):
            raise ValueError(f'{field_name} 必须是整数')

        if value < min_value or value > max_value:
            raise ValueError(f'{field_name} 必须在 {min_value}-{max_value} 之间')

        return value

    def _get_subject_max_scores(self, exam):
        exam_subjects = ExamSubject.objects.filter(exam=exam)
        max_scores = {subject.subject_code: float(subject.max_score) for subject in exam_subjects}

        default_config = SUBJECT_DEFAULT_MAX_SCORES.get(exam.get_grade_level_from_cohort(), {})
        for subject_code, _ in SCORE_SUBJECT_CHOICES:
            if subject_code not in max_scores:
                max_scores[subject_code] = float(default_config.get(subject_code, 100))

        return max_scores

    def _filter_scores(self, request):
        scores = Score.objects.select_related('student', 'student__current_class', 'exam').order_by(
            'student__student_id', 'exam__date', 'subject'
        )

        student_id_filter = request.query_params.get('student_id_filter')
        student_name_filter = request.query_params.get('student_name_filter')
        exam_filter = request.query_params.get('exam_filter')
        subject_filter = request.query_params.get('subject_filter')
        subject_filters = request.query_params.getlist('subject_filters')
        subject_multi = request.query_params.getlist('subject')
        grade_filter = request.query_params.get('grade_filter')
        class_filter = request.query_params.get('class_filter')
        academic_year_filter = request.query_params.get('academic_year_filter')
        date_from_filter = request.query_params.get('date_from_filter')
        date_to_filter = request.query_params.get('date_to_filter')

        if subject_filter and isinstance(subject_filter, str) and ',' in subject_filter:
            subject_filters.extend([x.strip() for x in subject_filter.split(',') if x.strip()])
        if subject_multi:
            subject_filters.extend([x for x in subject_multi if x])

        if student_id_filter:
            scores = scores.filter(student__student_id__icontains=student_id_filter)
        if student_name_filter:
            scores = scores.filter(student__name__icontains=student_name_filter)
        if exam_filter:
            scores = scores.filter(exam__pk=exam_filter)
        if subject_filters:
            scores = scores.filter(subject__in=list(set(subject_filters)))
        elif subject_filter:
            scores = scores.filter(subject=subject_filter)
        if grade_filter:
            # grade_filter 现在传入 cohort 格式（如"初中2023级"），直接用 student__cohort 过滤
            scores = scores.filter(student__cohort=grade_filter)
        if class_filter:
            scores = scores.filter(student__current_class__class_name=class_filter)
        if academic_year_filter:
            scores = scores.filter(exam__academic_year=academic_year_filter)
        if date_from_filter:
            scores = scores.filter(exam__date__gte=date_from_filter)
        if date_to_filter:
            scores = scores.filter(exam__date__lte=date_to_filter)

        return scores

    def _aggregate_rows(self, scores):
        grade_label_map = {value: label for value, label in GRADE_LEVEL_CHOICES}
        aggregated_data = defaultdict(lambda: {
            'student_obj': None,
            'class_obj': None,
            'exam_obj': None,
            'scores': {},
            'total_score': 0.0,
            'grade_rank': None,
        })

        for score in scores:
            key = (score.student.pk, score.exam.pk)
            if aggregated_data[key]['student_obj'] is None:
                aggregated_data[key]['student_obj'] = score.student
                aggregated_data[key]['class_obj'] = score.student.current_class
                aggregated_data[key]['exam_obj'] = score.exam
                aggregated_data[key]['grade_rank'] = score.total_score_rank_in_grade
            aggregated_data[key]['scores'][score.subject] = float(score.score_value)
            aggregated_data[key]['total_score'] += float(score.score_value)

        rows = []
        for key, data in aggregated_data.items():
            student = data['student_obj']
            class_obj = data['class_obj']
            exam = data['exam_obj']
            rows.append({
                'record_key': f"{student.pk}_{exam.pk}",
                'student_id': student.pk,
                'exam_id': exam.pk,
                'student': {
                    'student_id': student.student_id,
                    'name': student.name,
                    'cohort': student.cohort or '',
                    'grade_level': student.grade_level,
                    'grade_level_display': grade_label_map.get(student.grade_level, student.grade_level),
                },
                'class': {
                    'class_name': class_obj.class_name if class_obj else None,
                },
                'exam': {
                    'id': exam.pk,
                    'name': exam.name,
                    'academic_year': exam.academic_year,
                    'date': exam.date.strftime('%Y-%m-%d') if exam.date else '',
                },
                'scores': data['scores'],
                'total_score': round(data['total_score'], 2),
                'grade_rank': data['grade_rank'],
            })
        return rows

    def _sort_rows(self, rows, request):
        sort_by = request.query_params.get('sort_by')
        subject_sort = request.query_params.get('subject_sort')
        sort_order = request.query_params.get('sort_order', 'desc')
        reverse = sort_order == 'desc'

        if subject_sort:
            if subject_sort == 'total_score':
                rows.sort(key=lambda x: float(x.get('total_score') or 0), reverse=reverse)
            elif subject_sort == 'grade_rank':
                rows.sort(
                    key=lambda x: (x.get('grade_rank') is None, x.get('grade_rank') if x.get('grade_rank') is not None else 999999),
                    reverse=reverse,
                )
            else:
                rows.sort(key=lambda x: float(x.get('scores', {}).get(subject_sort, -1)), reverse=reverse)
            return rows

        if sort_by == 'total_score_desc':
            rows.sort(key=lambda x: float(x.get('total_score') or 0), reverse=True)
        elif sort_by == 'total_score_asc':
            rows.sort(key=lambda x: float(x.get('total_score') or 0), reverse=False)
        elif sort_by == 'student_name':
            rows.sort(key=lambda x: x.get('student', {}).get('name') or '')
        elif sort_by == 'exam_date':
            rows.sort(key=lambda x: x.get('exam', {}).get('date') or '', reverse=True)
        elif sort_by == 'grade_rank':
            rows.sort(key=lambda x: x.get('grade_rank') if x.get('grade_rank') is not None else 999999)

        return rows

    def _build_export_workbook(self, rows):
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "成绩导出"

        all_subjects = [subject_code for subject_code, _ in SCORE_SUBJECT_CHOICES]
        headers = ["学号", "学生姓名", "届别", "年级", "班级", "考试名称", "学年", "考试日期"] + all_subjects
        sheet.append(headers)

        for row in rows:
            output_row = [
                row['student']['student_id'],
                row['student']['name'],
                row['student']['cohort'],
                row['student']['grade_level_display'],
                row['class']['class_name'] or "N/A",
                row['exam']['name'],
                row['exam']['academic_year'] or "N/A",
                row['exam']['date'] or "",
            ]
            for subject in all_subjects:
                value = row['scores'].get(subject)
                output_row.append(value if value is not None else "-")
            sheet.append(output_row)

        return workbook

    def _build_query_export_workbook(self, rows, all_subjects):
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "成绩查询导出"

        headers = ["学号", "学生姓名", "入学级别", "年级", "班级", "考试名称", "学年", "考试日期"] + all_subjects + ["总分", "年级排名"]
        sheet.append(headers)

        for row in rows:
            output_row = [
                row['student']['student_id'],
                row['student']['name'],
                row['student']['cohort'],
                row['student']['grade_level_display'],
                row['class']['class_name'] or "N/A",
                row['exam']['name'],
                row['exam']['academic_year'] or "N/A",
                row['exam']['date'] or "",
            ]

            for subject in all_subjects:
                value = row['scores'].get(subject)
                output_row.append(value if value is not None else "-")

            output_row.append(row.get('total_score', 0))
            output_row.append(row.get('grade_rank') if row.get('grade_rank') is not None else "-")
            sheet.append(output_row)

        return workbook

    @action(detail=False, methods=['get'], url_path='options')
    def options(self, request):
        grade_level_filter = request.query_params.get('grade_level')
        exams = Exam.objects.all()
        if grade_level_filter:
            exams = exams.filter(grade_level=grade_level_filter)
        exams = exams.order_by('-academic_year', '-date', 'name')[:100]
        academic_year_filter = request.query_params.get('academic_year')
        all_exams_qs = Exam.objects.all()
        if grade_level_filter:
            all_exams_qs = all_exams_qs.filter(grade_level=grade_level_filter)
        academic_year_values = sorted(
            {
                value
                for value in all_exams_qs.exclude(academic_year__isnull=True)
                .exclude(academic_year='')
                .values_list('academic_year', flat=True)
                if value
            },
            reverse=True,
        )
        return Response({
            'exams': [
                {
                    'value': str(exam.pk),
                    'label': f"{exam.academic_year} {exam.name} ({exam.get_grade_level_display()})"
                }
                for exam in exams
            ],
            'grade_levels': [{'value': value, 'label': label} for value, label in COHORT_CHOICES],
            'class_name_choices': [{'value': value, 'label': label} for value, label in CLASS_NAME_CHOICES],
            'subjects': [{'value': value, 'label': label} for value, label in SCORE_SUBJECT_CHOICES],
            'academic_years': [{'value': value, 'label': value} for value in academic_year_values],
            'sort_by_options': [
                {'value': '', 'label': '--- 默认排序 ---'},
                {'value': 'total_score_desc', 'label': '总分降序'},
                {'value': 'total_score_asc', 'label': '总分升序'},
                {'value': 'student_name', 'label': '学生姓名'},
                {'value': 'exam_date', 'label': '考试日期'},
                {'value': 'grade_rank', 'label': '年级排名'},
            ],
            'all_subjects': [value for value, _ in SCORE_SUBJECT_CHOICES],
            'per_page_options': [10, 20, 50, 100],
        })

    @action(detail=False, methods=['get'], url_path='student-search')
    def student_search(self, request):
        query = (request.query_params.get('q') or '').strip()
        if not query:
            return Response({'results': []})

        students = Student.objects.select_related('current_class').filter(
            models.Q(name__icontains=query)
            | models.Q(student_id__icontains=query)
            | models.Q(current_class__class_name__icontains=query)
        ).order_by('name', 'student_id')[:20]

        return Response({
            'results': [
                {
                    'id': student.pk,
                    'student_id': student.student_id,
                    'name': student.name,
                    'grade_level': student.grade_level,
                    'grade_level_display': student.get_grade_level_display() if student.grade_level else '',
                    'class_name': student.current_class.class_name if student.current_class else '',
                    'display': f"{student.name} ({student.student_id}) - {student.get_grade_level_display() if student.grade_level else ''}{student.current_class.class_name if student.current_class else ''}",
                }
                for student in students
            ]
        })

    @action(detail=False, methods=['get'], url_path='student-analysis-data')
    def student_analysis_data(self, request):
        """获取学生个人成绩分析数据（前后端分离 API 版本）"""
        student_id = request.query_params.get('student_id')
        exam_ids = request.query_params.get('exam_ids', '')
        exam_id = request.query_params.get('exam_id')

        if not student_id:
            return Response({'success': False, 'error': '缺少学生ID参数'}, status=status.HTTP_400_BAD_REQUEST)

        if exam_ids:
            exam_id_list = [item.strip() for item in exam_ids.split(',') if item.strip()]
        elif exam_id:
            exam_id_list = [exam_id]
        else:
            student_exam_ids = Score.objects.filter(student_id=student_id).values_list('exam_id', flat=True).distinct()
            exam_id_list = [str(item) for item in student_exam_ids]

        try:
            student = Student.objects.select_related('current_class').get(id=student_id)
            exams = Exam.objects.filter(id__in=exam_id_list).order_by('date', 'id')

            if not exams.exists():
                return Response({'success': False, 'error': '未找到指定的考试'}, status=status.HTTP_404_NOT_FOUND)

            analysis_data = {
                'student_info': {
                    'id': student.id,
                    'student_id': student.student_id,
                    'name': student.name,
                    'grade_level': student.grade_level,
                    'class_name': student.current_class.class_name if student.current_class else '未分班',
                },
                'exams': [],
                'subjects': [],
                'trend_data': {},
                'summary': {
                    'total_exams': exams.count(),
                    'subjects_count': 0,
                }
            }

            all_subjects = set()

            def get_subject_order(subject):
                for index, (subject_code, subject_name) in enumerate(SCORE_SUBJECT_CHOICES):
                    if subject_code == subject or subject_name == subject:
                        return index
                return 999

            for exam in exams:
                scores = Score.objects.filter(student=student, exam=exam).select_related('exam_subject')
                scores_list = list(scores)
                scores_list.sort(key=lambda score: get_subject_order(score.subject))
                exam_subject_max_scores = {
                    item.subject_code: float(item.max_score)
                    for item in exam.exam_subjects.all()
                }

                exam_data = {
                    'id': exam.id,
                    'name': exam.name,
                    'academic_year': exam.academic_year,
                    'exam_date': exam.date.strftime('%Y-%m-%d') if exam.date else None,
                    'grade_level': exam.get_grade_level_display(),
                    'scores': [],
                    'total_score': 0,
                    'average_score': 0,
                    'grade_total_rank': None,
                    'class_total_rank': None,
                }

                total_score = 0
                valid_scores = 0

                for score in scores_list:
                    subject_name = score.subject
                    all_subjects.add(subject_name)

                    score_value = float(score.score_value) if score.score_value else 0

                    if score.exam_subject and score.exam_subject.max_score is not None:
                        full_score = float(score.exam_subject.max_score)
                    elif subject_name in exam_subject_max_scores:
                        full_score = exam_subject_max_scores[subject_name]
                    else:
                        max_score = score.get_max_score()
                        full_score = float(max_score) if max_score else 0

                    percentage = round((score_value / full_score) * 100, 1) if full_score > 0 else 0

                    exam_data['scores'].append({
                        'subject_name': subject_name,
                        'score_value': score_value,
                        'full_score': full_score,
                        'grade_rank': score.grade_rank_in_subject,
                        'class_rank': score.class_rank_in_subject,
                        'percentage': percentage,
                    })

                    if score.score_value is not None:
                        total_score += float(score.score_value)
                        valid_scores += 1

                    if subject_name not in analysis_data['trend_data']:
                        analysis_data['trend_data'][subject_name] = {
                            'class_ranks': [],
                            'grade_ranks': [],
                            'scores': [],
                            'exam_names': [],
                            'exam_ids': [],
                        }

                    analysis_data['trend_data'][subject_name]['class_ranks'].append(score.class_rank_in_subject)
                    analysis_data['trend_data'][subject_name]['grade_ranks'].append(score.grade_rank_in_subject)
                    analysis_data['trend_data'][subject_name]['scores'].append(score_value)
                    analysis_data['trend_data'][subject_name]['exam_names'].append(exam.name)
                    analysis_data['trend_data'][subject_name]['exam_ids'].append(exam.id)

                exam_data['total_score'] = round(total_score, 1)
                if valid_scores > 0:
                    exam_data['average_score'] = round(total_score / valid_scores, 1)

                if scores_list:
                    first_score = scores_list[0]
                    exam_data['grade_total_rank'] = first_score.total_score_rank_in_grade
                    exam_data['class_total_rank'] = first_score.total_score_rank_in_class

                analysis_data['exams'].append(exam_data)

            analysis_data['subjects'] = sorted(list(all_subjects), key=get_subject_order)
            analysis_data['summary']['subjects_count'] = len(all_subjects)
            analysis_data['trend_data']['total'] = {
                'class_ranks': [item['class_total_rank'] for item in analysis_data['exams']],
                'grade_ranks': [item['grade_total_rank'] for item in analysis_data['exams']],
                'scores': [item['total_score'] for item in analysis_data['exams']],
                'exam_names': [item['name'] for item in analysis_data['exams']],
                'exam_ids': [item['id'] for item in analysis_data['exams']],
            }

            return Response({'success': True, 'data': analysis_data})

        except Student.DoesNotExist:
            return Response({'success': False, 'error': '学生不存在'}, status=status.HTTP_404_NOT_FOUND)
        except Exception as e:
            return Response({'success': False, 'error': f'服务器错误: {str(e)}'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    @action(detail=False, methods=['get'], url_path='class-analysis-single')
    def class_analysis_single(self, request):
        """获取单班级成绩分析数据（前后端分离 API 版本）"""
        exam_id = request.query_params.get('exam')
        grade_level = request.query_params.get('grade_level')
        academic_year = request.query_params.get('academic_year', '')
        class_name_param = request.query_params.get('class_name')
        selected_classes = request.query_params.getlist('selected_classes')

        if not exam_id:
            return Response({'success': False, 'error': '缺少考试ID参数'}, status=status.HTTP_400_BAD_REQUEST)

        class_id = None
        if selected_classes:
            if 'all' in selected_classes:
                return Response({'success': False, 'error': '该接口仅支持单班级分析，不支持 all'}, status=status.HTTP_400_BAD_REQUEST)
            class_id = selected_classes[0]
        elif class_name_param:
            if ',' in class_name_param:
                return Response({'success': False, 'error': '该接口仅支持单班级分析，请只传入一个班级'}, status=status.HTTP_400_BAD_REQUEST)
            class_id = class_name_param

        if not class_id:
            return Response({'success': False, 'error': '缺少班级参数'}, status=status.HTTP_400_BAD_REQUEST)

        try:
            exam = Exam.objects.get(id=exam_id)
        except Exam.DoesNotExist:
            return Response({'success': False, 'error': '考试不存在'}, status=status.HTTP_404_NOT_FOUND)

        try:
            target_class = Class.objects.get(id=class_id)
        except Class.DoesNotExist:
            return Response({'success': False, 'error': '班级不存在'}, status=status.HTTP_404_NOT_FOUND)

        if grade_level and target_class.cohort != grade_level:
            return Response({'success': False, 'error': '班级与年级参数不一致'}, status=status.HTTP_400_BAD_REQUEST)

        try:
            scores = Score.objects.filter(
                exam=exam,
                student__current_class=target_class
            )

            analysis_result = analyze_single_class(scores, target_class, exam)
            chart_data = json.loads(analysis_result.get('chart_data_json', '{}') or '{}')

            subject_stats_dict = analysis_result.get('subject_stats', {}) or {}
            ordered_subject_stats = []
            for subject_code, _ in SCORE_SUBJECT_CHOICES:
                if subject_code in subject_stats_dict:
                    item = subject_stats_dict[subject_code]
                    ordered_subject_stats.append({
                        'code': subject_code,
                        'name': item.get('name', subject_code),
                        'avg_score': float(item.get('avg_score', 0) or 0),
                        'actual_max_score': float(item.get('actual_max_score', 0) or 0),
                        'actual_min_score': float(item.get('actual_min_score', 0) or 0),
                        'count': int(item.get('count', 0) or 0),
                        'exam_max_score': float(item.get('exam_max_score', 0) or 0),
                    })

            data = {
                'selected_exam': {
                    'id': exam.id,
                    'name': exam.name,
                    'academic_year': exam.academic_year,
                    'grade_level': exam.grade_level,
                    'grade_level_display': exam.get_grade_level_display(),
                },
                'selected_grade': grade_level or target_class.grade_level,
                'academic_year': academic_year or exam.academic_year,
                'target_class': {
                    'id': target_class.id,
                    'grade_level': target_class.grade_level,
                    'class_name': target_class.class_name,
                },
                'total_students': int(analysis_result.get('total_students', 0) or 0),
                'class_avg_total': float(analysis_result.get('class_avg_total', 0) or 0),
                'class_max_total': float(analysis_result.get('class_max_total', 0) or 0),
                'class_min_total': float(analysis_result.get('class_min_total', 0) or 0),
                'subject_stats': ordered_subject_stats,
                'student_rankings': analysis_result.get('student_rankings', []) or [],
                'chart_data': chart_data,
            }

            return Response({'success': True, 'data': data})
        except Exception as e:
            return Response({'success': False, 'error': f'服务器错误: {str(e)}'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    @action(detail=False, methods=['get'], url_path='class-analysis-multi')
    def class_analysis_multi(self, request):
        """获取多班级成绩对比分析数据（前后端分离 API 版本）"""
        exam_id = request.query_params.get('exam')
        grade_level = request.query_params.get('grade_level')
        academic_year = request.query_params.get('academic_year', '')
        class_name_param = request.query_params.get('class_name', '')
        selected_classes_param = request.query_params.getlist('selected_classes')

        if not exam_id:
            return Response({'success': False, 'error': '缺少考试ID参数'}, status=status.HTTP_400_BAD_REQUEST)

        class_id_values = [value for value in selected_classes_param if value and value != 'all']
        if not class_id_values and class_name_param:
            class_id_values = [item.strip() for item in class_name_param.split(',') if item.strip() and item.strip() != 'all']

        if len(class_id_values) < 2:
            return Response({'success': False, 'error': '多班级分析至少需要选择2个班级'}, status=status.HTTP_400_BAD_REQUEST)

        try:
            selected_class_ids = [int(value) for value in class_id_values]
        except ValueError:
            return Response({'success': False, 'error': '班级参数格式错误'}, status=status.HTTP_400_BAD_REQUEST)

        try:
            exam = Exam.objects.get(id=exam_id)
        except Exam.DoesNotExist:
            return Response({'success': False, 'error': '考试不存在'}, status=status.HTTP_404_NOT_FOUND)

        selected_classes = list(Class.objects.filter(id__in=selected_class_ids))
        if len(selected_classes) < 2:
            return Response({'success': False, 'error': '有效班级不足，无法进行多班级分析'}, status=status.HTTP_400_BAD_REQUEST)

        if grade_level:
            invalid_classes = [item for item in selected_classes if item.cohort != grade_level]
            if invalid_classes:
                return Response({'success': False, 'error': '存在与年级参数不一致的班级'}, status=status.HTTP_400_BAD_REQUEST)

        selected_classes = sorted(selected_classes, key=lambda item: selected_class_ids.index(item.id))

        try:
            analysis_result = analyze_multiple_classes(selected_classes, exam)
            chart_data = json.loads(analysis_result.get('chart_data_json', '{}') or '{}')

            data = {
                'selected_exam': {
                    'id': exam.id,
                    'name': exam.name,
                    'academic_year': exam.academic_year,
                    'grade_level': exam.grade_level,
                    'grade_level_display': exam.get_grade_level_display(),
                },
                'selected_grade': grade_level or exam.grade_level,
                'academic_year': academic_year or exam.academic_year,
                'selected_classes': [f"{item.grade_level}{item.class_name}" for item in selected_classes],
                'class_statistics': analysis_result.get('class_statistics', []) or [],
                'subjects': analysis_result.get('subjects', []) or [],
                'total_students': int(analysis_result.get('total_students', 0) or 0),
                'subject_count': int(analysis_result.get('subject_count', 0) or 0),
                'highest_avg': float(analysis_result.get('highest_avg', 0) or 0),
                'chart_data': chart_data,
            }

            return Response({'success': True, 'data': data})
        except Exception as e:
            return Response({'success': False, 'error': f'服务器错误: {str(e)}'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    @action(detail=False, methods=['get'], url_path='class-analysis-grade')
    def class_analysis_grade(self, request):
        """获取年级成绩分析数据（前后端分离 API 版本）"""
        exam_id = request.query_params.get('exam')
        grade_level = request.query_params.get('grade_level')
        academic_year = request.query_params.get('academic_year', '')

        if not exam_id:
            return Response({'success': False, 'error': '缺少考试ID参数'}, status=status.HTTP_400_BAD_REQUEST)

        if not grade_level:
            return Response({'success': False, 'error': '缺少年级参数'}, status=status.HTTP_400_BAD_REQUEST)

        try:
            exam = Exam.objects.get(id=exam_id)
        except Exam.DoesNotExist:
            return Response({'success': False, 'error': '考试不存在'}, status=status.HTTP_404_NOT_FOUND)

        try:
            analysis_result = analyze_grade(exam, grade_level)
            chart_data = json.loads(analysis_result.get('chart_data_json', '{}') or '{}')

            data = {
                'selected_exam': {
                    'id': exam.id,
                    'name': exam.name,
                    'academic_year': exam.academic_year,
                    'grade_level': exam.grade_level,
                    'grade_level_display': exam.get_grade_level_display(),
                },
                'selected_grade': grade_level,
                'academic_year': academic_year or exam.academic_year,
                'total_students': int(analysis_result.get('total_students', 0) or 0),
                'total_classes': int(analysis_result.get('total_classes', 0) or 0),
                'grade_avg_score': float(analysis_result.get('grade_avg_score', 0) or 0),
                'excellent_rate': float(analysis_result.get('excellent_rate', 0) or 0),
                'class_statistics': analysis_result.get('class_statistics', []) or [],
                'subjects': analysis_result.get('subjects', []) or [],
                'total_max_score': float(analysis_result.get('total_max_score', 0) or 0),
                'chart_data': chart_data,
            }

            return Response({'success': True, 'data': data})
        except Exception as e:
            return Response({'success': False, 'error': f'服务器错误: {str(e)}'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    @action(detail=False, methods=['post'], url_path='manual-add')
    def manual_add(self, request):
        student_id = request.data.get('student_id')
        exam_id = request.data.get('exam_id')
        scores = request.data.get('scores', {})

        if not student_id or not exam_id:
            return Response({'success': False, 'message': '学生和考试为必填项'}, status=status.HTTP_400_BAD_REQUEST)

        try:
            student = Student.objects.get(pk=student_id)
        except Student.DoesNotExist:
            return Response({'success': False, 'message': '选择的学生不存在'}, status=status.HTTP_400_BAD_REQUEST)

        try:
            exam = Exam.objects.get(pk=exam_id)
        except Exam.DoesNotExist:
            return Response({'success': False, 'message': '选择的考试不存在'}, status=status.HTTP_400_BAD_REQUEST)

        valid_scores = {}
        for subject_code, _ in SCORE_SUBJECT_CHOICES:
            raw_value = scores.get(subject_code)
            if raw_value in [None, '']:
                continue
            try:
                valid_scores[subject_code] = float(raw_value)
            except (TypeError, ValueError):
                return Response({'success': False, 'message': f'{subject_code} 分数格式不正确'}, status=status.HTTP_400_BAD_REQUEST)

        if not valid_scores:
            return Response({'success': False, 'message': '请至少输入一个科目的成绩'}, status=status.HTTP_400_BAD_REQUEST)

        existing_subjects = []
        subject_name_map = {code: name for code, name in SCORE_SUBJECT_CHOICES}
        for subject_code in valid_scores.keys():
            if Score.objects.filter(student=student, exam=exam, subject=subject_code).exists():
                existing_subjects.append(subject_name_map.get(subject_code, subject_code))

        if existing_subjects:
            return Response({
                'success': False,
                'code': 'duplicate_scores',
                'message': f"以下科目的成绩已存在：{', '.join(existing_subjects)}。",
                'duplicate_subjects': existing_subjects,
                'student_id': student.pk,
                'exam_id': exam.pk,
            }, status=status.HTTP_400_BAD_REQUEST)

        exam_subject_map = {s.subject_code: s for s in exam.exam_subjects.all()}

        created_count = 0
        with transaction.atomic():
            for subject_code, score_value in valid_scores.items():
                score = Score(
                    student=student,
                    exam=exam,
                    subject=subject_code,
                    score_value=score_value,
                    exam_subject=exam_subject_map.get(subject_code)
                )
                score.save()
                created_count += 1

        try:
            update_all_rankings_async.delay(exam.pk, student.grade_level)
        except Exception:
            pass

        return Response({
            'success': True,
            'message': f'成功添加 {created_count} 个科目的成绩',
            'created_count': created_count,
        })

    @action(detail=False, methods=['get'], url_path='batch-edit-detail')
    def batch_edit_detail(self, request):
        student_id = request.query_params.get('student') or request.query_params.get('student_id')
        exam_id = request.query_params.get('exam') or request.query_params.get('exam_id')

        if not student_id or not exam_id:
            return Response({'success': False, 'message': '缺少必要参数：学生ID或考试ID'}, status=status.HTTP_400_BAD_REQUEST)

        try:
            student = Student.objects.select_related('current_class').get(pk=student_id)
        except Student.DoesNotExist:
            return Response({'success': False, 'message': '学生不存在'}, status=status.HTTP_404_NOT_FOUND)

        try:
            exam = Exam.objects.get(pk=exam_id)
        except Exam.DoesNotExist:
            return Response({'success': False, 'message': '考试不存在'}, status=status.HTTP_404_NOT_FOUND)

        existing_scores = {
            score.subject: float(score.score_value)
            for score in Score.objects.filter(student=student, exam=exam)
        }

        return Response({
            'success': True,
            'student': {
                'id': student.pk,
                'name': student.name,
                'student_id': student.student_id,
                'grade_level': student.grade_level,
                'grade_level_display': student.get_grade_level_display() if student.grade_level else '',
                'class_name': student.current_class.class_name if student.current_class else '',
            },
            'exam': {
                'id': exam.pk,
                'name': exam.name,
                'academic_year': exam.academic_year,
                'date': exam.date.strftime('%Y-%m-%d') if exam.date else '',
            },
            'subjects': [{'value': code, 'label': label} for code, label in SCORE_SUBJECT_CHOICES],
            'existing_scores': existing_scores,
            'subject_max_scores': self._get_subject_max_scores(exam),
        })

    @action(detail=False, methods=['post'], url_path='batch-edit-save')
    def batch_edit_save(self, request):
        student_id = request.data.get('student_id')
        exam_id = request.data.get('exam_id')
        scores = request.data.get('scores', {})

        if not student_id or not exam_id:
            return Response({'success': False, 'message': '缺少必要参数：学生ID或考试ID'}, status=status.HTTP_400_BAD_REQUEST)

        try:
            student = Student.objects.get(pk=student_id)
        except Student.DoesNotExist:
            return Response({'success': False, 'message': '学生不存在'}, status=status.HTTP_404_NOT_FOUND)

        try:
            exam = Exam.objects.get(pk=exam_id)
        except Exam.DoesNotExist:
            return Response({'success': False, 'message': '考试不存在'}, status=status.HTTP_404_NOT_FOUND)

        subject_max_scores = self._get_subject_max_scores(exam)
        subject_name_map = {code: name for code, name in SCORE_SUBJECT_CHOICES}
        exam_subject_map = {subject.subject_code: subject for subject in exam.exam_subjects.all()}

        updated_count = 0
        created_count = 0
        deleted_count = 0

        try:
            with transaction.atomic():
                for subject_code, _ in SCORE_SUBJECT_CHOICES:
                    raw_value = scores.get(subject_code)

                    if raw_value in [None, '']:
                        deleted_count += Score.objects.filter(
                            student=student,
                            exam=exam,
                            subject=subject_code
                        ).delete()[0]
                        continue

                    try:
                        score_value = float(raw_value)
                    except (TypeError, ValueError):
                        return Response({'success': False, 'message': f"{subject_name_map.get(subject_code, subject_code)} 的分数格式不正确"}, status=status.HTTP_400_BAD_REQUEST)

                    max_score = float(subject_max_scores.get(subject_code, 100))
                    if score_value < 0 or score_value > max_score:
                        return Response({
                            'success': False,
                            'message': f"{subject_name_map.get(subject_code, subject_code)} 的分数必须在0-{max_score:g}分之间"
                        }, status=status.HTTP_400_BAD_REQUEST)

                    score_obj, created = Score.objects.update_or_create(
                        student=student,
                        exam=exam,
                        subject=subject_code,
                        defaults={
                            'score_value': score_value,
                            'exam_subject': exam_subject_map.get(subject_code)
                        }
                    )

                    score_obj.clean()
                    score_obj.save()

                    if created:
                        created_count += 1
                    else:
                        updated_count += 1

        except Exception as e:
            return Response({'success': False, 'message': str(e)}, status=status.HTTP_400_BAD_REQUEST)

        try:
            update_all_rankings_async.delay(exam.pk, student.grade_level)
        except Exception:
            pass

        return Response({
            'success': True,
            'message': '成功修改成绩！',
            'created_count': created_count,
            'updated_count': updated_count,
            'deleted_count': deleted_count,
        })

    def list(self, request, *args, **kwargs):
        rows = self._aggregate_rows(self._filter_scores(request))
        rows = self._sort_rows(rows, request)

        subject_order = [value for value, _ in SCORE_SUBJECT_CHOICES]
        if request.query_params.get('dynamic_subjects') in ['1', 'true', 'True']:
            subject_set = set()
            for row in rows:
                subject_set.update(row.get('scores', {}).keys())
            all_subjects = [subject for subject in subject_order if subject in subject_set]
        else:
            all_subjects = subject_order

        page = request.query_params.get('page', '1')
        page_size = request.query_params.get('page_size', '100')
        try:
            page_size = int(page_size)
        except (TypeError, ValueError):
            page_size = 100
        page_size = max(10, min(100, page_size))

        paginator = Paginator(rows, page_size)
        page_obj = paginator.get_page(page)

        return Response({
            'count': paginator.count,
            'num_pages': paginator.num_pages,
            'current_page': page_obj.number,
            'has_previous': page_obj.has_previous(),
            'has_next': page_obj.has_next(),
            'previous_page': page_obj.previous_page_number() if page_obj.has_previous() else None,
            'next_page': page_obj.next_page_number() if page_obj.has_next() else None,
            'start_index': page_obj.start_index() if paginator.count else 0,
            'end_index': page_obj.end_index() if paginator.count else 0,
            'page_size': page_size,
            'results': list(page_obj),
            'all_subjects': all_subjects,
        })

    @action(detail=False, methods=['get'], url_path='download-template')
    def download_template(self, request):
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "成绩导入模板"

        headers = ["学号", "学生姓名"] + [name for name, _ in SCORE_SUBJECT_CHOICES]
        sheet.append(headers)

        row_a = ['S001', '张三'] + [''] * len(SCORE_SUBJECT_CHOICES)
        if '语文' in headers:
            row_a[headers.index('语文')] = 85.5
        if '数学' in headers:
            row_a[headers.index('数学')] = 92.0
        sheet.append(row_a)

        row_b = ['S002', '李四'] + [''] * len(SCORE_SUBJECT_CHOICES)
        if '英语' in headers:
            row_b[headers.index('英语')] = 78.0
        if '物理' in headers:
            row_b[headers.index('物理')] = 90.0
        sheet.append(row_b)

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="score_import_template.xlsx"'
        workbook.save(response)
        return response

    @action(detail=False, methods=['post'], url_path='batch-delete-selected')
    def batch_delete_selected(self, request):
        selected_records = request.data.get('selected_records', [])
        if not selected_records:
            return Response({'success': False, 'message': '没有选择任何记录'}, status=status.HTTP_400_BAD_REQUEST)

        total_deleted = 0
        affected_exam_ids = set()

        for record in selected_records:
            try:
                student_id, exam_id = str(record).split('_')
            except ValueError:
                continue
            deleted_count = Score.objects.filter(student_id=student_id, exam_id=exam_id).delete()[0]
            total_deleted += deleted_count
            if deleted_count > 0:
                affected_exam_ids.add(int(exam_id))

        for exam_id in affected_exam_ids:
            try:
                update_all_rankings_async.delay(exam_id)
            except Exception:
                pass

        return Response({
            'success': True,
            'deleted_count': total_deleted,
            'message': f'成功删除 {total_deleted} 条成绩记录' if total_deleted else '没有找到对应的成绩记录'
        })

    @action(detail=False, methods=['post'], url_path='batch-delete-filtered')
    def batch_delete_filtered(self, request):
        """按当前筛选条件批量删除成绩（与旧 score_batch_delete_filtered 行为对齐）"""
        filtered_scores = self._filter_scores(request)
        delete_count = filtered_scores.count()

        if delete_count == 0:
            return Response({
                'success': True,
                'deleted_count': 0,
                'message': '没有符合筛选条件的成绩记录'
            })

        affected_exam_ids = list(filtered_scores.values_list('exam_id', flat=True).distinct())
        filtered_scores.delete()

        for exam_id in affected_exam_ids:
            try:
                update_all_rankings_async.delay(exam_id)
            except Exception:
                pass

        return Response({
            'success': True,
            'deleted_count': delete_count,
            'message': f'成功删除 {delete_count} 条符合筛选条件的成绩记录'
        })

    @action(detail=False, methods=['get'], url_path='select-all-record-keys')
    def select_all_record_keys(self, request):
        score_pairs = self._filter_scores(request).values_list('student_id', 'exam_id').distinct()
        record_keys = [f"{student_id}_{exam_id}" for student_id, exam_id in score_pairs]

        return Response({
            'success': True,
            'count': len(record_keys),
            'record_keys': record_keys,
        })

    @action(detail=False, methods=['post'], url_path='batch-export-selected')
    def batch_export_selected(self, request):
        selected_records = request.data.get('selected_records', [])
        if not selected_records:
            return Response({'success': False, 'message': '没有选择任何记录'}, status=status.HTTP_400_BAD_REQUEST)

        selected_rows = []
        for record in selected_records:
            try:
                student_id, exam_id = str(record).split('_')
            except ValueError:
                continue
            scores = Score.objects.filter(student_id=student_id, exam_id=exam_id).select_related(
                'student', 'student__current_class', 'exam'
            )
            selected_rows.extend(self._aggregate_rows(scores))

        workbook = self._build_export_workbook(selected_rows)
        timestamp = datetime.datetime.now().strftime('%Y%m%d_%H%M%S')
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="聚合成绩导出_{timestamp}.xlsx"'
        workbook.save(response)
        return response

    @action(detail=False, methods=['get'], url_path='batch-export')
    def batch_export(self, request):
        rows = self._aggregate_rows(self._filter_scores(request))
        workbook = self._build_export_workbook(rows)
        timestamp = datetime.datetime.now().strftime('%Y%m%d_%H%M%S')
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="筛选成绩导出_{timestamp}.xlsx"'
        workbook.save(response)
        return response

    @action(detail=False, methods=['get'], url_path='query-export')
    def query_export(self, request):
        rows = self._aggregate_rows(self._filter_scores(request))
        rows = self._sort_rows(rows, request)

        subject_order = [value for value, _ in SCORE_SUBJECT_CHOICES]
        if request.query_params.get('dynamic_subjects') in ['1', 'true', 'True']:
            subject_set = set()
            for row in rows:
                subject_set.update(row.get('scores', {}).keys())
            all_subjects = [subject for subject in subject_order if subject in subject_set]
        else:
            all_subjects = subject_order

        workbook = self._build_query_export_workbook(rows, all_subjects)
        timestamp = datetime.datetime.now().strftime('%Y%m%d_%H%M%S')
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="成绩查询导出_{timestamp}.xlsx"'
        workbook.save(response)
        return response

    @action(detail=False, methods=['post'], url_path='batch-import', parser_classes=[MultiPartParser, FormParser])
    def batch_import(self, request):
        start_time = timezone.now()
        excel_file = request.FILES.get('excel_file')
        exam_id = request.data.get('exam')

        if not exam_id:
            return Response({'success': False, 'message': '请选择考试'}, status=status.HTTP_400_BAD_REQUEST)
        if not excel_file:
            return Response({'success': False, 'message': '请选择Excel文件'}, status=status.HTTP_400_BAD_REQUEST)
        if not excel_file.name.endswith(('.xlsx', '.xls')):
            return Response({'success': False, 'message': '文件格式不正确，请上传 .xlsx 或 .xls 文件'}, status=status.HTTP_400_BAD_REQUEST)

        try:
            exam = Exam.objects.get(pk=exam_id)
        except Exam.DoesNotExist:
            return Response({'success': False, 'message': '考试不存在'}, status=status.HTTP_400_BAD_REQUEST)

        try:
            workbook = openpyxl.load_workbook(excel_file)
            sheet = workbook.active
            headers = [cell.value for cell in sheet[1]]

            exam_subject_map = {
                subject.subject_code: subject
                for subject in exam.exam_subjects.all()
            }
            max_score_map = {
                subject.subject_code: float(subject.max_score)
                for subject in exam.exam_subjects.all()
            }

            if not max_score_map:
                default_config = SUBJECT_DEFAULT_MAX_SCORES.get(exam.get_grade_level_from_cohort(), {})
                for subject_code, _ in SCORE_SUBJECT_CHOICES:
                    if subject_code in default_config:
                        max_score_map[subject_code] = float(default_config[subject_code])

            subject_codes = [subject_code for subject_code, _ in SCORE_SUBJECT_CHOICES if subject_code in headers]

            student_ids = set()
            excel_rows = []
            for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                if not any(row):
                    continue
                row_data = dict(zip(headers, row))
                sid = row_data.get('学号')
                if sid:
                    sid = str(sid).strip()
                    student_ids.add(sid)
                excel_rows.append((row_idx, row_data))

            students = Student.objects.filter(student_id__in=student_ids).select_related('current_class')
            students_map = {student.student_id: student for student in students}

            imported_count = 0
            failed_count = 0
            error_details = []

            with transaction.atomic():
                for row_idx, row_data in excel_rows:
                    student_id = str(row_data.get('学号') or '').strip()
                    student_name = str(row_data.get('学生姓名') or '').strip()

                    row_errors = []
                    if not student_id:
                        row_errors.append('缺少学号')
                    student = students_map.get(student_id)
                    if not student:
                        row_errors.append(f'学号 {student_id} 对应的学生不存在')

                    changed_any = False
                    if not row_errors:
                        for subject_code in subject_codes:
                            raw_score = row_data.get(subject_code)
                            if raw_score in [None, '']:
                                continue
                            try:
                                score_value = float(raw_score)
                            except (TypeError, ValueError):
                                row_errors.append(f'{subject_code} 分数格式错误')
                                continue

                            if score_value < 0:
                                row_errors.append(f'{subject_code} 分数不能为负数')
                                continue

                            max_score = max_score_map.get(subject_code)
                            if max_score is not None and score_value > max_score:
                                row_errors.append(f'{subject_code} 分数 {score_value} 超过满分 {max_score}')
                                continue

                            Score.objects.update_or_create(
                                student=student,
                                exam=exam,
                                subject=subject_code,
                                defaults={
                                    'score_value': score_value,
                                    'exam_subject': exam_subject_map.get(subject_code)
                                }
                            )
                            changed_any = True

                    if row_errors:
                        failed_count += 1
                        error_details.append({
                            'row': row_idx,
                            'student_id': student_id,
                            'student_name': student_name,
                            'errors': row_errors,
                        })
                    elif changed_any:
                        imported_count += 1

            try:
                update_all_rankings_async.delay(exam.pk)
            except Exception:
                pass

            execution_time = (timezone.now() - start_time).total_seconds()
            return Response({
                'success': True,
                'message': '导入完成',
                'imported_count': imported_count,
                'failed_count': failed_count,
                'error_details': error_details,
                'execution_time': round(execution_time, 2),
            })
        except Exception as e:
            return Response({
                'success': False,
                'message': f'文件处理失败：{str(e)}'
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

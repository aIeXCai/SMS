from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.views.decorators.http import require_http_methods, require_POST
from django.db import transaction
from django.core.exceptions import ValidationError
from django.db.models import Sum, Count, Q, F, Case, When, IntegerField, Avg, Max, Min
from django.core.paginator import Paginator
from django.http import HttpResponse, JsonResponse
import openpyxl
from datetime import datetime
from decimal import Decimal
import io, json, statistics, re, time
from collections import defaultdict

# 从新的models导入
from ..models import (
    Exam, ExamSubject, Score, Student, Class, 
    SUBJECT_CHOICES, SUBJECT_DEFAULT_MAX_SCORES,
    ACADEMIC_YEAR_CHOICES, GRADE_LEVEL_CHOICES, CLASS_NAME_CHOICES
)
from ..config import PAGINATION_CONFIG, TABLE_CONFIG, IMPORT_CONFIG
from ..forms import ScoreForm, ScoreBatchUploadForm, ScoreAddForm, ScoreQueryForm, ScoreAnalysisForm

# --- 成績管理 Views ---

# <!-- 基础成绩管理 -->
# 成績列表 (Read)
def score_list(request):
    # 获取筛选参数
    student_id_filter = request.GET.get('student_id_filter')
    student_name_filter = request.GET.get('student_name_filter')
    exam_filter = request.GET.get('exam_filter')
    subject_filter = request.GET.get('subject_filter')
    grade_filter = request.GET.get('grade_filter')
    class_filter = request.GET.get('class_filter')
    
    # 优化查询：使用select_related减少数据库查询次数
    scores = Score.objects.select_related(
        'student', 
        'student__current_class', 
        'exam'
    ).order_by('student__student_id', 'exam__date', 'subject')

    # 应用筛选
    if student_id_filter:
        scores = scores.filter(student__student_id__icontains=student_id_filter)
    if student_name_filter:
        scores = scores.filter(student__name__icontains=student_name_filter)
    if exam_filter:
        scores = scores.filter(exam__pk=exam_filter)
    if subject_filter:
        scores = scores.filter(subject=subject_filter)
    if grade_filter:
        scores = scores.filter(student__grade_level=grade_filter)
    if class_filter:
        scores = scores.filter(student__current_class__class_name=class_filter)

    # --- 优化的数据聚合逻辑 ---
    # 限制查询结果数量，避免内存溢出
    scores = scores[:2000]  # 限制最多处理2000条记录
    
    aggregated_data = defaultdict(lambda: {
        'student_obj': None,
        'class_obj': None,
        'exam_obj': None,
        'scores': {}
    })

    # 获取所有科目用于表头显示
    all_subjects = [choice[0] for choice in SUBJECT_CHOICES]

    # 聚合成绩数据
    for score in scores:
        key = (score.student.pk, score.exam.pk)
        
        if aggregated_data[key]['student_obj'] is None:
            aggregated_data[key]['student_obj'] = score.student
            aggregated_data[key]['class_obj'] = score.student.current_class
            aggregated_data[key]['exam_obj'] = score.exam
        
        # 将科目成绩添加到scores字典中
        aggregated_data[key]['scores'][score.subject] = score.score_value

    # 转换为列表格式供模板使用
    final_display_rows = []
    for key, data in aggregated_data.items():
        row_data_obj = type('obj', (object,), data)
        final_display_rows.append(row_data_obj)

    # 添加分页功能 - 使用配置文件设置
    per_page = request.GET.get('per_page', PAGINATION_CONFIG['DEFAULT_PER_PAGE'])
    try:
        per_page = int(per_page)
        # 使用配置文件中的限制范围
        if per_page < PAGINATION_CONFIG['MIN_PER_PAGE']:
            per_page = PAGINATION_CONFIG['MIN_PER_PAGE']
        elif per_page > PAGINATION_CONFIG['MAX_PER_PAGE']:
            per_page = PAGINATION_CONFIG['MAX_PER_PAGE']
    except (ValueError, TypeError):
        per_page = PAGINATION_CONFIG['DEFAULT_PER_PAGE']
    
    paginator = Paginator(final_display_rows, per_page)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    # 获取筛选选项数据（优化查询）
    # 只获取有成绩记录的学生，避免加载所有学生数据
    students = Student.objects.filter(
        pk__in=Score.objects.values_list('student_id', flat=True).distinct()
    ).select_related('current_class').order_by('name', 'student_id')[:100]  # 限制数量
    
    # 获取所有考试用于批量导入功能，按时间倒序排列
    exams = Exam.objects.all().order_by('-academic_year', '-date', 'name')[:100]  # 适当限制数量

    context = {
        'aggregated_scores': page_obj,  # 使用分页对象
        'all_subjects': all_subjects,   # 所有科目列表
        'students': students,
        'exams': exams,
        'subjects': SUBJECT_CHOICES,
        'grade_levels': GRADE_LEVEL_CHOICES,
        'class_name_choices': CLASS_NAME_CHOICES,
        
        # 分页信息
        'page_obj': page_obj,
        'is_paginated': page_obj.has_other_pages(),
        'per_page': per_page,  # 当前每页显示数量
        'per_page_options': PAGINATION_CONFIG['PER_PAGE_OPTIONS'],  # 使用配置文件中的选项
        
        # 保持筛选状态
        'selected_student_id_filter': student_id_filter,
        'selected_student_name_filter': student_name_filter,
        'selected_exam_filter': exam_filter,
        'selected_subject_filter': subject_filter,
        'selected_grade_filter': grade_filter,
        'selected_class_filter': class_filter,
    }
    return render(request, 'scores/score_list.html', context)

# 新增成績 (Create)
@require_http_methods(["GET", "POST"])
def score_add(request):
    if request.method == 'POST':
        form = ScoreAddForm(request.POST)
        if form.is_valid():
            try:
                with transaction.atomic():
                    student = form.cleaned_data['student']
                    exam = form.cleaned_data['exam']
                    
                    created_count = 0
                    for subject_code, subject_name in SUBJECT_CHOICES:
                        score_field = f'score_{subject_code}'
                        score_value = form.cleaned_data.get(score_field)
                        
                        if score_value is not None:
                            Score.objects.create(
                                student=student,
                                exam=exam,
                                subject=subject_code,
                                score_value=score_value
                            )
                            created_count += 1
                    
                    # 提交异步任务：更新排名
                    try:
                        from school_management.students_grades.tasks import update_all_rankings_async
                        job = update_all_rankings_async.delay(exam.pk, student.grade_level)
                        messages.success(request, f"成功添加 {created_count} 个科目的成绩！排名计算已提交到后台处理（任务ID: {job.id}）")
                    except Exception as e:
                        messages.warning(request, f'成绩添加成功，但排名更新任务提交失败: {e}')
                    
                    return redirect('students_grades:score_list')
            except Exception as e:
                messages.error(request, f"添加成绩失败：{e}")
        else:
            messages.error(request, "添加成绩失败，请检查表单数据。")
    else:
        form = ScoreAddForm()
    
    # Create a dictionary of score fields for easier template access
    score_fields = {}
    for subject_code, subject_name in SUBJECT_CHOICES:
        field_name = f'score_{subject_code}'
        if field_name in form.fields:
            score_fields[subject_code] = form[field_name]
    
    context = {
        'form': form,
        'title': '新增成绩',
        'subjects': SUBJECT_CHOICES,
        'score_fields': score_fields
    }
    
    return render(request, 'scores/score_form.html', context)

# 編輯成績 (Update)
@require_http_methods(["GET", "POST"])
def score_edit(request, pk):
    score = get_object_or_404(Score, pk=pk)
    if request.method == 'POST':
        form = ScoreForm(request.POST, instance=score)
        if form.is_valid():
            try:
                updated_score = form.save()
                messages.success(request, "成绩信息更新成功！")
                
                # 提交异步任务：更新排名
                try:
                    from school_management.students_grades.tasks import update_all_rankings_async
                    job = update_all_rankings_async.delay(updated_score.exam.pk, updated_score.student.grade_level)
                    messages.success(request, f'成绩更新成功！排名计算已提交到后台处理（任务ID: {job.id}）')
                except Exception as e:
                    messages.warning(request, f'成绩更新成功，但排名更新任务提交失败: {e}')
                    
                return redirect('students_grades:score_list')
            except Exception as e:
                messages.error(request, f"更新成绩失败：{e}。请检查是否已存在该学生在该考试该科目的成绩。")
        else:
            messages.error(request, "更新成绩信息失败，请检查表单数据。")
    else:
        form = ScoreForm(instance=score)
    
    return render(request, 'scores/score_form.html', {'form': form, 'title': '編輯成績'})

# 批量编辑学生在某考试中的所有科目成绩
@require_http_methods(["GET", "POST"])
def score_batch_edit(request):
    student_id = request.GET.get('student')
    exam_id = request.GET.get('exam')
    
    if not student_id or not exam_id:
        messages.error(request, '缺少必要的参数：学生ID或考试ID')
        return redirect('students_grades:score_list')
    
    try:
        student = Student.objects.get(pk=student_id)
        exam = Exam.objects.get(pk=exam_id)
    except (Student.DoesNotExist, Exam.DoesNotExist):
        messages.error(request, '学生或考试不存在')
        return redirect('students_grades:score_list')
    
    if request.method == 'POST':
        # 处理表单提交
        updated_count = 0
        created_count = 0
        validation_errors = []  # 收集验证错误
        
        try:
            with transaction.atomic():
                for subject_code, subject_name in SUBJECT_CHOICES:
                    score_value = request.POST.get(f'score_{subject_code}')
                    
                    if score_value and score_value.strip():  # 如果有输入分数
                        try:
                            score_value = float(score_value)
                            score_obj, created = Score.objects.update_or_create(
                                student=student,
                                exam=exam,
                                subject=subject_code,
                                defaults={'score_value': score_value}
                            )
                            # 手动调用clean方法进行验证
                            try:
                                score_obj.clean()
                            except ValidationError as e:
                                # 收集验证错误信息
                                error_msg = str(e.message) if hasattr(e, 'message') else str(e)
                                validation_errors.append(f"{subject_name}：{error_msg}")
                                continue
                            
                            if created:
                                created_count += 1
                            else:
                                updated_count += 1
                        except ValueError:
                            validation_errors.append(f'{subject_name} 的分数格式不正确')
                            continue
                    else:
                        # 如果分数为空，删除已存在的成绩记录
                        Score.objects.filter(
                            student=student,
                            exam=exam,
                            subject=subject_code
                        ).delete()
                
                # 如果有验证错误，抛出异常回滚事务
                if validation_errors:
                    raise ValidationError(validation_errors)
        
        except ValidationError:
            # 将验证错误信息合并为一个消息
            error_message = "\n".join(validation_errors)
            messages.error(request, error_message)
            return render(request, 'scores/score_batch_edit.html', {
                'student': student,
                'exam': exam,
                'existing_scores': get_existing_scores(student, exam),
                'subjects': SUBJECT_CHOICES,
                'subject_max_scores': get_subject_max_scores(exam)  # 添加满分信息
            })
        
        if created_count > 0 or updated_count > 0:
            messages.success(request, f'成功修改成绩！')
            # 提交异步任务：更新排名
            try:
                from school_management.students_grades.tasks import update_all_rankings_async
                job = update_all_rankings_async.delay(exam.pk, student.grade_level)
                messages.info(request, f'排名计算已提交到后台处理（任务ID: {job.id}）')
            except Exception as e:
                messages.warning(request, f'成绩保存成功，但排名更新任务提交失败: {e}')
        else:
            messages.info(request, '没有检测到任何更改')
        
        return redirect('students_grades:score_list')
    
    # GET请求：显示编辑表单
    existing_scores = get_existing_scores(student, exam)
    
    context = {
        'student': student,
        'exam': exam,
        'existing_scores': existing_scores,
        'subjects': SUBJECT_CHOICES,
        'subject_max_scores': get_subject_max_scores(exam)  # 添加满分信息
    }
    return render(request, 'scores/score_batch_edit.html', context)

# 辅助函数：获取学生在某考试中的现有成绩
def get_existing_scores(student, exam):
    scores = Score.objects.filter(student=student, exam=exam)
    score_dict = {}
    for score in scores:
        score_dict[score.subject] = score.score_value
    return score_dict

# 辅助函数：获取考试中每个科目的满分
def get_subject_max_scores(exam):
    """
    获取指定考试中每个科目的满分配置
    返回格式：{'chinese': 150, 'math': 150, 'english': 120, ...}
    """
    exam_subjects = ExamSubject.objects.filter(exam=exam)
    max_scores = {}
    
    for exam_subject in exam_subjects:
        max_scores[exam_subject.subject_code] = exam_subject.max_score
    
    # 对于没有配置的科目，使用默认满分
    for subject_code, subject_name in SUBJECT_CHOICES:
        if subject_code not in max_scores:
            default_config = SUBJECT_DEFAULT_MAX_SCORES.get(exam.grade_level, {})
            max_scores[subject_code] = default_config.get(subject_code, 100)
    
    return max_scores


# <!--批量操作-->
# 批量導入成績 (Excel) - AJAX版本
@require_POST
def score_batch_import_ajax(request):
    """优化版AJAX批量导入成绩接口"""
    start_time = time.time()
    timing_info = {}
    
    try:
        form = ScoreBatchUploadForm(request.POST, request.FILES)
        if not form.is_valid():
            return JsonResponse({
                'success': False,
                'message': '表单验证失败，请检查文件格式和考试选择。',
                'errors': form.errors
            })
        
        timing_info['form_validation'] = time.time() - start_time
        
        excel_file = request.FILES['excel_file']
        selected_exam = form.cleaned_data['exam']
        
        step_start = time.time()
        workbook = openpyxl.load_workbook(excel_file)
        sheet = workbook.active
        headers = [cell.value for cell in sheet[1]]
        timing_info['excel_loading'] = time.time() - step_start
        
        # 第一步：收集所有学号，进行批量查询
        step_start = time.time()
        all_student_ids = set()
        excel_data = []
        
        for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            row_data = dict(zip(headers, row))
            student_id = row_data.get("学号")
            if student_id:
                all_student_ids.add(str(student_id).strip())
                excel_data.append((row_idx, row_data))
        
        timing_info['data_collection'] = time.time() - step_start
        
        # 第二步：批量查询所有学生信息 - 性能优化关键点1
        step_start = time.time()
        students_dict = {}
        if all_student_ids:
            students_qs = Student.objects.filter(
                student_id__in=all_student_ids
            ).select_related('current_class')  # 使用select_related减少查询
            
            for student in students_qs:
                students_dict[student.student_id] = student
        
        timing_info['student_query'] = time.time() - step_start
        
        # 第三步：批量查询现有成绩记录 - 性能优化关键点2
        step_start = time.time()
        existing_scores = {}
        if all_student_ids:
            existing_scores_qs = Score.objects.filter(
                student__student_id__in=all_student_ids,
                exam=selected_exam
            ).select_related('student')
            
            for score in existing_scores_qs:
                key = (score.student.student_id, score.subject)
                existing_scores[key] = score
        
        timing_info['existing_scores_query'] = time.time() - step_start
        
        # 获取科目列表
        all_valid_subjects = {choice[0] for choice in SUBJECT_CHOICES}
        
        # 数据验证和收集阶段
        scores_to_create = []
        scores_to_update = []
        error_details = []
        successful_students = set()
        failed_students = set()
        
        # 第四步：处理Excel数据
        for row_idx, row_data in excel_data:
            student_id_from_excel = str(row_data.get("学号", "")).strip()
            student_name_from_excel = str(row_data.get("学生姓名", "")).strip()
            
            errors_in_row = []
            
            # 验证学生是否存在（使用缓存的查询结果）
            if student_id_from_excel not in students_dict:
                errors_in_row.append(f"学号 '{student_id_from_excel}' 对应的学生不存在")
                failed_students.add(student_id_from_excel)
                continue
            
            student_obj = students_dict[student_id_from_excel]
            
            # 验证姓名是否匹配
            if student_name_from_excel and student_obj.name != student_name_from_excel:
                errors_in_row.append(f"学号 '{student_id_from_excel}' 对应的学生姓名不匹配")
                failed_students.add(student_id_from_excel)
                continue
            
            # 处理该学生的所有科目成绩
            student_has_valid_scores = False
            for col_header, score_value_from_excel in row_data.items():
                if col_header in all_valid_subjects:
                    # 验证分数
                    if score_value_from_excel is not None and str(score_value_from_excel).strip() != '':
                        try:
                            score_value = float(score_value_from_excel)
                            if not (0 <= score_value <= 200):
                                errors_in_row.append(f"{col_header}分数超出有效范围(0-200)")
                                continue
                        except ValueError:
                            errors_in_row.append(f"{col_header}分数格式不正确")
                            continue
                        
                        # 准备成绩数据 - 性能优化关键点3：批量准备而不是逐个保存
                        score_key = (student_id_from_excel, col_header)
                        if score_key in existing_scores:
                            # 更新现有成绩
                            existing_score = existing_scores[score_key]
                            existing_score.score_value = score_value
                            scores_to_update.append(existing_score)
                        else:
                            # 创建新成绩
                            scores_to_create.append(Score(
                                student=student_obj,
                                exam=selected_exam,
                                subject=col_header,
                                score_value=score_value
                            ))
                        
                        student_has_valid_scores = True
            
            # 记录处理结果
            if errors_in_row:
                failed_students.add(student_id_from_excel)
                error_details.append({
                    'row': row_idx,
                    'student_id': student_id_from_excel,
                    'student_name': student_name_from_excel,
                    'errors': errors_in_row
                })
            elif student_has_valid_scores:
                successful_students.add(student_id_from_excel)
        
        # 第五步：批量数据库操作 - 性能优化关键点4
        step_start = time.time()
        with transaction.atomic():
            # 批量创建新成绩记录
            if scores_to_create:
                Score.objects.bulk_create(
                    scores_to_create, 
                    batch_size=1000,  # 分批处理，避免内存问题
                    ignore_conflicts=False
                )
            
            # 批量更新现有成绩记录
            if scores_to_update:
                Score.objects.bulk_update(
                    scores_to_update, 
                    ['score_value'], 
                    batch_size=1000
                )
        
        timing_info['database_operations'] = time.time() - step_start
        
        # 统计结果
        imported_count = len(successful_students)
        failed_count = len(failed_students)
        
        # 排名更新（异步处理优先，失败时跳过而不是同步处理）
        if imported_count > 0:
            # 短暂延迟确保数据库事务完全提交
            time.sleep(0.1)
            
            ranking_update_status = "skipped"
            try:
                # 检查是否安装了django-rq和Redis可用性
                try:
                    import django_rq
                    from ..tasks import update_all_rankings_async
                    
                    # 获取异步队列
                    queue = django_rq.get_queue('default')
                    
                    # 测试Redis连接
                    queue.connection.ping()  # 这会抛出异常如果Redis不可用
                    
                    # 提交异步任务：更新完整排名（总分和单科的年级、班级排名）
                    job = queue.enqueue(
                        update_all_rankings_async, 
                        selected_exam.pk,
                        job_timeout=1200  # 20分钟超时
                    )
                    
                    # 记录异步任务ID
                    print(f"异步完整排名更新任务已提交: job_id={job.id}")
                    ranking_update_status = "async_submitted"
                    
                except ImportError:
                    # django-rq未安装
                    print("django-rq未安装，跳过排名更新")
                    ranking_update_status = "django_rq_not_installed"
                    
                except Exception as redis_error:
                    # Redis连接失败或其他异步任务问题
                    print(f"Redis/异步任务不可用，跳过排名更新: {redis_error}")
                    ranking_update_status = "redis_unavailable"
                    
            except Exception as e:
                # 任何其他异常都跳过排名更新
                print(f"排名更新跳过，原因: {e}")
                ranking_update_status = "error_skipped"
        
        # 计算执行时间
        execution_time = time.time() - start_time
        
        # 生成详细的时间报告
        timing_report = f"总耗时: {execution_time:.2f}s"
        for step, duration in timing_info.items():
            timing_report += f" | {step}: {duration:.2f}s"
        
        # 返回结果
        if imported_count > 0:
            message = f"上传成功！成功导入 {imported_count} 个学生，{timing_report}"
            if failed_count > 0:
                message += f"，失败 {failed_count} 个学生"
            
            # 添加排名更新状态信息
            if 'ranking_update_status' in locals():
                if ranking_update_status == "async_submitted":
                    message += "。排名更新任务已提交至后台队列"
                elif ranking_update_status == "redis_unavailable":
                    message += "。注意：Redis未运行，排名更新已跳过"
                elif ranking_update_status == "django_rq_not_installed":
                    message += "。注意：django-rq未安装，排名更新已跳过"
                else:
                    message += "。注意：排名更新已跳过"
            
            return JsonResponse({
                'success': True,
                'message': message,
                'imported_count': imported_count,
                'failed_count': failed_count,
                'execution_time': round(execution_time, 2),
                'timing_details': timing_info,
                'ranking_update_status': locals().get('ranking_update_status', 'unknown'),
                'error_details': error_details[:20]  # 只返回前20个错误详情，避免响应过大
            })
        else:
            return JsonResponse({
                'success': False,
                'message': f"上传失败！失败 {failed_count} 个学生，耗时 {execution_time:.2f} 秒",
                'imported_count': 0,
                'failed_count': failed_count,
                'execution_time': round(execution_time, 2),
                'error_details': error_details[:20]
            })
            
    except Exception as e:
        execution_time = time.time() - start_time if 'start_time' in locals() else 0
        return JsonResponse({
            'success': False,
            'message': f"文件处理失败：{str(e)}，耗时 {execution_time:.2f} 秒",
            'imported_count': 0,
            'failed_count': 0,
            'execution_time': round(execution_time, 2)
        })

# 下載成績導入模板
def download_score_import_template(request):
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "成績導入模板"

    # 定義 Excel 表頭 (包括固定列和所有科目列)
    headers = ["学号", "学生姓名"] # 固定列
    
    # 按照 SUBJECT_CHOICES 的順序添加科目列
    for subject_name, _ in SUBJECT_CHOICES:
        headers.append(subject_name)
    
    sheet.append(headers)

    # 添加一些提示或示例數據
    # 學生 A 的語文和數學成績
    row_a = ['S001', '張三'] + [''] * len(SUBJECT_CHOICES) # 初始化空列表
    row_a[headers.index('语文')] = 85.5
    row_a[headers.index('数学')] = 92.0
    sheet.append(row_a)

    # 學生 B 的英語和物理成績
    row_b = ['S002', '李四'] + [''] * len(SUBJECT_CHOICES)
    row_b[headers.index('英语')] = 78.0
    row_b[headers.index('物理')] = 90.0
    sheet.append(row_b)

    # 添加備註和數據驗證提示
    # 可以在學號列添加提示
    sheet.cell(row=1, column=headers.index("学号") + 1).comment = openpyxl.comments.Comment("请填写学生学号，系统将通过此学号查找学生。必填。", "System")
    sheet.cell(row=1, column=headers.index("学生姓名") + 1).comment = openpyxl.comments.Comment("此列仅供参考，不参与导入判断。建议填写以方便核对。", "System")

    # 為每個科目列添加分數格式提示
    for subject_name, _ in SUBJECT_CHOICES:
        if subject_name in headers: # 確保科目在頭部列表中
            col_idx = headers.index(subject_name) + 1 # Excel 列是從 1 開始的
            sheet.cell(row=1, column=col_idx).comment = openpyxl.comments.Comment("成绩可以是整数或小数 (例如 85 或 92.5)。范围 0-100。留空表示无该科目成绩。", "System")

            # 可以在這裡添加數據驗證（雖然 openpyxl 的數據驗證比較基礎，且可能不被所有 Excel 軟體完美支持）
            # from openpyxl.worksheet.datavalidation import DataValidation
            # dv = DataValidation(type="whole", operator="between", formula1=0, formula2=100)
            # sheet.add_data_validation(dv)
            # dv.add('{}2:{}1000'.format(openpyxl.utils.get_column_letter(col_idx), openpyxl.utils.get_column_letter(col_idx)))


    excel_file = io.BytesIO()
    workbook.save(excel_file)
    excel_file.seek(0)

    response = HttpResponse(
        excel_file.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="score_import_template_pivot.xlsx"' # 更改文件名以區分
    return response

# 批量導出筛选后的成绩
def score_batch_export(request):
    # 获取筛选参数（与score_list相同的逻辑）
    scores = Score.objects.all()
    
    student_id_filter = request.GET.get('student_id_filter')
    student_name_filter = request.GET.get('student_name_filter')
    exam_filter = request.GET.get('exam_filter')
    subject_filter = request.GET.get('subject_filter')
    grade_filter = request.GET.get('grade_filter')
    class_filter = request.GET.get('class_filter')

    # 应用相同的筛选逻辑
    if student_id_filter:
        scores = scores.filter(student__student_id__icontains=student_id_filter)
    if student_name_filter:
        scores = scores.filter(student__name__icontains=student_name_filter)
    if exam_filter:
        scores = scores.filter(exam__pk=exam_filter)
    if subject_filter:
        scores = scores.filter(subject=subject_filter)
    if grade_filter:
        scores = scores.filter(student__grade_level=grade_filter)
    if class_filter:
        scores = scores.filter(student__current_class__class_name=class_filter)
    
    # 聚合数据逻辑（与score_list相同）
    aggregated_data = defaultdict(lambda: {
        'student_obj': None,
        'class_obj': None,
        'exam_obj': None,
        'scores': {}
    })
    
    # 获取所有科目用于表头
    all_subjects = [choice[0] for choice in SUBJECT_CHOICES]
    subject_names = {choice[0]: choice[1] for choice in SUBJECT_CHOICES}
    
    # 聚合成绩数据
    for score in scores.select_related('student', 'exam', 'student__current_class'):
        key = (score.student.pk, score.exam.pk)
        
        if aggregated_data[key]['student_obj'] is None:
            aggregated_data[key]['student_obj'] = score.student
            aggregated_data[key]['class_obj'] = score.student.current_class
            aggregated_data[key]['exam_obj'] = score.exam
        
        # 将科目成绩添加到scores字典中
        aggregated_data[key]['scores'][score.subject] = score.score_value
    
    # 创建Excel文件（与上面相同的逻辑）
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "筛选成绩导出"
    
    # 设置表头
    headers = [
        "学号", "学生姓名", "年级", "班级", "考试名称", "学年", "考试日期"
    ]
    for subject_code in all_subjects:
        headers.append(subject_names[subject_code])
    
    sheet.append(headers)
    
    # 添加聚合数据
    for key, data in aggregated_data.items():
        student = data['student_obj']
        exam = data['exam_obj']
        class_obj = data['class_obj']
        scores = data['scores']
        
        row = [
            student.student_id,
            student.name,
            student.get_grade_level_display() if student.grade_level else "N/A",
            class_obj.class_name if class_obj else "N/A",
            exam.name,
            exam.academic_year or "N/A",
            exam.date.strftime('%Y-%m-%d')
        ]
        
        # 添加各科目成绩
        for subject_code in all_subjects:
            score_value = scores.get(subject_code)
            if score_value is not None:
                row.append(float(score_value))
            else:
                row.append("-")
        
        sheet.append(row)
    
    # 设置样式和列宽（与上面相同）
    base_widths = [15, 15, 10, 10, 20, 15, 12]
    subject_widths = [10] * len(all_subjects)
    column_widths = base_widths + subject_widths
    
    for i, width in enumerate(column_widths, 1):
        sheet.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width
    
    # 生成文件名
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    filename = f"筛选成绩导出_{timestamp}.xlsx"
    
    # 返回Excel文件
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    workbook.save(response)
    return response

# 批量删除筛选后的成绩
@require_POST
def score_batch_delete_filtered(request):
    # 获取筛选参数
    scores = Score.objects.all()
    
    student_id_filter = request.POST.get('student_id_filter')
    student_name_filter = request.POST.get('student_name_filter')
    exam_filter = request.POST.get('exam_filter')
    subject_filter = request.POST.get('subject_filter')
    grade_filter = request.POST.get('grade_filter')
    class_filter = request.POST.get('class_filter')

    # 应用相同的筛选逻辑
    if student_id_filter:
        scores = scores.filter(student__student_id__icontains=student_id_filter)
    if student_name_filter:
        scores = scores.filter(student__name__icontains=student_name_filter)
    if exam_filter:
        scores = scores.filter(exam__pk=exam_filter)
    if subject_filter:
        scores = scores.filter(subject=subject_filter)
    if grade_filter:
        scores = scores.filter(student__grade_level=grade_filter)
    if class_filter:
        scores = scores.filter(student__current_class__class_name=class_filter)

    # 执行删除
    count = scores.count()
    if count > 0:
        # 获取删除记录的考试ID，用于后续排名更新
        exam_ids = set(scores.values_list('exam_id', flat=True).distinct())
        
        scores.delete()
        messages.success(request, f'成功删除 {count} 条成绩记录')
        
        # 异步更新相关考试的排名
        from school_management.students_grades.tasks import update_all_rankings_async
        for exam_id in exam_ids:
            try:
                update_all_rankings_async.delay(exam_id)
                messages.info(request, f'已提交考试 {exam_id} 的排名更新任务到后台处理')
            except Exception as e:
                messages.warning(request, f'排名更新任务提交失败: {str(e)}')
    else:
        messages.info(request, '没有找到符合条件的成绩记录')
    
    return redirect('students_grades:score_list')


# 导出选中的成绩记录
@require_POST
def score_batch_export_selected(request):
    selected_records = request.POST.getlist('selected_records')
    
    if not selected_records:
        messages.error(request, '没有选择任何记录')
        return redirect('students_grades:score_list')
    
    # 解析选中的记录并聚合数据
    aggregated_data = defaultdict(lambda: {
        'student_obj': None,
        'class_obj': None,
        'exam_obj': None,
        'scores': {}
    })
    
    # 获取所有科目用于表头
    all_subjects = [choice[0] for choice in SUBJECT_CHOICES]
    subject_names = {choice[0]: choice[1] for choice in SUBJECT_CHOICES}
    
    for record in selected_records:
        try:
            student_id, exam_id = record.split('_')
            scores = Score.objects.filter(
                student_id=student_id,
                exam_id=exam_id
            ).select_related('student', 'exam', 'student__current_class')
            
            key = (student_id, exam_id)
            
            for score in scores:
                if aggregated_data[key]['student_obj'] is None:
                    aggregated_data[key]['student_obj'] = score.student
                    aggregated_data[key]['class_obj'] = score.student.current_class
                    aggregated_data[key]['exam_obj'] = score.exam
                
                # 将科目成绩添加到scores字典中
                aggregated_data[key]['scores'][score.subject] = score.score_value
                
        except ValueError:
            continue
    
    if not aggregated_data:
        messages.error(request, '没有找到对应的成绩记录')
        return redirect('students_grades:score_list')
    
    # 创建Excel文件
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "聚合成绩导出"
    
    # 设置表头 - 与页面显示一致
    headers = [
        "学号", "学生姓名", "年级", "班级", "考试名称", "学年", "考试日期"
    ]
    # 添加所有科目列
    for subject_code in all_subjects:
        headers.append(subject_names[subject_code])
    
    sheet.append(headers)
    
    # 添加聚合数据
    for key, data in aggregated_data.items():
        student = data['student_obj']
        exam = data['exam_obj']
        class_obj = data['class_obj']
        scores = data['scores']
        
        row = [
            student.student_id,
            student.name,
            student.get_grade_level_display() if student.grade_level else "N/A",
            class_obj.class_name if class_obj else "N/A",
            exam.name,
            exam.academic_year or "N/A",
            exam.date.strftime('%Y-%m-%d')
        ]
        
        # 添加各科目成绩
        for subject_code in all_subjects:
            score_value = scores.get(subject_code)
            if score_value is not None:
                row.append(float(score_value))
            else:
                row.append("-")  # 没有成绩的科目显示为 "-"
        
        sheet.append(row)
    
    # 设置列宽
    base_widths = [15, 15, 10, 10, 20, 15, 12]  # 基础信息列宽
    subject_widths = [10] * len(all_subjects)    # 科目列宽
    column_widths = base_widths + subject_widths
    
    for i, width in enumerate(column_widths, 1):
        sheet.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width
    
    # 设置表头样式
    from openpyxl.styles import Font, PatternFill, Alignment
    
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    for col in range(1, len(headers) + 1):
        cell = sheet.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # 生成文件名
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    filename = f"聚合成绩导出_{timestamp}.xlsx"
    
    # 返回Excel文件
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    workbook.save(response)
    return response

# 删除选中的成绩记录
@require_POST
def score_batch_delete_selected(request):
    selected_records = request.POST.getlist('selected_records')
    
    if not selected_records:
        messages.error(request, '没有选择任何记录')
        return redirect('students_grades:score_list')
    
    # 解析选中的记录并删除
    total_deleted = 0
    affected_exam_ids = set()
    
    for record in selected_records:
        try:
            student_id, exam_id = record.split('_')
            deleted_count = Score.objects.filter(
                student_id=student_id,
                exam_id=exam_id
            ).delete()[0]
            total_deleted += deleted_count
            if deleted_count > 0:
                affected_exam_ids.add(int(exam_id))
        except ValueError:
            continue
    
    if total_deleted > 0:
        messages.success(request, f'成功删除 {total_deleted} 条成绩记录')
        
        # 异步更新受影响考试的排名
        from school_management.students_grades.tasks import update_all_rankings_async
        for exam_id in affected_exam_ids:
            try:
                update_all_rankings_async.delay(exam_id)
                messages.info(request, f'已提交考试 {exam_id} 的排名更新任务到后台处理')
            except Exception as e:
                messages.warning(request, f'排名更新任务提交失败: {str(e)}')
    else:
        messages.info(request, '没有找到对应的成绩记录')
    
    return redirect('students_grades:score_list')



# <!-- 成绩查询 -->
# 成绩查询主页面
def score_query(request):
    form = ScoreQueryForm()
    
    context = {
        'form': form,
        'page_title': '成绩查询',
    }
    
    return render(request, 'scores/score_query.html', context)

# 成绩查询结果页面
# 处理查询请求并显示结果，包含排名计算和科目排序
def score_query_results(request):
    form = ScoreQueryForm(request.GET)
    results = []
    total_count = 0
    
    if form.is_valid():
        # 获取查询参数
        student_name = form.cleaned_data.get('student_name')
        student_id = form.cleaned_data.get('student_id')
        grade_level = form.cleaned_data.get('grade_level')
        class_name = form.cleaned_data.get('class_name')
        exam = form.cleaned_data.get('exam')
        academic_year = form.cleaned_data.get('academic_year')
        subjects = form.cleaned_data.get('subject')  # 现在是列表
        date_from = form.cleaned_data.get('date_from')
        date_to = form.cleaned_data.get('date_to')
        sort_by = form.cleaned_data.get('sort_by')
        
        # 获取科目排序参数
        subject_sort = request.GET.get('subject_sort')
        sort_order = request.GET.get('sort_order', 'desc')  # 默认降序
        
        # 构建查询条件
        queryset = Score.objects.select_related(
            'student', 'exam', 'student__current_class'
        )
        
        # 应用筛选条件
        if student_name:
            queryset = queryset.filter(student__name__icontains=student_name)
        
        if student_id:
            queryset = queryset.filter(student__student_id__icontains=student_id)
        
        if grade_level:
            queryset = queryset.filter(student__grade_level=grade_level)
        
        if class_name:
            queryset = queryset.filter(student__current_class__class_name=class_name)
        
        if exam:
            queryset = queryset.filter(exam=exam)
        
        if academic_year:
            queryset = queryset.filter(exam__academic_year=academic_year)
        
        # 科目筛选 - 支持多选
        subjects = form.cleaned_data.get('subject')  # 现在是列表
        if subjects:
            queryset = queryset.filter(subject__in=subjects)
        
        if date_from:
            queryset = queryset.filter(exam__date__gte=date_from)
        
        if date_to:
            queryset = queryset.filter(exam__date__lte=date_to)
        
        # 聚合数据：按学生-考试组合
        results = calculate_scores_with_ranking(queryset, sort_by, subject_sort, sort_order)
        total_count = len(results)
        
        # 分页处理
        paginator = Paginator(results, 100)  # 每页100条记录
        page_number = request.GET.get('page')
        page_obj = paginator.get_page(page_number)
        
        context = {
            'form': form,
            'results': page_obj,
            'total_count': total_count,
            'page_title': '查询结果',
            'has_results': True,
        }
    else:
        context = {
            'form': form,
            'page_title': '成绩查询',
            'has_results': False,
        }
    
    return render(request, 'scores/score_query_results.html', context)

# 计算成绩和排名的辅助函数
# 计算成绩总分和年级排名，支持按科目、总分和排名排序
def calculate_scores_with_ranking(queryset, sort_by=None, subject_sort=None, sort_order='desc'):
    # 获取科目顺序
    def get_subject_order(subject_name):
        subject_dict = dict(SUBJECT_CHOICES)
        for i, (code, name) in enumerate(SUBJECT_CHOICES):
            if name == subject_name:
                return i
        return 999  # 未知科目排在最后
    
    # 聚合数据：按学生-考试组合
    student_exam_data = defaultdict(lambda: {
        'student': None,
        'exam': None,
        'class_obj': None,
        'scores': {},
        'total_score': 0,
        'subject_count': 0
    })
    
    # 收集所有科目
    all_subjects = set()
    
    for score in queryset:
        key = (score.student.pk, score.exam.pk)
        data = student_exam_data[key]
        
        # 设置基本信息
        if not data['student']:
            data['student'] = score.student
            data['exam'] = score.exam
            data['class_obj'] = score.student.current_class
        
        # 获取科目显示名称
        subject_name = dict(SUBJECT_CHOICES).get(score.subject, score.subject)
        data['scores'][subject_name] = score.score_value
        data['total_score'] += score.score_value or 0
        data['subject_count'] += 1
        all_subjects.add(subject_name)
    
    # 按SUBJECT_CHOICES顺序排列科目
    all_subjects = sorted(all_subjects, key=get_subject_order)
    
    # 转换为列表并使用数据库中的排名
    results = []
    for data in student_exam_data.values():
        data['all_subjects'] = all_subjects
        
        # 从数据库中获取排名（取该学生该考试任意一条成绩记录的排名）
        sample_score = queryset.filter(
            student=data['student'],
            exam=data['exam']
        ).first()
        
        if sample_score and sample_score.total_score_rank_in_grade:
            data['grade_rank'] = sample_score.total_score_rank_in_grade
        else:
            data['grade_rank'] = None
            
        results.append(data)
    
    # 应用排序
    if subject_sort:
        reverse_order = (sort_order == 'desc')
        
        if subject_sort == 'total_score':
            # 按总分排序
            results.sort(key=lambda x: x['total_score'], reverse=reverse_order)
        elif subject_sort == 'grade_rank':
            # 按年级排名排序
            def get_rank_value(result):
                rank = result.get('grade_rank')
                return rank if rank is not None else 999  # 没有排名的排在最后
            
            results.sort(key=get_rank_value, reverse=reverse_order)
        elif subject_sort in all_subjects:
            # 按指定科目排序
            def get_subject_score(result):
                score = result['scores'].get(subject_sort)
                return score if score is not None else -1  # 没有成绩的排在最后
            
            results.sort(key=get_subject_score, reverse=reverse_order)
    elif sort_by:
        # 应用其他排序
        if sort_by == 'total_score_desc':
            results.sort(key=lambda x: x['total_score'], reverse=True)
        elif sort_by == 'total_score_asc':
            results.sort(key=lambda x: x['total_score'], reverse=False)
        elif sort_by == 'student_name':
            results.sort(key=lambda x: x['student'].name)
        elif sort_by == 'exam_date':
            results.sort(key=lambda x: x['exam'].date, reverse=True)
        elif sort_by == 'grade_rank':
            results.sort(key=lambda x: x.get('grade_rank', 999))
    
    return results


# 成绩查询结果导出
def score_query_export(request):    
    form = ScoreQueryForm(request.GET)
    if not form.is_valid():
        messages.error(request, '查询参数无效，无法导出')
        return redirect('students_grades:score_query')
    
    # 获取查询参数（与score_query_results相同的逻辑）
    student_name = form.cleaned_data.get('student_name')
    student_id = form.cleaned_data.get('student_id')
    grade_level = form.cleaned_data.get('grade_level')
    class_name = form.cleaned_data.get('class_name')
    exam = form.cleaned_data.get('exam')
    academic_year = form.cleaned_data.get('academic_year')
    subject = form.cleaned_data.get('subject')
    date_from = form.cleaned_data.get('date_from')
    date_to = form.cleaned_data.get('date_to')
    sort_by = form.cleaned_data.get('sort_by')
    
    # 构建查询条件
    queryset = Score.objects.select_related(
        'student', 'exam', 'student__current_class'
    )
    
    # 应用筛选条件
    if student_name:
        queryset = queryset.filter(student__name__icontains=student_name)
    
    if student_id:
        queryset = queryset.filter(student__student_id__icontains=student_id)
    
    if grade_level:
        queryset = queryset.filter(student__grade_level=grade_level)
    
    if class_name:
        queryset = queryset.filter(student__current_class__class_name=class_name)
    
    if exam:
        queryset = queryset.filter(exam=exam)
    
    if academic_year:
            queryset = queryset.filter(exam__academic_year=academic_year)
        
    # 科目筛选
    if subject:
        queryset = queryset.filter(subject=subject)
    
    if date_from:
        queryset = queryset.filter(exam__date__gte=date_from)
    
    if date_to:
        queryset = queryset.filter(exam__date__lte=date_to)
    
    # 计算结果
    results = calculate_scores_with_ranking(queryset, sort_by)
    
    # 创建Excel文件
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "成绩查询结果"
    
    # 设置表头
    headers = [
        "学号", "学生姓名", "年级", "班级", "考试名称", "学年", "考试日期"
    ]
    
    # 添加科目列 - 按照SUBJECT_CHOICES顺序
    all_subjects = set()
    for result in results:
        # result 是字典，使用 get 方法访问 all_subjects
        subjects = result.get('all_subjects', [])
        all_subjects.update(subjects)
    
    # 按照SUBJECT_CHOICES的顺序排列科目
    def get_subject_order(subject):
        """获取科目在SUBJECT_CHOICES中的顺序"""
        for index, (choice_value, choice_display) in enumerate(SUBJECT_CHOICES):
            if choice_value == subject:
                return index
        return 999  # 如果科目不在SUBJECT_CHOICES中，放到最后
    
    ordered_subjects = sorted(all_subjects, key=get_subject_order)
    
    subject_names = dict(SUBJECT_CHOICES)
    for subject_code in ordered_subjects:
        headers.append(subject_names.get(subject_code, subject_code))
    
    # 添加总分和年级排名列
    headers.extend(["总分", "年级排名"])
    
    sheet.append(headers)
    
    # 添加数据行
    for result in results:
        # result 是字典，使用 get 方法访问各个字段
        student = result.get('student')
        class_obj = result.get('class_obj')
        exam = result.get('exam')
        scores = result.get('scores', {})
        
        row = [
            student.student_id if student else "",
            student.name if student else "",
            student.get_grade_level_display() if student and student.grade_level else "N/A",
            class_obj.class_name if class_obj else "N/A",
            exam.name if exam else "",
            exam.academic_year or "N/A" if exam else "N/A",
            exam.date.strftime('%Y-%m-%d') if exam and exam.date else ""
        ]
        
        # 添加各科目成绩 - 按照SUBJECT_CHOICES顺序
        for subject_code in ordered_subjects:
            score_value = scores.get(subject_code, "")
            row.append(score_value)
        
        # 添加总分和年级排名
        row.append(result.get('total_score', ""))
        row.append(result.get('grade_rank', "") or "")
        
        sheet.append(row)
    
    # 设置列宽
    for column in sheet.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 20)
        sheet.column_dimensions[column_letter].width = adjusted_width
    
    # 生成文件名
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    filename = f"成绩查询结果_{timestamp}.xlsx"
    
    # 返回Excel文件
    excel_file = io.BytesIO()
    workbook.save(excel_file)
    excel_file.seek(0)
    
    response = HttpResponse(
        excel_file.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


# AJAX端点：根据输入的关键字搜索学生
# 支持按姓名、学号、年级班级进行模糊搜索
@require_http_methods(["GET"])
def search_students_ajax(request):
    query = request.GET.get('q', '').strip()
    
    if not query:
        return JsonResponse({'students': []})
    
    # 构建搜索条件
    search_conditions = Q()
    
    # 按姓名搜索
    search_conditions |= Q(name__icontains=query)
    
    # 按学号搜索
    search_conditions |= Q(student_id__icontains=query)
    
    # 按年级班级搜索（通过关联的Class模型）
    search_conditions |= Q(current_class__grade_level__icontains=query)
    search_conditions |= Q(current_class__class_name__icontains=query)
    
    # 执行搜索，限制结果数量避免性能问题
    students = Student.objects.filter(search_conditions).select_related('current_class').order_by('name', 'student_id')[:20]
    
    # 构建返回数据
    student_list = []
    for student in students:
        class_info = ''
        if student.current_class:
            class_info = f"{student.current_class.grade_level}{student.current_class.class_name}"
        
        student_list.append({
            'id': student.id,
            'text': f"{student.name}-{student.student_id}-{class_info}",
            'name': student.name,
            'student_id': student.student_id,
            'class_info': class_info
        })
    
    return JsonResponse({'students': student_list})

# <!-- 成绩排名更新 -->
# 成绩排名更新视图
# 保持向后兼容的函数名
def update_grade_rankings(exam_id, grade_level=None):
    """向后兼容函数，调用tasks中的异步排名更新函数"""
    try:
        from ..tasks import update_all_rankings_async
        # 同步调用异步函数（用于向后兼容）
        return update_all_rankings_async(exam_id, grade_level)
    except ImportError:
        # 如果无法导入异步任务，返回错误信息
        return {
            'success': False,
            'message': '无法导入异步任务模块，请检查Redis和django-rq配置'
        }


# <!-- 成绩分析 -->
# 成绩分析主界面
def score_analysis(request):
    form = ScoreAnalysisForm()
    
    exams = Exam.objects.all().order_by('-academic_year', '-date', 'name')
    
    # 获取所有班级数据，按年级和班级名称排序
    # 使用自定义排序来正确处理班级编号（避免'10班'排在'2班'前面的问题）
    all_classes = Class.objects.all().order_by('grade_level')
    
    # 对班级进行自定义排序，提取班级编号进行数字排序
    def extract_class_number(class_obj):
        class_name = class_obj.class_name
        # 提取班级编号，例如从'1班'中提取'1'
        match = re.search(r'(\d+)', class_name)
        return int(match.group(1)) if match else 0
    
    all_classes = sorted(all_classes, key=lambda x: (x.grade_level, extract_class_number(x)))

    context = {
        'form': form,
        'page_title': '成绩分析',
        'exams' : exams,
        'all_classes': all_classes,
    }
    
    return render(request, 'scores/score_analysis.html', context)

# 班级/年级成绩分析
def score_analysis_class(request):
    form = ScoreAnalysisForm(request.GET or None)
    
    # 检查是否从score_analysis页面传递了参数
    from_analysis_page = bool(request.GET.get('academic_year') and 
                             request.GET.get('exam') and 
                             request.GET.get('grade_level'))
    
    # 获取可选班级列表（根据年级筛选）
    available_classes = []
    if request.GET.get('grade_level'):
        available_classes = Class.objects.filter(
            grade_level=request.GET.get('grade_level')
        ).annotate(
            student_count=Count('student')
        ).order_by('class_name')
    
    # 处理从score_analysis页面传递的班级选择
    url_selected_classes = request.GET.getlist('selected_classes')
    auto_analysis = False
    
    if from_analysis_page and url_selected_classes:
        # 如果有班级选择，直接进行分析
        auto_analysis = True
        
        # 直接从 URL 参数获取数据，不使用表单验证
        academic_year = request.GET.get('academic_year')
        exam_id = request.GET.get('exam')
        grade_level = request.GET.get('grade_level')
        
        try:
            exam = Exam.objects.get(id=exam_id)
        except Exam.DoesNotExist:
            messages.error(request, f'考试不存在')
            return render(request, 'scores/score_analysis_class.html', {
                'form': form,
                'available_classes': available_classes,
                'page_title': '班级成绩分析',
                'analysis_type': 'class',
                'from_analysis_page': from_analysis_page,
                'show_class_selection': from_analysis_page and available_classes,
                'auto_analysis': auto_analysis
            })
        
        # 根据传递的班级选择确定分析模式
        if 'all' in url_selected_classes:
            # 场景3：所有班级 - 年级整体分析
            analysis_mode = 'grade_overall'
            scores = Score.objects.filter(
                exam=exam,
                student__current_class__grade_level=grade_level
            )
            
            # 计算总满分
            total_max_score = 0
            exam_subjects = exam.exam_subjects.all()
            for exam_subject in exam_subjects:
                total_max_score += exam_subject.max_score
            
            context = {
                'form': form,
                'available_classes': available_classes,
                'page_title': '年级整体成绩分析',
                'analysis_type': 'class',
                'from_analysis_page': from_analysis_page,
                'show_class_selection': False,
                'auto_analysis': True,
                'analysis_mode': analysis_mode,
                'selected_exam': exam,
                'selected_grade': grade_level,
                'total_students': scores.values('student').distinct().count(),
                'total_scores': scores.count(),
                'total_max_score': total_max_score,
            }
            return render(request, 'scores/score_analysis_class.html', context)
            
        elif len(url_selected_classes) == 1:
            # 场景1：单个班级分析
            class_id = url_selected_classes[0]
            try:
                target_class = Class.objects.get(id=class_id)
                
                scores = Score.objects.filter(
                    exam=exam,
                    student__current_class=target_class
                )
                
                analysis_result = _analyze_single_class(scores, target_class, exam)
                
                context = {
                    'form': form,
                    'available_classes': available_classes,
                    'page_title': f'{target_class.class_name} 成绩分析',
                    'analysis_type': 'class',
                    'from_analysis_page': from_analysis_page,
                    'show_class_selection': False,
                    'auto_analysis': True,
                    'analysis_mode': 'single_class',
                    'selected_exam': exam,
                    'selected_grade': grade_level,
                    'target_class': target_class,
                    **analysis_result
                }
                return render(request, 'scores/score_analysis_class.html', context)
                
            except Class.DoesNotExist:
                messages.error(request, f'班级 {class_id} 不存在')
                
        else:
            # 场景2：多班级对比分析
            analysis_mode = 'class_comparison'
            selected_class_ids = [int(class_id) for class_id in url_selected_classes]
            selected_classes = Class.objects.filter(id__in=selected_class_ids)
            
            # 获取多班级对比分析数据
            analysis_result = _analyze_multiple_classes(selected_classes, exam)
            
            context = {
                'form': form,
                'available_classes': available_classes,
                'page_title': '多班级成绩对比分析',
                'analysis_type': 'class',
                'from_analysis_page': from_analysis_page,
                'show_class_selection': False,
                'auto_analysis': True,
                'analysis_mode': analysis_mode,
                'selected_exam': exam,
                'selected_grade': grade_level,
                'selected_classes': selected_classes,
                **analysis_result
            }
            return render(request, 'scores/score_analysis_class_multi.html', context)
    
    context = {
        'form': form,
        'available_classes': available_classes,
        'page_title': '班级成绩分析',
        'analysis_type': 'class',
        'from_analysis_page': from_analysis_page,
        'show_class_selection': from_analysis_page and available_classes,
        'auto_analysis': auto_analysis
    }
    
    if form.is_valid():
        academic_year = form.cleaned_data['academic_year']
        exam = form.cleaned_data['exam']
        grade_level = form.cleaned_data['grade_level']
        
        # 判断分析模式
        all_classes_selected = request.GET.get('class_selection') == 'all'
        selected_classes = form.cleaned_data.get('selected_classes', [])
        
        # 如果selected_classes为空，尝试从class_name字段获取（向后兼容）
        if not selected_classes and form.cleaned_data.get('class_name'):
            try:
                class_obj = Class.objects.get(
                    grade_level=grade_level,
                    class_name=form.cleaned_data['class_name']
                )
                selected_classes = [class_obj]
            except Class.DoesNotExist:
                selected_classes = []
        
        if all_classes_selected or not selected_classes:
            # 年级整体分析
            analysis_mode = 'grade_overall'
            analysis_scope = f"年级整体分析：{grade_level}"
            filters = {
                'exam': exam,
                'student__current_class__grade_level': grade_level,
            }
        elif len(selected_classes) == 1:
            # 单班级详细分析
            analysis_mode = 'single_class'
            target_class = selected_classes[0]
            analysis_scope = f"单班级详细分析：{target_class.grade_level}{target_class.class_name}"
            filters = {
                'exam': exam,
                'student__current_class': target_class,
            }
        else:
            # 多班级对比分析
            analysis_mode = 'class_comparison'
            class_names = [f"{c.grade_level}{c.class_name}" for c in selected_classes]
            analysis_scope = f"班级对比分析：{', '.join(class_names)}"
            filters = {
                'exam': exam,
                'student__current_class__in': selected_classes,
            }
        
        # 获取成绩数据并进行相应分析
        scores = Score.objects.filter(**filters).select_related(
            'student', 'exam', 'student__current_class'
        )
        
        # 单班级详细分析的数据处理
        if analysis_mode == 'single_class':
            analysis_data = _analyze_single_class(scores, target_class, exam)
            context.update(analysis_data)
        elif analysis_mode == 'class_comparison':
            # 多班级对比分析的数据处理
            analysis_data = _analyze_multiple_classes(selected_classes, exam)
            context.update(analysis_data)
        elif analysis_mode == 'grade_overall':
            # 年级整体分析需要计算总满分
            total_max_score = 0
            exam_subjects = exam.exam_subjects.all()
            for exam_subject in exam_subjects:
                total_max_score += exam_subject.max_score
            context['total_max_score'] = total_max_score
        
        context.update({
            'scores': scores,
            'analysis_mode': analysis_mode,
            'analysis_scope': analysis_scope,
            'selected_exam': exam,
            'selected_grade': grade_level,
            'selected_classes': selected_classes,
            'academic_year': academic_year,
        })
    
    # 根据分析模式选择不同的模板
    if context.get('analysis_mode') == 'class_comparison':
        return render(request, 'scores/score_analysis_class_multi.html', context)
    else:
        return render(request, 'scores/score_analysis_class.html', context)


# 学生个人成绩分析
def score_analysis_student(request):
    """学生个人成绩分析页面"""
    form = ScoreAnalysisForm(request.GET or None)

    context = {
        'form': form,
        'page_title': '个人成绩分析',
        'analysis_type': 'student'
    }
    
    # 获取URL参数中的筛选条件
    academic_year = request.GET.get('academic_year')
    exam_id = request.GET.get('exam')
    grade_level = request.GET.get('grade_level')
    class_id = request.GET.get('class_name')  # 注意：这里传递的实际是班级ID
    
    # 年级映射：从数字/英文转换为中文
    grade_mapping = {
        '10': '高一',
        '11': '高二', 
        '12': '高三',
        '7': '初一',
        '8': '初二',
        '9': '初三',
        'Grade10': '高一',
        'Grade11': '高二',
        'Grade12': '高三',
        'Grade7': '初一',
        'Grade8': '初二',
        'Grade9': '初三',
    }
    
    # 转换年级格式
    if grade_level and grade_level in grade_mapping:
        grade_level = grade_mapping[grade_level]
    
    # 获取班级对象和班级名称
    selected_class = None
    class_name = None
    if class_id:
        try:
            selected_class = Class.objects.get(id=class_id)
            class_name = selected_class.class_name
        except Class.DoesNotExist:
            pass
    
    # 添加调试信息到上下文
    context['debug_info'] = {
        'academic_year': academic_year,
        'exam_id': exam_id,
        'grade_level': grade_level,
        'class_id': class_id,
        'class_name': class_name,
        'selected_class': selected_class,
        'all_params': dict(request.GET.items())
    }
    
    # 如果有筛选条件，获取对应的学生列表
    if grade_level:
        # 构建查询条件
        filters = {'grade_level': grade_level}
        
        if selected_class:
            filters['current_class'] = selected_class
        
        # 获取学生列表，按学号排序
        students = Student.objects.filter(**filters).select_related('current_class').order_by('student_id')
        
        # 获取考试列表（根据年级和学年筛选）
        exam_filters = {'grade_level': grade_level}
        if academic_year:
            exam_filters['academic_year'] = academic_year
        
        exams = Exam.objects.filter(**exam_filters).order_by('-date')
        
        context.update({
            'students': students,
            'exams': exams,
            'academic_year': academic_year,
            'selected_grade': grade_level,
            'selected_class': f"{selected_class.grade_level}{selected_class.class_name}" if selected_class else None,
            'student_count': students.count(),
        })
        
        # 如果指定了考试ID，获取考试对象
        if exam_id:
            try:
                exam = Exam.objects.get(pk=exam_id)
                context['selected_exam'] = exam
            except Exam.DoesNotExist:
                pass
    else:
        # 如果没有筛选条件，获取所有学生和考试
        students = Student.objects.all().select_related('current_class').order_by('grade_level', 'student_id')
        exams = Exam.objects.all().order_by('-date')
        
        context.update({
            'students': students,
            'exams': exams,
            'student_count': students.count(),
        })
    
    return render(request, 'scores/score_analysis_student.html', context)

# AJAX接口：根据年级获取班级列表
def get_classes_by_grade(request):
    """根据年级获取班级列表的AJAX接口"""
    grade_level = request.GET.get('grade_level')
    
    if not grade_level:
        return JsonResponse({'error': '年级参数不能为空'}, status=400)
    
    try:
        classes = Class.objects.filter(grade_level=grade_level)
        
        # 自定义排序函数：提取班级编号进行数字排序
        def extract_class_number(class_obj):
            """从班级名称中提取数字进行排序，例如从'1班'中提取'1'"""
            import re
            class_name = class_obj.class_name
            # 提取班级编号，例如从'1班'中提取'1'
            match = re.search(r'(\d+)', class_name)
            return int(match.group(1)) if match else 0
        
        # 按班级编号数字排序
        sorted_classes = sorted(classes, key=extract_class_number)
        
        classes_data = [
            {
                'id': cls.id,
                'class_name': cls.class_name,
                'grade_level': cls.grade_level,
                'display_name': f"{cls.grade_level}{cls.class_name}"
            }
            for cls in sorted_classes
        ]
        
        return JsonResponse({
            'classes': classes_data,
            'count': len(classes_data)
        })
    except Exception as e:
        return JsonResponse({'error': f'获取班级列表失败: {str(e)}'}, status=500)

# AJAX接口：根据班级获取学生列表
def get_students_by_class(request):
    """根据年级和班级获取学生列表的AJAX接口"""
    grade_level = request.GET.get('grade_level')
    class_name = request.GET.get('class_name')
    class_id = request.GET.get('class_id')
    
    if not grade_level:
        return JsonResponse({'error': '年级参数不能为空'}, status=400)
    
    try:
        # 构建查询条件
        filters = {'grade_level': grade_level}
        
        # 支持按班级ID或班级名称查询
        if class_id:
            filters['current_class_id'] = class_id
        elif class_name:
            filters['current_class__class_name'] = class_name
        
        students = Student.objects.filter(**filters).select_related('current_class').order_by('student_id', 'name')
        students_data = [
            {
                'id': student.id,
                'student_id': student.student_id,
                'name': student.name,
                'class_name': student.current_class.class_name if student.current_class else '未分班',
                'grade_level': student.grade_level
            }
            for student in students
        ]
        
        return JsonResponse({
            'students': students_data,
            'count': len(students_data)
        })
    except Exception as e:
        return JsonResponse({'error': f'获取学生列表失败: {str(e)}'}, status=500)

# 年级成绩分析
def score_analysis_grade(request):
    form = ScoreAnalysisForm(request.GET or None)
    
    # 检查是否从score_analysis页面传递了参数
    from_analysis_page = bool(request.GET.get('academic_year') and 
                             request.GET.get('exam') and 
                             request.GET.get('grade_level'))
    
    context = {
        'form': form,
        'page_title': '年级成绩分析',
        'analysis_type': 'grade',
        'from_analysis_page': from_analysis_page,
        'total_max_score': 0,  # 添加默认值，避免模板错误
    }
    
    if from_analysis_page:
        # 直接从 URL 参数获取数据，不使用表单验证
        academic_year = request.GET.get('academic_year')
        exam_id = request.GET.get('exam')
        grade_level = request.GET.get('grade_level')
        
        try:
            exam = Exam.objects.get(id=exam_id)
            
            # 获取年级整体分析数据
            analysis_result = _analyze_grade(exam, grade_level)
            
            context.update({
                'selected_exam': exam,
                'selected_grade': grade_level,
                'academic_year': academic_year,
                **analysis_result
            })
        except Exam.DoesNotExist:
            messages.error(request, f'考试不存在')
    elif form.is_valid():
        # 处理表单提交的情况
        academic_year = form.cleaned_data['academic_year']
        exam = form.cleaned_data['exam']
        grade_level = form.cleaned_data['grade_level']
        
        try:
            # 获取年级整体分析数据
            analysis_result = _analyze_grade(exam, grade_level)
            
            context.update({
                'selected_exam': exam,
                'selected_grade': grade_level,
                'academic_year': academic_year,
                **analysis_result
            })
        except Exception as e:
            messages.error(request, f'数据分析失败: {str(e)}')
    
    return render(request, 'scores/score_analysis_grade.html', context)

# 单班级详细数据分析辅助函数
def _analyze_single_class(scores, target_class, exam):
    # 基础统计信息
    total_students = scores.values('student').distinct().count()
    
    # 按科目统计成绩
    subject_stats = {}
    score_distribution = {}
    student_scores = {}
    
    for subject_code, subject_name in SUBJECT_CHOICES:
        subject_scores = scores.filter(subject=subject_code)
        if subject_scores.exists():
            stats = subject_scores.aggregate(
                avg_score=Avg('score_value'),
                max_score=Max('score_value'),
                min_score=Min('score_value'),
                count=Count('score_value')
            )
            subject_stats[subject_code] = {
                'name': subject_name,
                'avg_score': round(float(stats['avg_score'] or 0), 2),
                'actual_max_score': float(stats['max_score'] or 0),  # 该班级实际最高分
                'actual_min_score': float(stats['min_score'] or 0),  # 该班级实际最低分
                'count': stats['count']
            }
            
            # 获取该科目的满分
            subject_max_score = 100  # 默认满分
            try:
                exam_subject = exam.exam_subjects.filter(subject=subject_code).first()
                if exam_subject:
                    subject_max_score = exam_subject.max_score
                else:
                    # 如果ExamSubject不存在，从SUBJECT_DEFAULT_MAX_SCORES中获取默认满分
                    grade_config = SUBJECT_DEFAULT_MAX_SCORES.get(exam.grade_level, {})
                    subject_max_score = grade_config.get(subject_code, 100)
            except:
                # 异常情况下，尝试从默认配置获取
                grade_config = SUBJECT_DEFAULT_MAX_SCORES.get(exam.grade_level, {})
                subject_max_score = grade_config.get(subject_code, 100)
            
            # 将考试设定的满分添加到subject_stats中（用于等级分布计算）
            subject_stats[subject_code]['exam_max_score'] = subject_max_score
            
            # 成绩分布统计（按等级百分比）
            grade_ranges = {
                '特优(95%+)': subject_scores.filter(score_value__gte=subject_max_score * 0.95).count(),
                '优秀(85%-95%)': subject_scores.filter(score_value__gte=subject_max_score * 0.85, score_value__lt=subject_max_score * 0.95).count(),
                '良好(70%-85%)': subject_scores.filter(score_value__gte=subject_max_score * 0.70, score_value__lt=subject_max_score * 0.85).count(),
                '及格(60%-70%)': subject_scores.filter(score_value__gte=subject_max_score * 0.60, score_value__lt=subject_max_score * 0.70).count(),
                '不及格(<60%)': subject_scores.filter(score_value__lt=subject_max_score * 0.60).count(),
            }
            score_distribution[subject_code] = grade_ranges
    
    # 学生总分排名 - 使用Django ORM聚合查询避免重复
    student_total_scores = []
    
    # 使用values和annotate来按学生分组并计算总分
    student_scores = scores.values('student__id', 'student__name').annotate(
        total_score=Sum('score_value'),
        subject_count=Count('subject')
    ).order_by('-total_score')
    
    # 转换为列表格式并添加排名和年级排名
    for i, student_data in enumerate(student_scores):
        student_id = student_data['student__id']
        
        # 获取该学生的年级排名（从数据库中的total_score_rank_in_grade字段获取）
        grade_rank = None
        sample_score = scores.filter(student_id=student_id).first()
        if sample_score:
            grade_rank = sample_score.total_score_rank_in_grade
        

        
        student_total_scores.append({
            'student_id': student_id,
            'student_name': student_data['student__name'],
            'total_score': float(student_data['total_score']),
            'subject_count': student_data['subject_count'],
            'rank': i + 1,
            'grade_rank': grade_rank
        })
    
    # 班级总体统计
    if student_total_scores:
        class_avg_total = sum([s['total_score'] for s in student_total_scores]) / len(student_total_scores)
        class_max_total = max([s['total_score'] for s in student_total_scores])
        class_min_total = min([s['total_score'] for s in student_total_scores])
    else:
        class_avg_total = class_max_total = class_min_total = 0
    
    # 计算总分满分
    total_max_score = 0
    exam_subjects = exam.exam_subjects.all()
    for exam_subject in exam_subjects:
        total_max_score += exam_subject.max_score
    
    # 计算等级分布（基于总分百分比）
    grade_distribution = {
        '特优(95%+)': 0,
        '优秀(85%-95%)': 0,
        '良好(70%-85%)': 0,
        '及格(60%-70%)': 0,
        '不及格(<60%)': 0
    }
    
    # 统计各等级人数
    for student_data in student_total_scores:
        if total_max_score > 0:
            percentage = student_data['total_score'] / total_max_score
            if percentage >= 0.95:
                grade_distribution['特优(95%+)'] += 1
            elif percentage >= 0.85:
                grade_distribution['优秀(85%-95%)'] += 1
            elif percentage >= 0.70:
                grade_distribution['良好(70%-85%)'] += 1
            elif percentage >= 0.60:
                grade_distribution['及格(60%-70%)'] += 1
            else:
                grade_distribution['不及格(<60%)'] += 1
    
    # 准备图表数据
    chart_data = {
        'subject_avg_scores': {
            'labels': [subject_stats[code]['name'] for code in subject_stats.keys()],
            'data': [float(subject_stats[code]['avg_score']) for code in subject_stats.keys()]
        },
        'subject_max_scores': [subject_stats[code]['actual_max_score'] for code in subject_stats.keys()],  # 添加各科目实际最高分
        'score_distribution': score_distribution,
        'student_total_scores': [{
            'student_id': s['student_id'],
            'student_name': s['student_name'],
            'total_score': float(s['total_score']),
            'subject_count': s['subject_count'],
            'rank': s['rank'],
            'grade_rank': s['grade_rank']
        } for s in student_total_scores[:10]],  # 只显示前10名
        'total_max_score': total_max_score,
        'grade_distribution': grade_distribution
    }
    
    return {
        'total_students': total_students,
        'subject_stats': subject_stats,
        'score_distribution': score_distribution,
        'student_rankings': student_total_scores,
        'class_avg_total': round(class_avg_total, 2),
        'class_max_total': float(class_max_total) if class_max_total else 0,
        'class_min_total': float(class_min_total) if class_min_total else 0,
        'chart_data_json': json.dumps(chart_data, ensure_ascii=False),
        'target_class': target_class
    }

# 多班级对比分析辅助函数
def _analyze_multiple_classes(selected_classes, exam):
    # 获取所有选中班级的成绩数据
    all_scores = Score.objects.filter(
        exam=exam,
        student__current_class__in=selected_classes
    ).select_related('student', 'student__current_class')
    
    # 获取考试科目
    exam_subjects = ExamSubject.objects.filter(exam=exam)
    subjects = [es.subject_code for es in exam_subjects]
    
    # 辅助函数：将Decimal转换为float
    def decimal_to_float(value):
        if isinstance(value, Decimal):
            return float(value)
        return value
    
    # 按班级分组统计数据
    class_statistics = []
    class_subject_averages = {}
    score_distributions = {}
    total_students = 0
    highest_avg = 0
    
    # 首先对班级进行数字排序
    def extract_class_number(class_obj):
        """提取班级编号进行数字排序"""
        class_name = class_obj.class_name
        match = re.search(r'(\d+)', class_name)
        return int(match.group(1)) if match else 0
    
    # 对传入的班级按数字顺序排序
    sorted_classes = sorted(selected_classes, key=extract_class_number)
    
    for class_obj in sorted_classes:
        class_scores = all_scores.filter(student__current_class=class_obj)
        
        # 基本统计信息
        student_count = class_scores.values('student').distinct().count()
        total_students += student_count
        
        # 计算各科目平均分
        subject_averages = []
        class_name = class_obj.class_name  # 只显示班级名称，不包含年级
        class_subject_averages[class_name] = []
        
        for subject in subjects:
            subject_scores = class_scores.filter(subject=subject)
            scores_list = [decimal_to_float(score.score_value) for score in subject_scores if score.score_value is not None]
            
            if scores_list:
                avg_score = statistics.mean(scores_list)
                subject_averages.append(round(avg_score, 2))
                class_subject_averages[class_name].append(round(avg_score, 2))
            else:
                subject_averages.append(0)
                class_subject_averages[class_name].append(0)
        
        # 计算总分统计
        student_totals = {}
        for score in class_scores:
            if score.student.id not in student_totals:
                student_totals[score.student.id] = 0
            if score.score_value is not None:
                student_totals[score.student.id] += decimal_to_float(score.score_value)
        
        total_scores_list = list(student_totals.values()) if student_totals else [0]
        avg_total = statistics.mean(total_scores_list) if total_scores_list else 0
        max_total = max(total_scores_list) if total_scores_list else 0
        min_total = min(total_scores_list) if total_scores_list else 0
        
        # 确保所有数值都是float类型
        avg_total = round(float(avg_total), 2)
        max_total = float(max_total)
        min_total = float(min_total)
        
        if avg_total > highest_avg:
            highest_avg = avg_total
        
        # 等级分布统计（基于百分比）
        score_dist = [0, 0, 0, 0, 0]  # 特优(95%+), 优秀(85%-95%), 良好(70%-85%), 及格(60%-70%), 不及格(<60%)
        
        # 计算满分
        exam_subjects = exam.exam_subjects.all()
        total_max_score = sum(es.max_score for es in exam_subjects)
        
        for total in total_scores_list:
            if total_max_score > 0:
                percentage = (total / total_max_score) * 100
                if percentage >= 95:
                    score_dist[0] += 1  # 特优(95%+)
                elif percentage >= 85:
                    score_dist[1] += 1  # 优秀(85%-95%)
                elif percentage >= 70:
                    score_dist[2] += 1  # 良好(70%-85%)
                elif percentage >= 60:
                    score_dist[3] += 1  # 及格(60%-70%)
                else:
                    score_dist[4] += 1  # 不及格(<60%)
        
        score_distributions[class_name] = score_dist
        
        class_statistics.append({
            'class_name': class_name,
            'student_count': student_count,
            'avg_total': decimal_to_float(avg_total),
            'max_total': decimal_to_float(max_total),
            'min_total': decimal_to_float(min_total),
            'subject_averages': subject_averages
        })
    
    # 准备图表数据 - 确保班级顺序正确
    sorted_class_names = [class_obj.class_name for class_obj in sorted_classes]
    chart_data = {
        'subjects': subjects,
        'classes': sorted_class_names,
        'class_subject_averages': class_subject_averages,
        'score_distributions': score_distributions
    }
    
    # 获取科目名称映射
    subject_names = []
    for subject_code in subjects:
        exam_subject = exam_subjects.filter(subject_code=subject_code).first()
        if exam_subject:
            subject_names.append(exam_subject.subject_name)
        else:
            # 从SUBJECT_CHOICES中获取名称
            for code, name in SUBJECT_CHOICES:
                if code == subject_code:
                    subject_names.append(name)
                    break
            else:
                subject_names.append(subject_code)
    
    return {
        'class_statistics': class_statistics,
        'subjects': [{'code': code, 'name': name} for code, name in zip(subjects, subject_names)],
        'total_students': total_students,
        'subject_count': len(subjects),
        'highest_avg': float(highest_avg),
        'chart_data_json': json.dumps(chart_data),
        'academic_year': exam.academic_year
    }

# 年级成绩分析辅助函数
def _analyze_grade(exam, grade_level):
    # 辅助函数：将Decimal转换为float
    def decimal_to_float(value):
        if value is None:
            return 0.0
        if isinstance(value, Decimal):
            return float(value)
        return value
    
    # 获取该年级的所有班级
    # 使用自定义排序，按班级名称中的数字排序
    classes = Class.objects.filter(grade_level=grade_level)
    classes = sorted(classes, key=lambda x: int(''.join(filter(str.isdigit, x.class_name))) if any(c.isdigit() for c in x.class_name) else 999)
    
    # 获取该年级该考试的所有成绩
    all_scores = Score.objects.filter(
        exam=exam,
        student__current_class__grade_level=grade_level
    ).select_related('student', 'student__current_class')
    
    # 获取考试科目
    exam_subjects = ExamSubject.objects.filter(exam=exam)
    subjects = [{'code': es.subject_code, 'name': dict(SUBJECT_CHOICES).get(es.subject_code, es.subject_code), 'max_score': es.max_score} for es in exam_subjects]
    
    # 基本统计
    total_students = all_scores.values('student').distinct().count()
    total_classes = len(classes)
    
    # 按班级分组统计
    class_statistics = []
    class_names = []
    class_averages = []
    class_grade_distribution = {}
    
    # 科目统计
    subject_stats = {}
    for subject in subjects:
        subject_code = subject['code']
        subject_scores = all_scores.filter(subject=subject_code)
        if subject_scores.exists():
            avg_score = subject_scores.aggregate(avg=Avg('score_value'))['avg']
            subject_stats[subject_code] = {
                'name': subject['name'],
                'avg_score': decimal_to_float(avg_score),
                'max_score': subject['max_score']
            }
    
    # 计算总分和年级平均分
    student_totals = {}
    for score in all_scores:
        if score.student.id not in student_totals:
            student_totals[score.student.id] = 0
        if score.score_value is not None:
            student_totals[score.student.id] += decimal_to_float(score.score_value)
    
    total_scores_list = list(student_totals.values()) if student_totals else [0]
    grade_avg_score = statistics.mean(total_scores_list) if total_scores_list else 0
    
    # 计算总分满分
    total_max_score = sum([subject['max_score'] for subject in subjects])
    
    # 计算年级优秀率（95%以上）
    # 如果 total_max_score 为 0（没有 ExamSubject），应返回 0 而不是将所有分数视作优秀
    if total_max_score > 0:
        excellent_count = sum(1 for score in total_scores_list if score >= total_max_score * 0.95)
        excellent_rate = (excellent_count / len(total_scores_list) * 100) if total_scores_list else 0
    else:
        excellent_count = 0
        excellent_rate = 0
    
    # 各班级详细统计
    for class_obj in classes:
        class_scores = all_scores.filter(student__current_class=class_obj)
        
        if not class_scores.exists():
            continue
            
        class_name = class_obj.class_name  # 只显示班级名称，不包含年级
        class_names.append(class_name)
        
        # 班级学生总分统计
        class_student_totals = {}
        for score in class_scores:
            if score.student.id not in class_student_totals:
                class_student_totals[score.student.id] = 0
            if score.score_value is not None:
                class_student_totals[score.student.id] += decimal_to_float(score.score_value)
        
        class_total_scores = list(class_student_totals.values()) if class_student_totals else [0]
        avg_total = statistics.mean(class_total_scores) if class_total_scores else 0
        max_total = max(class_total_scores) if class_total_scores else 0
        min_total = min(class_total_scores) if class_total_scores else 0
        
        class_averages.append(decimal_to_float(avg_total))
        
        # 计算各等级人数和比例（按照新的5级划分）
        student_count = len(class_total_scores)
        if total_max_score > 0:
            class_excellent_plus_count = sum(1 for score in class_total_scores if score >= total_max_score * 0.95)  # 特优(95%+)
            class_excellent_count = sum(1 for score in class_total_scores if total_max_score * 0.85 <= score < total_max_score * 0.95)  # 优秀(85%-95%)
            class_good_count = sum(1 for score in class_total_scores if total_max_score * 0.70 <= score < total_max_score * 0.85)  # 良好(70%-85%)
            class_pass_count = sum(1 for score in class_total_scores if total_max_score * 0.60 <= score < total_max_score * 0.70)  # 及格(60%-70%)
            class_fail_count = sum(1 for score in class_total_scores if score < total_max_score * 0.60)  # 不及格(<60%)

            class_excellent_rate = ((class_excellent_plus_count + class_excellent_count) / student_count * 100) if student_count > 0 else 0  # 优秀率（包含特优+优秀）
            class_good_rate = ((class_good_count + class_excellent_count + class_excellent_plus_count) / student_count * 100) if student_count > 0 else 0  # 良好率
            class_pass_rate = ((class_pass_count + class_good_count + class_excellent_count + class_excellent_plus_count) / student_count * 100) if student_count > 0 else 0  # 及格率
        else:
            # 没有满分定义时，各等级计数和比率应为 0（避免把 0 视为满分导致所有学生被计为优秀）
            class_excellent_plus_count = class_excellent_count = class_good_count = class_pass_count = class_fail_count = 0
            class_excellent_rate = class_good_rate = class_pass_rate = 0

        # 各科目平均分
        subject_averages = []
        for subject in subjects:
            subject_code = subject['code']
            class_subject_scores = class_scores.filter(subject=subject_code)
            if class_subject_scores.exists():
                avg = class_subject_scores.aggregate(avg=Avg('score_value'))['avg']
                subject_averages.append(decimal_to_float(avg))
            else:
                subject_averages.append(0)

        class_statistics.append({
            'class_name': class_name,
            'student_count': student_count,
            'avg_total': decimal_to_float(avg_total),
            'max_total': decimal_to_float(max_total),
            'min_total': decimal_to_float(min_total),
            'excellent_rate': class_excellent_rate,
            'good_rate': class_good_rate,
            'pass_rate': class_pass_rate,
            'subject_averages': subject_averages
        })

        # 班级等级分布（用于堆叠柱状图）- 5个等级（从底部到顶部：不及格→及格→良好→优秀→特优）
        class_grade_distribution[class_name] = [class_fail_count, class_pass_count, class_good_count, class_excellent_count, class_excellent_plus_count]
    
    # 年级成绩分布（总分分段统计）
    score_ranges = ['特优(95%+)', '优秀(85%-95%)', '良好(70%-85%)', '及格(60%-70%)', '不及格(<60%)']
    score_distribution = [0, 0, 0, 0, 0]
    
    if total_max_score > 0:
        for total_score in total_scores_list:
            if total_score >= total_max_score * 0.95:
                score_distribution[0] += 1
            elif total_score >= total_max_score * 0.85:
                score_distribution[1] += 1
            elif total_score >= total_max_score * 0.70:
                score_distribution[2] += 1
            elif total_score >= total_max_score * 0.60:
                score_distribution[3] += 1
            else:
                score_distribution[4] += 1
    
    # 科目难度系数计算（平均分/满分）
    difficulty_coefficients = []
    subject_names = []
    subject_averages = []
    subject_max_scores = []
    
    for subject in subjects:
        subject_code = subject['code']
        if subject_code in subject_stats:
            subject_names.append(subject_stats[subject_code]['name'])
            subject_averages.append(subject_stats[subject_code]['avg_score'])
            subject_max_scores.append(subject['max_score'])
            difficulty_coefficient = subject_stats[subject_code]['avg_score'] / subject['max_score']
            difficulty_coefficients.append(difficulty_coefficient)
    
    # 准备图表数据
    chart_data = {
        'class_names': class_names,
        'class_averages': class_averages,
        'subject_names': subject_names,
        'subject_averages': subject_averages,
        'subject_max_scores': subject_max_scores,
        'score_ranges': score_ranges,
        'score_distribution': score_distribution,
        'class_grade_distribution': class_grade_distribution,
        'difficulty_coefficients': difficulty_coefficients,
        'total_max_score': total_max_score,
        'total_scores': total_scores_list,  # 添加所有学生的总分数据
    }
    
    return {
        'total_students': total_students,
        'total_classes': total_classes,
        'grade_avg_score': decimal_to_float(grade_avg_score),
        'excellent_rate': excellent_rate,
        'class_statistics': class_statistics,
        'subjects': subjects,
        'total_max_score': total_max_score,
        'chart_data_json': json.dumps(chart_data, ensure_ascii=False)
    }


# 学生个人成绩分析AJAX接口
def get_student_analysis_data(request):
    """获取学生个人成绩分析数据的AJAX接口"""
    if request.method != 'GET':
        return JsonResponse({'error': '不支持的请求方法'}, status=405)
    
    student_id = request.GET.get('student_id')
    exam_ids = request.GET.get('exam_ids', '')  # 支持多考试ID，用逗号分隔
    exam_id = request.GET.get('exam_id')  # 兼容单考试ID
    
    if not student_id:
        return JsonResponse({'error': '缺少学生ID参数'}, status=400)
        
    # 处理考试ID参数
    if exam_ids:
        exam_id_list = [id.strip() for id in exam_ids.split(',') if id.strip()]
    elif exam_id:
        exam_id_list = [exam_id]
    else:
        return JsonResponse({'error': '缺少考试ID参数'}, status=400)
    
    try:
        student = Student.objects.get(id=student_id)
        # 按考试日期升序排列，确保趋势图横坐标按时间顺序（从早到晚）展示
        # 使用 (date, id) 做稳定排序，避免同一天考试顺序不确定
        exams = Exam.objects.filter(id__in=exam_id_list).order_by('date', 'id')
        
        if not exams.exists():
            return JsonResponse({'error': '未找到指定的考试'}, status=404)
        
        # 构建多考试分析数据
        analysis_data = {
            'student_info': {
                'id': student.id,
                'student_id': student.student_id,
                'name': student.name,
                'grade_level': student.grade_level,
                'class_name': student.current_class.class_name if student.current_class else '未分班',
            },
            'exams': [],
            'subjects': [],
            'trend_data': {},
            'summary': {
                'total_exams': exams.count(),
                'subjects_count': 0,
            }
        }
        
        # 收集所有科目
        all_subjects = set()
        
        # 处理每个考试的数据
        for exam in exams:            
            # 获取该学生在该次考试的所有成绩
            scores = Score.objects.filter(
                student=student,
                exam=exam
            )
            
            # 按SUBJECT_CHOICES顺序排序
            def get_subject_order_for_score(score):
                """获取成绩对象的科目在SUBJECT_CHOICES中的顺序"""
                for index, (subject_code, subject_name) in enumerate(SUBJECT_CHOICES):
                    if subject_code == score.subject or subject_name == score.subject:
                        return index
                return 999  # 未知科目排在最后
            
            scores_list = list(scores)
            scores_list.sort(key=get_subject_order_for_score)
            
            exam_data = {
                'id': exam.id,
                'name': exam.name,
                'academic_year': exam.academic_year,
                'exam_date': exam.date.strftime('%Y-%m-%d') if exam.date else None,
                'grade_level': exam.get_grade_level_display(),  # 添加年级信息
                'scores': [],
                'total_score': 0,
                'average_score': 0,
                'grade_total_rank': None,
                'class_total_rank': None,
            }
            
            total_score = 0
            valid_scores = 0
            
            for score in scores_list:
                subject_name = score.subject
                all_subjects.add(subject_name)
                
                # 获取该科目的正确满分
                max_score = score.get_max_score()
                
                score_data = {
                    'subject_name': subject_name,
                    'score_value': float(score.score_value) if score.score_value else 0,
                    'full_score': float(max_score),  # 使用该科目的实际满分
                    'grade_rank': score.grade_rank_in_subject,
                    'class_rank': score.class_rank_in_subject,
                    'percentage': 0
                }
                
                # 计算百分比
                if score_data['full_score'] > 0:
                    score_data['percentage'] = round((score_data['score_value'] / score_data['full_score']) * 100, 1)
                
                exam_data['scores'].append(score_data)
                
                # 累计总分
                if score.score_value:
                    total_score += float(score.score_value)
                    valid_scores += 1
                
                # 构建趋势数据
                if subject_name not in analysis_data['trend_data']:
                    analysis_data['trend_data'][subject_name] = {
                        'class_ranks': [],
                        'grade_ranks': [],
                        'scores': [],
                        'exam_names': []
                    }
                
                analysis_data['trend_data'][subject_name]['class_ranks'].append(score.class_rank_in_subject)
                analysis_data['trend_data'][subject_name]['grade_ranks'].append(score.grade_rank_in_subject)
                analysis_data['trend_data'][subject_name]['scores'].append(score_data['score_value'])
                analysis_data['trend_data'][subject_name]['exam_names'].append(exam.name)
            
            # 计算考试总分和平均分
            exam_data['total_score'] = round(total_score, 1)
            if valid_scores > 0:
                exam_data['average_score'] = round(total_score / valid_scores, 1)
            
            # 获取总分排名
            if scores_list:
                first_score = scores_list[0]
                exam_data['grade_total_rank'] = first_score.total_score_rank_in_grade
                exam_data['class_total_rank'] = first_score.total_score_rank_in_class
            
            analysis_data['exams'].append(exam_data)
        
        # 设置科目列表和汇总信息 - 按SUBJECT_CHOICES顺序排序
        def get_subject_order(subject):
            """获取科目在SUBJECT_CHOICES中的顺序"""
            for index, (subject_code, subject_name) in enumerate(SUBJECT_CHOICES):
                if subject_code == subject or subject_name == subject:
                    return index
            return 999  # 未知科目排在最后
        
        analysis_data['subjects'] = sorted(list(all_subjects), key=get_subject_order)
        analysis_data['summary']['subjects_count'] = len(all_subjects)
        
        # 构建总分趋势数据
        analysis_data['trend_data']['total'] = {
            'class_ranks': [exam_data['class_total_rank'] for exam_data in analysis_data['exams']],
            'grade_ranks': [exam_data['grade_total_rank'] for exam_data in analysis_data['exams']],
            'scores': [exam_data['total_score'] for exam_data in analysis_data['exams']],
            'exam_names': [exam_data['name'] for exam_data in analysis_data['exams']]
        }
        
        return JsonResponse({
            'success': True,
            'data': analysis_data
        })
        
    except Student.DoesNotExist:
        return JsonResponse({'error': '学生不存在'}, status=404)
    except Exam.DoesNotExist:
        return JsonResponse({'error': '考试不存在'}, status=404)
    except Exception as e:
        return JsonResponse({'error': f'服务器错误: {str(e)}'}, status=500)

# 学生个人成绩详细分析页面
def score_analysis_student_detail(request):
    """学生个人成绩详细分析页面，显示图表和数据分析"""
    
    # 获取筛选参数
    grade_level = request.GET.get('grade_level')
    class_name = request.GET.get('class_name')
    student_id = request.GET.get('student_id')
    
    if not all([grade_level, class_name, student_id]):
        messages.error(request, '缺少必要的分析参数')
        return redirect('students_grades:student_analysis')
    
    # 查找学生
    try:
        # 根据隐藏字段的ID值查找学生
        student = Student.objects.select_related('current_class').get(pk=student_id)
    except Student.DoesNotExist:
        messages.error(request, '学生不存在')
        return redirect('students_grades:student_analysis')
    
    # 获取该学生的所有考试成绩
    scores = Score.objects.filter(
        student=student
    ).select_related('exam').order_by('-exam__date')
    
    # 获取该学生参与的所有考试
    exams = Exam.objects.filter(
        id__in=scores.values_list('exam_id', flat=True).distinct()
    ).order_by('-date')
    
    context = {
        'student': student,
        'scores': scores,
        'exams': exams,
        'subjects': SUBJECT_CHOICES,
        'page_title': f'{student.name} - 个人成绩分析',
        'selected_grade': grade_level,
        'selected_class': class_name,
        'selected_student_id': student_id,
    }
    
    return render(request, 'scores/score_analysis_student_detail.html', context)

# AJAX接口：获取所有年级列表
def get_grades_ajax(request):
    """获取系统中所有年级的AJAX接口"""
    try:
        # 从学生表中获取所有不重复的年级
        grades = Student.objects.values_list('grade_level', flat=True).distinct().order_by('grade_level')
        
        # 过滤掉空值并构建返回数据
        grade_list = []
        grade_display_map = {
            'Grade7': '初一',
            'Grade8': '初二', 
            'Grade9': '初三',
            'Grade10': '高一',
            'Grade11': '高二',
            'Grade12': '高三'
        }
        
        for grade in grades:
            if grade:  # 过滤掉None或空字符串
                display_name = grade_display_map.get(grade, grade)
                grade_list.append({
                    'value': grade,
                    'display_name': display_name
                })
        
        # 如果没有找到年级数据，返回默认年级
        if not grade_list:
            grade_list = [
                {'value': 'Grade10', 'display_name': '高一'},
                {'value': 'Grade11', 'display_name': '高二'},
                {'value': 'Grade12', 'display_name': '高三'}
            ]
        
        return JsonResponse({
            'grades': grade_list,
            'count': len(grade_list)
        })
    except Exception as e:
        return JsonResponse({'error': f'获取年级列表失败: {str(e)}'}, status=500)

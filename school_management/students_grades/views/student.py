import datetime

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from django.db import transaction
from django.http import HttpResponse
from django.utils import timezone
from django_filters.rest_framework import DjangoFilterBackend
from rest_framework import filters, permissions, status, viewsets
from rest_framework.decorators import action
from rest_framework.parsers import FormParser, MultiPartParser
from rest_framework.response import Response

from school_management.users.permissions import IsAdminOrStaff

from ..models.score import Score
from ..models.student import (
    Student,
    Class,
    STATUS_CHOICES,
    GRADE_LEVEL_CHOICES,
    CLASS_NAME_CHOICES,
    COHORT_CHOICES,
)
from ..serializers import StudentSerializer
from ..tasks import update_all_rankings_async

class StudentViewSet(viewsets.ModelViewSet):
    """
    学生管理 ViewSet
    提供学生的 CRUD 操作

    这个 ViewSet 同时也提供一些辅助 API 以方便前端在单页应用中获取数据。
    """
    queryset = Student.objects.all().select_related('current_class')
    serializer_class = StudentSerializer
    permission_classes = [permissions.IsAuthenticated]
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    # cohort 存储 cohort 格式（如"初中2023级"），grade_level 保留用于展示
    filterset_fields = ['status', 'gender', 'current_class__cohort', 'current_class__class_name']
    search_fields = ['name', 'student_id', 'current_class__class_name']
    ordering_fields = ['student_id', 'name', 'entry_date']
    ordering = ['student_id']

    def get_permissions(self):
        """
        根据操作类型设置不同的权限
        """
        if self.action in [
            'create', 'update', 'partial_update', 'destroy',
            'batch_delete', 'batch_update_status', 'batch_promote', 'batch_import'
        ]:
            permission_classes = [permissions.IsAuthenticated, IsAdminOrStaff]
        else:
            # 查看操作所有登录用户都可以
            permission_classes = [permissions.IsAuthenticated]
        
        return [permission() for permission in permission_classes]

    def perform_destroy(self, instance):
        # 获取该学生参与的所有考试ID，用于后续排名更新
        affected_exam_ids = list(Score.objects.filter(student=instance).values_list('exam_id', flat=True).distinct())
        instance.delete()
        
        # 异步更新受影响考试的排名
        if affected_exam_ids:
            for exam_id in affected_exam_ids:
                try:
                    update_all_rankings_async.delay(exam_id)
                except Exception:
                    pass  # 静默处理任务提交错误

    def perform_update(self, serializer):
        # 检查是否变更为毕业状态，如果是则补充毕业日期
        instance = serializer.save()
        if instance.status == '毕业' and not instance.graduation_date:
            instance.graduation_date = timezone.now().date()
            instance.save(update_fields=['graduation_date'])

    @action(detail=False, methods=['post'], url_path='batch-delete')
    def batch_delete(self, request):
        """批量删除学生"""
        student_ids = request.data.get('student_ids', [])
        if not student_ids:
            return Response({'success': False, 'message': '没有选择任何学生。'}, status=status.HTTP_400_BAD_REQUEST)

        try:
            with transaction.atomic():
                students_to_delete = Student.objects.filter(pk__in=student_ids)
                affected_exam_ids = list(Score.objects.filter(student__in=students_to_delete).values_list('exam_id', flat=True).distinct())
                
                deleted_count, _ = students_to_delete.delete()
                
                if affected_exam_ids:
                    for exam_id in affected_exam_ids:
                        try:
                            update_all_rankings_async.delay(exam_id)
                        except Exception:
                            pass
                            
                return Response({
                    'success': True,
                    'message': f'成功删除 {deleted_count} 名学生。'
                })
        except Exception as e:
            return Response({'success': False, 'message': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    @action(detail=False, methods=['post'], url_path='batch-update-status')
    def batch_update_status(self, request):
        """批量修改学生状态（支持批量毕业）"""
        student_ids = request.data.get('student_ids', [])
        new_status = request.data.get('status', '')
        
        if not student_ids or not new_status:
            return Response({'success': False, 'message': '缺少必要参数。'}, status=status.HTTP_400_BAD_REQUEST)
            
        try:
            with transaction.atomic():
                students_to_update = Student.objects.filter(pk__in=student_ids)
                
                if new_status == '毕业':
                    # 只更新状态不是毕业的，设置当前的 graduation_date
                    updated_count = students_to_update.exclude(status='毕业').update(
                        status=new_status, 
                        graduation_date=timezone.now().date()
                    )
                    # 已经是毕业的无需改毕业时间，但也保障新状态一致性，不过这句其实无异于不变。
                    students_to_update.filter(status='毕业').update(status=new_status)
                else:
                    updated_count = students_to_update.update(status=new_status)
                    
                return Response({
                    'success': True, 
                    'message': f'成功更新 {updated_count} 名学生的状态为 "{new_status}"。'
                })
        except Exception as e:
            return Response({'success': False, 'message': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    @action(detail=False, methods=['get'], url_path='stats')
    def stats(self, request):
        """返回学生统计数据，用于前端展示总数/状态分布等。"""
        # 与原 student_list 视图保持一致：统计所有学生（不受过滤影响）
        all_students = Student.objects.all()
        total_students = all_students.count()
        active_students = all_students.filter(status='在读').count()
        graduated_students = all_students.filter(status='毕业').count()
        suspended_students = all_students.filter(status='休学').count()

        return Response({
            'total_students': total_students,
            'active_students': active_students,
            'graduated_students': graduated_students,
            'suspended_students': suspended_students,
            'status_choices': [c[0] for c in STATUS_CHOICES],
            'grade_level_choices': [c[0] for c in GRADE_LEVEL_CHOICES],  # deprecated
            'cohort_choices': [c[0] for c in COHORT_CHOICES],
            'class_name_choices': [c[0] for c in CLASS_NAME_CHOICES],
        })


    @action(detail=False, methods=['post'], url_path='batch-promote')
    def batch_promote(self, request):
        """
        批量升年级
        期望 payload:
        {
            "student_ids": [1, 2, 3],
            "target_grade_level": "二年级",
            "current_grade_level": "(可选)一年级",
            "auto_create_classes": true
        }
        """
        from django.db import transaction
        student_ids = request.data.get('student_ids', [])
        target_grade = request.data.get('target_grade_level')
        from_grade = request.data.get('current_grade_level')
        auto_create = request.data.get('auto_create_classes', False)

        if not student_ids:
            return Response({"success": False, "message": "未选择任何学生"}, status=400)
        if not target_grade:
            return Response({"success": False, "message": "请选择目标年级"}, status=400)
        if from_grade and from_grade == target_grade:
            return Response({"success": False, "message": "目标年级不能与当前年级相同"}, status=400)

        updated_count = 0
        errors = []

        with transaction.atomic():
            students = Student.objects.filter(pk__in=student_ids)
            if from_grade:
                students = students.filter(current_class__grade_level=from_grade)

            for student in students:
                try:
                    current_class = student.current_class
                    if current_class:
                        target_class_obj = Class.objects.filter(
                            class_name=current_class.class_name,
                            grade_level=target_grade
                        ).first()

                        if not target_class_obj and auto_create:
                            target_class_obj = Class.objects.create(
                                class_name=current_class.class_name,
                                grade_level=target_grade
                            )

                        if target_class_obj:
                            student.current_class = target_class_obj
                            student.grade_level = target_grade
                            student.save()
                            updated_count += 1
                        else:
                            errors.append(f"学生 {student.name} 所在的班级在目标年级中不存在，且未开启自动创建。")
                    else:
                        errors.append(f"学生 {student.name} 当前无班级，跳过。")
                except Exception as e:
                    errors.append(f"学生 {student.name} 失败: {str(e)}")

        return Response({
            "success": updated_count > 0,
            "message": f"成功将 {updated_count} 名学生升入 {target_grade}",
            "updated_count": updated_count,
            "errors": errors
        })

    @action(detail=False, methods=['get'], url_path='download-template')
    def download_template(self, request):
        """下载批量导入模板（新版：学段+届别格式）"""
        from ..models.student import SECTION_CHOICES, COHORT_YEAR_CHOICES

        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "学生导入模板"

        # 样式定义
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill("solid", fgColor="00897B")
        header_alignment = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )

        headers = [
            "学号 (必填)", "姓名 (必填)", "性别 (男/女)", "出生日期 (YYYY-MM-DD)",
            "学段 (初中/高中)", "届别年份 (纯数字，如2026)", "年级 (初一/初二/初三/高一/高二/高三)",
            "班级名称 (1班-20班)", "在校状态 (在读/转学/休学/复学/毕业)", "身份证号码", "学籍号",
            "家庭地址", "监护人姓名", "监护人联系电话",
            "入学日期 (YYYY-MM-DD)", "毕业日期 (YYYY-MM-DD)"
        ]

        # 添加表头
        for col, header in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        # 添加示例数据
        example_row = [
            "S001", "张三", "男", "2010-05-15",
            "初中", "2026", "初一", "1班", "在读", "", "",
            "", "张父", "13800138000", "2023-09-01", ""
        ]
        for col, value in enumerate(example_row, 1):
            cell = sheet.cell(row=2, column=col, value=value)
            cell.alignment = header_alignment
            cell.border = thin_border

        # 设置列宽
        col_widths = [15, 10, 8, 18, 15, 18, 22, 15, 20, 15, 15, 20, 12, 15, 15, 15]
        for col, width in enumerate(col_widths, 1):
            sheet.column_dimensions[chr(64 + col) if col <= 26 else 'A' + chr(64 + col - 26)].width = width

        # 设置行高
        sheet.row_dimensions[1].height = 25

        # 添加提示信息
        section_options = ', '.join([f"{s[0]}({s[1]})" for s in SECTION_CHOICES])
        sheet.cell(row=2, column=5).comment = openpyxl.comments.Comment(
            f"请填写: {section_options}", "System")

        cohort_options = ', '.join([c[0] for c in COHORT_YEAR_CHOICES])
        sheet.cell(row=2, column=6).comment = openpyxl.comments.Comment(
            f"请填写纯数字，如: {cohort_options}", "System")

        grade_level_options = ', '.join([g[0] for g in GRADE_LEVEL_CHOICES])
        sheet.cell(row=2, column=7).comment = openpyxl.comments.Comment(
            f"请填写: {grade_level_options}", "System")

        class_name_options = ', '.join([choice[0] for choice in CLASS_NAME_CHOICES])
        sheet.cell(row=2, column=8).comment = openpyxl.comments.Comment(
            f"请填写: {class_name_options}", "System")

        status_options = ', '.join([s[0] for s in STATUS_CHOICES])
        sheet.cell(row=2, column=9).comment = openpyxl.comments.Comment(
            f"请填写: {status_options}", "System")

        date_format_text = "日期格式必须是 YYYY-MM-DD，例如 2006-01-23"
        sheet.cell(row=2, column=4).comment = openpyxl.comments.Comment(date_format_text, "System")
        sheet.cell(row=2, column=15).comment = openpyxl.comments.Comment(date_format_text, "System")
        sheet.cell(row=2, column=16).comment = openpyxl.comments.Comment(date_format_text, "System")

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="student_import_template.xlsx"'
        workbook.save(response)
        return response

    @action(detail=False, methods=['post'], url_path='batch-import', parser_classes=[MultiPartParser, FormParser])
    def batch_import(self, request):
        """批量导入学生"""
        if 'file' not in request.FILES:
            return Response({'success': False, 'message': "请选择正确的 Excel 文件。"}, status=400)
        
        excel_file = request.FILES['file']
        
        if not excel_file.name.endswith(('.xlsx', '.xls')):
            return Response({'success': False, 'message': "文件格式不正确，请上传 .xlsx 或 .xls 文件。"}, status=400)

        try:
            workbook = openpyxl.load_workbook(excel_file)
            sheet = workbook.active
            
            header = [cell.value for cell in sheet[1]]
            
            header_mapping = {
                "学号 (必填)": "student_id",
                "姓名 (必填)": "name",
                "性别 (男/女)": "gender",
                "出生日期 (YYYY-MM-DD)": "date_of_birth",
                "年级 (初一/初二/初三/高一/高二/高三)": "grade_level",
                "学段 (初中/高中)": "_section",
                "届别年份 (纯数字，如2026)": "_cohort_year",
                "班级名称 (1班-20班)": "class_name",
                "在校状态 (在读/转学/休学/复学/毕业)": "status",
                "身份证号码": "id_card_number",
                "学籍号": "student_enrollment_number",
                "家庭地址": "home_address",
                "监护人姓名": "guardian_name",
                "监护人联系电话": "guardian_contact_phone",
                "入学日期 (YYYY-MM-DD)": "entry_date",
                "毕业日期 (YYYY-MM-DD)": "graduation_date",
            }

            imported_count = 0
            failed_rows = []
            success_messages = []
            error_messages = []
            warning_messages = []

            # 检测Excel中是否有重复学号
            excel_student_ids = []
            duplicate_rows = []

            for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                if not any(row):
                    continue
                row_data = dict(zip(header, row))
                sid = row_data.get('学号 (必填)') or row_data.get('学号')
                if sid:
                    sid = str(sid).strip()
                    if sid in excel_student_ids:
                        duplicate_rows.append({'row': row_idx, 'student_id': sid})
                    excel_student_ids.append(sid)

            if duplicate_rows:
                warning_messages.append(f"检测到Excel中有重复学号：{', '.join([r['student_id'] for r in duplicate_rows])}")

            from django.db import transaction

            for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                if not any(row):
                    continue

                try:
                    with transaction.atomic():
                        row_data = dict(zip(header, row))
                        student_data = {}
                        for excel_header, model_field in header_mapping.items():
                            if excel_header in row_data:
                                student_data[model_field] = row_data[excel_header]

                        if not student_data.get('student_id') or not student_data.get('name'):
                            raise ValueError("学号和姓名为必填字段")

                        if student_data.get('student_id'):
                            student_data['student_id'] = str(student_data['student_id']).strip()

                        if student_data.get('id_card_number'):
                            student_data['id_card_number'] = str(student_data['id_card_number']).strip()
                        # 学籍号：空字符串也设为None，避免唯一约束冲突
                        enrollment = student_data.get('student_enrollment_number')
                        if enrollment and str(enrollment).strip():
                            student_data['student_enrollment_number'] = str(enrollment).strip()
                        else:
                            student_data['student_enrollment_number'] = None

                        for date_field in ['date_of_birth', 'entry_date', 'graduation_date']:
                            if date_field in student_data:
                                date_value = student_data[date_field]
                                if date_value is None or date_value == '' or str(date_value).strip() == '':
                                    student_data[date_field] = None
                                    continue
                                
                                try:
                                    if isinstance(date_value, datetime.datetime):
                                        student_data[date_field] = date_value.date()
                                    elif isinstance(date_value, datetime.date):
                                        pass
                                    else:
                                        date_str = str(date_value).strip()
                                        if date_str:
                                            for date_format in ['%Y-%m-%d', '%Y/%m/%d', '%Y.%m.%d']:
                                                try:
                                                    student_data[date_field] = datetime.datetime.strptime(date_str, date_format).date()
                                                    break
                                                except ValueError:
                                                    continue
                                            else:
                                                student_data[date_field] = None
                                                warning_messages.append(f"第 {row_idx} 行的 '{date_field}' 日期格式不正确 (值: '{date_value}')")
                                        else:
                                            student_data[date_field] = None
                                except Exception as e:
                                    student_data[date_field] = None
                                    warning_messages.append(f"第 {row_idx} 行的 '{date_field}' 日期处理出错: {e}")
                            else:
                                student_data[date_field] = None

                        if 'gender' in student_data and student_data['gender']:
                            gender_value = str(student_data['gender']).strip()
                            gender_mapping = {
                                '男': '男', 'M': '男', 'Male': '男', 'male': '男', '1': '男',
                                '女': '女', 'F': '女', 'Female': '女', 'female': '女', '0': '女'
                            }
                            student_data['gender'] = gender_mapping.get(gender_value, gender_value)
                            if student_data['gender'] not in ['男', '女']:
                                warning_messages.append(f"第 {row_idx} 行的性别值 '{gender_value}' 无效，已设置为空")
                                student_data['gender'] = None

                        # 处理学段 + 届别年份 → cohort
                        section = student_data.pop('_section', None)
                        cohort_year = student_data.pop('_cohort_year', None)
                        class_name = student_data.pop('class_name', None)
                        current_class_obj = None

                        # grade_level 直接从 Excel 读取（传统格式如"初一"）
                        # cohort = section + cohort_year + "级"（如"初中2024级"）
                        if section and cohort_year:
                            cohort_year_str = str(cohort_year).strip()
                            cohort_value = f"{section}{cohort_year_str}级"
                            student_data['cohort'] = cohort_value

                        if class_name:
                            # 使用 grade_level + class_name 作为查找键，确保正确匹配
                            try:
                                current_class_obj, created = Class.objects.get_or_create(
                                    grade_level=student_data.get('grade_level', ''),
                                    class_name=class_name,
                                    defaults={
                                        'grade_level': student_data.get('grade_level', ''),
                                        'cohort': student_data.get('cohort', '')
                                    }
                                )
                                # 如果已存在但 cohort 为空，自动补全 cohort
                                if not created and not current_class_obj.cohort and student_data.get('cohort'):
                                    current_class_obj.cohort = student_data.get('cohort')
                                    current_class_obj.save(update_fields=['cohort'])
                                    warning_messages.append(
                                        f"第 {row_idx} 行：班级 {student_data.get('grade_level', '')}{class_name} 的 cohort 已自动补全为 {student_data.get('cohort')}"
                                    )
                                if created:
                                    success_messages.append(f"自动创建班级：{student_data.get('grade_level', '')}{class_name}")
                            except Exception as e:
                                raise ValueError(f"班级创建/查找失败: {e}")

                        student_data['current_class'] = current_class_obj

                        student_id = student_data.get('student_id')
                        student_obj, created = Student.objects.update_or_create(
                            student_id=student_id,
                            defaults=student_data
                        )

                        if created:
                            success_messages.append(f"成功新增学生：{student_obj.name} ({student_obj.student_id})")
                        else:
                            success_messages.append(f"成功更新学生：{student_obj.name} ({student_obj.student_id})")
                        imported_count += 1
                        
                except Exception as e:
                    failed_rows.append({'row': row_idx, 'error': str(e)})
                    error_messages.append(f"第 {row_idx} 行学生导入失败: {e}")
                    continue

            return Response({
                'success': True,
                'imported_count': imported_count,
                'failed_count': len(failed_rows),
                'success_messages': success_messages,
                'error_messages': error_messages,
                'warning_messages': warning_messages,
                'failed_rows': failed_rows
            })

        except Exception as e:
            return Response({
                'success': False,
                'message': f"解析文件时出现严重错误: {str(e)}"
            }, status=500)



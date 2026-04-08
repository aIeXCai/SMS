from django.utils import timezone
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.series import DataPoint
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


class StudentAnalysisExportService:
    """个人成绩分析导出服务。"""

    THEME_HEADER_BG = "0000897B"
    THEME_HEADER_FONT = "00FFFFFF"
    THEME_SECTION_BG = "00E0F2F1"
    THEME_BORDER = "00D9D9D9"
    POSITIVE_FILL = "00C6EFCE"
    POSITIVE_FONT = "00006100"
    NEGATIVE_FILL = "00FFC7CE"
    NEGATIVE_FONT = "009C0006"
    CHART_MAIN = "00897B"
    CHART_SECONDARY = "26A69A"
    CHART_SCORE_TOTAL = "1565C0"
    CHART_SCORE_ROLLING = "EF6C00"
    CHART_RANK = "1E88E5"
    CHART_NEUTRAL = "90A4AE"
    CHART_NEGATIVE_SOFT = "D9A3A3"

    @staticmethod
    def _write_chart_notice(sheet, cell, message, is_error=False):
        notice_cell = sheet[cell]
        notice_cell.value = message
        notice_cell.font = Font(
            size=10,
            bold=True,
            color=("00B71C1C" if is_error else "00616161"),
        )

    @classmethod
    def _apply_trend_line_style(cls, line_chart, color):
        line_chart.graphicalProperties.line.solidFill = color
        line_chart.graphicalProperties.line.width = 22000

    @staticmethod
    def _set_series_line_style(series, color, width=22000):
        series.graphicalProperties.line.solidFill = color
        series.graphicalProperties.line.width = width
        series.marker.symbol = "triangle"
        series.marker.size = 7

    @staticmethod
    def _set_series_data_label(series, position="t"):
        series.dLbls = DataLabelList()
        series.dLbls.showVal = True
        series.dLbls.showCatName = False
        series.dLbls.showSerName = False
        series.dLbls.showLegendKey = False
        series.dLbls.dLblPos = position

    @staticmethod
    def _auto_width_for_column(values, factor=1.5, min_width=10, max_width=48):
        max_len = max((len(str(item)) for item in values if item not in [None, ""]), default=min_width)
        return max(min_width, min(max_width, int(max_len * factor) + 2))

    @staticmethod
    def _to_number(value):
        try:
            if value is None:
                return None
            return float(value)
        except (TypeError, ValueError):
            return None

    @staticmethod
    def format_number(value, digits=1):
        if value is None:
            return "-"
        if isinstance(value, int):
            return str(value)

        try:
            numeric = float(value)
        except (TypeError, ValueError):
            return "-"

        if abs(numeric - int(numeric)) < 1e-9:
            return str(int(numeric))
        return f"{numeric:.{digits}f}"

    @staticmethod
    def format_change(value):
        if value is None:
            return "-"

        try:
            number = int(value)
        except (TypeError, ValueError):
            return "-"

        return f"+{number}" if number > 0 else str(number)

    @classmethod
    def build_payload(cls, analysis_data, subject_choices):
        exams = list(analysis_data.get("exams") or [])
        if not exams:
            raise ValueError("该学生暂无可导出分析数据")

        exams.sort(key=lambda item: ((item.get("exam_date") or ""), item.get("id") or 0))

        subject_priority = {subject_code: index for index, (subject_code, _) in enumerate(subject_choices)}
        subjects = list(analysis_data.get("subjects") or [])
        subjects.sort(key=lambda item: subject_priority.get(item, 999))

        def extract_subject_score(exam_item, subject_name):
            for score_item in exam_item.get("scores") or []:
                if score_item.get("subject_name") == subject_name:
                    return score_item
            return None

        first_exam = exams[0]
        latest_exam = exams[-1]
        prev_exam = exams[-2] if len(exams) > 1 else None

        total_scores = [cls._to_number(item.get("total_score")) for item in exams if cls._to_number(item.get("total_score")) is not None]
        avg_total_score = round(sum(total_scores) / len(total_scores), 1) if total_scores else None
        best_exam = max(exams, key=lambda item: cls._to_number(item.get("total_score")) if cls._to_number(item.get("total_score")) is not None else -1)
        worst_exam = min(exams, key=lambda item: cls._to_number(item.get("total_score")) if cls._to_number(item.get("total_score")) is not None else 10 ** 9)

        grade_rank_values = [item.get("grade_total_rank") for item in exams if item.get("grade_total_rank") is not None]
        best_rank_exam = min(exams, key=lambda item: item.get("grade_total_rank") if item.get("grade_total_rank") is not None else 10 ** 9)
        worst_rank_exam = max(exams, key=lambda item: item.get("grade_total_rank") if item.get("grade_total_rank") is not None else -1)

        latest_grade_rank = latest_exam.get("grade_total_rank")
        first_grade_rank = first_exam.get("grade_total_rank")
        prev_grade_rank = prev_exam.get("grade_total_rank") if prev_exam else None
        latest_total_score = cls._to_number(latest_exam.get("total_score"))
        prev_total_score = cls._to_number(prev_exam.get("total_score")) if prev_exam else None

        latest_subject_snapshot = []
        for subject_name in subjects:
            subject_score = extract_subject_score(latest_exam, subject_name)
            if not subject_score:
                latest_subject_snapshot.append(
                    {
                        "subject_name": subject_name,
                        "score_value": "-",
                        "class_rank": "-",
                        "grade_rank": "-",
                    }
                )
                continue

            latest_subject_snapshot.append(
                {
                    "subject_name": subject_name,
                    "score_value": subject_score.get("score_value") if subject_score.get("score_value") is not None else "-",
                    "class_rank": subject_score.get("class_rank") if subject_score.get("class_rank") is not None else "-",
                    "grade_rank": subject_score.get("grade_rank") if subject_score.get("grade_rank") is not None else "-",
                }
            )

        return {
            "student_info": analysis_data.get("student_info") or {},
            "subjects": subjects,
            "exams": exams,
            "overview": {
                "analysis_time": timezone.localtime(timezone.now()).strftime("%Y-%m-%d %H:%M"),
                "total_exams": len(exams),
                "avg_total_score": avg_total_score,
                "best_exam": best_exam,
                "worst_exam": worst_exam,
                "best_grade_rank_exam": best_rank_exam if grade_rank_values else None,
                "worst_grade_rank_exam": worst_rank_exam if grade_rank_values else None,
                "score_range": (
                    round(max(total_scores) - min(total_scores), 1)
                    if len(total_scores) >= 2
                    else (0 if len(total_scores) == 1 else None)
                ),
                "latest_exam": latest_exam,
                "latest_total_score": latest_total_score,
                "latest_grade_rank": latest_grade_rank,
                "latest_class_rank": latest_exam.get("class_total_rank"),
                "delta_total_score": (
                    round(latest_total_score - prev_total_score, 1)
                    if latest_total_score is not None and prev_total_score is not None
                    else None
                ),
                "delta_grade_rank": (
                    (prev_grade_rank - latest_grade_rank)
                    if prev_grade_rank is not None and latest_grade_rank is not None
                    else None
                ),
                "first_to_latest_rank_change": (
                    (first_grade_rank - latest_grade_rank)
                    if first_grade_rank is not None and latest_grade_rank is not None
                    else None
                ),
                "latest_subject_snapshot": latest_subject_snapshot,
            },
        }

    @classmethod
    def build_overview_sheet(cls, workbook, payload):
        sheet = workbook.active
        sheet.title = "总览"

        student_info = payload.get("student_info") or {}
        overview = payload.get("overview") or {}

        title_font = Font(size=14, bold=True, color=cls.THEME_HEADER_FONT)
        section_font = Font(size=11, bold=True)
        label_font = Font(size=10, bold=True)
        value_font = Font(size=10)
        title_fill = PatternFill(fill_type="solid", fgColor=cls.THEME_HEADER_BG)
        section_fill = PatternFill(fill_type="solid", fgColor=cls.THEME_SECTION_BG)
        border = Border(
            left=Side(style="thin", color=cls.THEME_BORDER),
            right=Side(style="thin", color=cls.THEME_BORDER),
            top=Side(style="thin", color=cls.THEME_BORDER),
            bottom=Side(style="thin", color=cls.THEME_BORDER),
        )

        sheet.column_dimensions["A"].width = 26
        sheet.column_dimensions["B"].width = 52

        row_index = 1

        def write_title(text):
            nonlocal row_index
            sheet.merge_cells(start_row=row_index, start_column=1, end_row=row_index, end_column=2)
            cell = sheet.cell(row=row_index, column=1, value=text)
            cell.font = title_font
            cell.fill = title_fill
            cell.border = border
            cell.alignment = Alignment(horizontal="center", vertical="center")
            sheet.cell(row=row_index, column=2).border = border
            sheet.cell(row=row_index, column=2).fill = title_fill
            row_index += 1

        def write_section(text):
            nonlocal row_index
            sheet.merge_cells(start_row=row_index, start_column=1, end_row=row_index, end_column=2)
            cell = sheet.cell(row=row_index, column=1, value=text)
            cell.font = section_font
            cell.fill = section_fill
            cell.alignment = Alignment(horizontal="left", vertical="center")
            cell.border = border
            sheet.cell(row=row_index, column=2).border = border
            row_index += 1

        def write_item(label, value):
            nonlocal row_index
            label_cell = sheet.cell(row=row_index, column=1, value=label)
            value_cell = sheet.cell(row=row_index, column=2, value=value if value is not None else "-")
            label_cell.font = label_font
            value_cell.font = value_font
            label_cell.alignment = Alignment(horizontal="left", vertical="center")
            value_cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            label_cell.border = border
            value_cell.border = border
            row_index += 1

        student_name = student_info.get("name") or "未知学生"
        student_no = student_info.get("student_id") or "-"
        write_title(f"个人成绩分析报告（{student_name} / {student_no}）")

        write_section("A. 学生信息区")
        write_item("姓名", student_name)
        write_item("学号", student_no)
        write_item("年级", student_info.get("grade_level") or "-")
        write_item("班级", student_info.get("class_name") or "-")
        write_item("分析时间", overview.get("analysis_time") or "-")
        write_item("分析覆盖考试数", overview.get("total_exams") or 0)

        best_exam = overview.get("best_exam") or {}
        worst_exam = overview.get("worst_exam") or {}
        write_section("B. 总分概览区")
        write_item("总分均分", cls.format_number(overview.get("avg_total_score")))
        write_item(
            "最好总分+排名",
            f"{cls.format_number(best_exam.get('total_score'))}（班级{best_exam.get('class_total_rank') or '-'} / 年级{best_exam.get('grade_total_rank') or '-'}）",
        )
        write_item(
            "最低总分+排名",
            f"{cls.format_number(worst_exam.get('total_score'))}（班级{worst_exam.get('class_total_rank') or '-'} / 年级{worst_exam.get('grade_total_rank') or '-'}）",
        )
        write_item("总分波动区间", cls.format_number(overview.get("score_range")))

        best_rank_exam = overview.get("best_grade_rank_exam") or {}
        worst_rank_exam = overview.get("worst_grade_rank_exam") or {}
        write_section("C. 排名概览区")
        write_item("最佳年级排名", f"{best_rank_exam.get('grade_total_rank') or '-'}（{best_rank_exam.get('name') or '-'}）")
        write_item("最差年级排名", f"{worst_rank_exam.get('grade_total_rank') or '-'}（{worst_rank_exam.get('name') or '-'}）")
        write_item("相对首场考试排名变化", cls.format_change(overview.get("first_to_latest_rank_change")))
        write_item("期初到期末排名净变化", cls.format_change(overview.get("first_to_latest_rank_change")))

        latest_exam = overview.get("latest_exam") or {}
        write_section("D. 最新考试快照区")
        write_item("最新考试", latest_exam.get("name") or "-")
        write_item("最新考试总分", cls.format_number(overview.get("latest_total_score")))
        write_item(
            "最新考试班级/年级排名",
            f"班级{overview.get('latest_class_rank') or '-'} / 年级{overview.get('latest_grade_rank') or '-'}",
        )
        write_item("较上次总分变化", cls.format_change(overview.get("delta_total_score")))
        write_item("较上次年级排名变化", cls.format_change(overview.get("delta_grade_rank")))

        write_section("E. 最新科目快照区")
        for item in overview.get("latest_subject_snapshot") or []:
            write_item(
                item.get("subject_name") or "-",
                f"{cls.format_number(item.get('score_value'))}（班排{item.get('class_rank')} / 级排{item.get('grade_rank')}）",
            )

    @staticmethod
    def _set_table_header(sheet, headers):
        header_font = Font(size=10, bold=True, color=StudentAnalysisExportService.THEME_HEADER_FONT)
        header_fill = PatternFill(fill_type="solid", fgColor=StudentAnalysisExportService.THEME_HEADER_BG)
        header_alignment = Alignment(horizontal="center", vertical="center")
        border = Border(
            left=Side(style="thin", color=StudentAnalysisExportService.THEME_BORDER),
            right=Side(style="thin", color=StudentAnalysisExportService.THEME_BORDER),
            top=Side(style="thin", color=StudentAnalysisExportService.THEME_BORDER),
            bottom=Side(style="thin", color=StudentAnalysisExportService.THEME_BORDER),
        )

        sheet.append(headers)
        for index in range(1, len(headers) + 1):
            cell = sheet.cell(row=1, column=index)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border

        return border

    @classmethod
    def _apply_rank_change_conditional_format(cls, sheet, column_index, start_row=2):
        if sheet.max_row < start_row:
            return

        column_letter = sheet.cell(row=1, column=column_index).column_letter
        cell_range = f"{column_letter}{start_row}:{column_letter}{sheet.max_row}"

        sheet.conditional_formatting.add(
            cell_range,
            CellIsRule(
                operator="greaterThan",
                formula=["0"],
                stopIfTrue=False,
                fill=PatternFill(fill_type="solid", fgColor=cls.POSITIVE_FILL),
                font=Font(color=cls.POSITIVE_FONT),
            ),
        )
        sheet.conditional_formatting.add(
            cell_range,
            CellIsRule(
                operator="lessThan",
                formula=["0"],
                stopIfTrue=False,
                fill=PatternFill(fill_type="solid", fgColor=cls.NEGATIVE_FILL),
                font=Font(color=cls.NEGATIVE_FONT),
            ),
        )

    @staticmethod
    def _normalize_rank(value):
        if value in [None, "", "-"]:
            return None
        try:
            return int(value)
        except (TypeError, ValueError):
            return None

    @staticmethod
    def _rank_change(prev_rank, curr_rank):
        if prev_rank is None or curr_rank is None:
            return "-"
        return prev_rank - curr_rank

    @classmethod
    def build_total_trend_sheet(cls, workbook, payload):
        sheet = workbook.create_sheet("总分趋势")
        headers = [
            "考试名称",
            "日期",
            "总分",
            "班级排名",
            "年级排名",
            "总分变化",
            "年级排名变化",
            "3次滚动均分",
            "相对首场排名变化",
        ]
        border = cls._set_table_header(sheet, headers)

        exams = payload.get("exams") or []
        first_rank = cls._normalize_rank(exams[0].get("grade_total_rank")) if exams else None

        numeric_totals = []
        for index, exam in enumerate(exams):
            prev_exam = exams[index - 1] if index > 0 else None

            current_total = cls._to_number(exam.get("total_score"))
            prev_total = cls._to_number(prev_exam.get("total_score")) if prev_exam else None

            current_grade_rank = cls._normalize_rank(exam.get("grade_total_rank"))
            prev_grade_rank = cls._normalize_rank(prev_exam.get("grade_total_rank")) if prev_exam else None
            current_class_rank = cls._normalize_rank(exam.get("class_total_rank"))

            numeric_totals.append(current_total)
            window_totals = [value for value in numeric_totals if value is not None][-3:]
            rolling_avg = round(sum(window_totals) / len(window_totals), 1) if window_totals else "-"

            row = [
                exam.get("name") or "-",
                exam.get("exam_date") or "-",
                current_total if current_total is not None else "-",
                current_class_rank if current_class_rank is not None else "-",
                current_grade_rank if current_grade_rank is not None else "-",
                (round(current_total - prev_total, 1) if current_total is not None and prev_total is not None else "-"),
                cls._rank_change(prev_grade_rank, current_grade_rank),
                rolling_avg,
                cls._rank_change(first_rank, current_grade_rank),
            ]
            sheet.append(row)

        sheet.freeze_panes = "A2"
        sheet.column_dimensions["A"].width = 32
        sheet.column_dimensions["B"].width = 14
        sheet.column_dimensions["C"].width = 10
        sheet.column_dimensions["D"].width = 10
        sheet.column_dimensions["E"].width = 10
        sheet.column_dimensions["F"].width = 10
        sheet.column_dimensions["G"].width = 12
        sheet.column_dimensions["H"].width = 12
        sheet.column_dimensions["I"].width = 14

        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=len(headers)):
            for cell in row:
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = border

        cls._apply_rank_change_conditional_format(sheet, 7)
        cls._apply_rank_change_conditional_format(sheet, 9)

        # 动态拉宽考试名称列，避免长考试名称拥挤。
        exam_name_values = [sheet.cell(row=r, column=1).value for r in range(1, sheet.max_row + 1)]
        sheet.column_dimensions["A"].width = cls._auto_width_for_column(exam_name_values, factor=1.6, min_width=18, max_width=52)

        # 图表：总分趋势 + 滚动均分（同图）
        # 图表：年级排名趋势（反向Y轴）
        try:
            numeric_exam_rows = [
                r for r in range(2, sheet.max_row + 1)
                if isinstance(sheet.cell(row=r, column=3).value, (int, float))
            ]

            if len(numeric_exam_rows) < 2:
                cls._write_chart_notice(sheet, "K2", "数据不足，未生成趋势图")
                return

            min_row = min(numeric_exam_rows)
            max_row = max(numeric_exam_rows)

            score_chart = LineChart()
            score_chart.title = "总分趋势（含滚动均分）"
            score_chart.y_axis.title = "分数"
            score_chart.x_axis.title = "考试"
            score_chart.height = 12
            score_chart.width = 30

            categories = Reference(sheet, min_col=1, min_row=min_row, max_row=max_row)
            score_ref = Reference(sheet, min_col=3, min_row=1, max_row=max_row)
            rolling_ref = Reference(sheet, min_col=8, min_row=1, max_row=max_row)

            score_chart.add_data(score_ref, titles_from_data=True)
            score_chart.add_data(rolling_ref, titles_from_data=True)
            score_chart.set_categories(categories)

            if len(score_chart.series) >= 1:
                cls._set_series_line_style(score_chart.series[0], cls.CHART_SCORE_TOTAL, 24000)
                cls._set_series_data_label(score_chart.series[0], "t")
            if len(score_chart.series) >= 2:
                cls._set_series_line_style(score_chart.series[1], cls.CHART_SCORE_ROLLING, 20000)
                cls._set_series_data_label(score_chart.series[1], "b")

            # 图表放在表格数据正下方
            first_chart_row = sheet.max_row + 3
            sheet.add_chart(score_chart, f"A{first_chart_row}")

            rank_numeric_rows = [
                r for r in range(2, sheet.max_row + 1)
                if isinstance(sheet.cell(row=r, column=5).value, int)
            ]
            if len(rank_numeric_rows) >= 2:
                rank_chart = LineChart()
                rank_chart.title = "年级排名趋势"
                rank_chart.y_axis.title = "排名"
                rank_chart.x_axis.title = "考试"
                rank_chart.y_axis.scaling.orientation = "maxMin"
                rank_chart.height = 12
                rank_chart.width = 30

                rank_min_row = min(rank_numeric_rows)
                rank_max_row = max(rank_numeric_rows)
                rank_categories = Reference(sheet, min_col=1, min_row=rank_min_row, max_row=rank_max_row)
                rank_ref = Reference(sheet, min_col=5, min_row=1, max_row=rank_max_row)
                rank_chart.add_data(rank_ref, titles_from_data=True)
                rank_chart.set_categories(rank_categories)

                if len(rank_chart.series) >= 1:
                    cls._set_series_line_style(rank_chart.series[0], cls.CHART_RANK, 22000)
                    cls._set_series_data_label(rank_chart.series[0], "b")

                # 第二张图放在第一张图正下方
                second_chart_row = first_chart_row + 28
                sheet.add_chart(rank_chart, f"A{second_chart_row}")
            else:
                cls._write_chart_notice(sheet, f"A{first_chart_row + 28}", "数据不足，未生成年级排名趋势图")

        except Exception:
            cls._write_chart_notice(sheet, f"A{sheet.max_row + 3}", "图表生成失败，请联系管理员", is_error=True)

    @classmethod
    def build_subject_detail_sheet(cls, workbook, payload):
        sheet = workbook.create_sheet("科目明细")
        subjects = payload.get("subjects") or []
        headers = ["日期", "考试名", "总分", "总分级排"]
        for subject in subjects:
            headers.extend([f"{subject}得分", f"{subject}班排", f"{subject}级排"])

        border = cls._set_table_header(sheet, headers)

        for exam in payload.get("exams") or []:
            score_map = {item.get("subject_name"): item for item in (exam.get("scores") or [])}
            row = [
                exam.get("exam_date") or "-",
                exam.get("name") or "-",
                cls._to_number(exam.get("total_score")) if cls._to_number(exam.get("total_score")) is not None else "-",
                cls._normalize_rank(exam.get("grade_total_rank")) if cls._normalize_rank(exam.get("grade_total_rank")) is not None else "-",
            ]

            for subject in subjects:
                score_item = score_map.get(subject)
                if not score_item:
                    row.extend(["-", "-", "-"])
                    continue

                score_value = cls._to_number(score_item.get("score_value"))
                class_rank = cls._normalize_rank(score_item.get("class_rank"))
                grade_rank = cls._normalize_rank(score_item.get("grade_rank"))
                row.extend([
                    score_value if score_value is not None else "-",
                    class_rank if class_rank is not None else "-",
                    grade_rank if grade_rank is not None else "-",
                ])

            sheet.append(row)

        sheet.freeze_panes = "A2"
        sheet.column_dimensions["A"].width = 14
        sheet.column_dimensions["B"].width = 32
        sheet.column_dimensions["C"].width = 10
        sheet.column_dimensions["D"].width = 10
        for col_index in range(5, len(headers) + 1):
            col_letter = sheet.cell(row=1, column=col_index).column_letter
            sheet.column_dimensions[col_letter].width = 10

        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=len(headers)):
            for cell in row:
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = border

    @classmethod
    def build_subject_trend_sheet(cls, workbook, payload):
        sheet = workbook.create_sheet("科目趋势")
        subjects = payload.get("subjects") or []
        headers = ["日期", "考试名", "总分级排", "总分变化"]
        for subject in subjects:
            headers.extend([f"{subject}级排", f"{subject}级排变化"])

        border = cls._set_table_header(sheet, headers)

        exams = payload.get("exams") or []
        for index, exam in enumerate(exams):
            prev_exam = exams[index - 1] if index > 0 else None

            current_total_rank = cls._normalize_rank(exam.get("grade_total_rank"))
            prev_total_rank = cls._normalize_rank(prev_exam.get("grade_total_rank")) if prev_exam else None

            current_map = {item.get("subject_name"): item for item in (exam.get("scores") or [])}
            prev_map = {item.get("subject_name"): item for item in (prev_exam.get("scores") or [])} if prev_exam else {}

            row = [
                exam.get("exam_date") or "-",
                exam.get("name") or "-",
                current_total_rank if current_total_rank is not None else "-",
                cls._rank_change(prev_total_rank, current_total_rank),
            ]

            for subject in subjects:
                current_subject_rank = cls._normalize_rank((current_map.get(subject) or {}).get("grade_rank"))
                prev_subject_rank = cls._normalize_rank((prev_map.get(subject) or {}).get("grade_rank"))
                row.extend([
                    current_subject_rank if current_subject_rank is not None else "-",
                    cls._rank_change(prev_subject_rank, current_subject_rank),
                ])

            sheet.append(row)

        sheet.freeze_panes = "A2"
        sheet.column_dimensions["A"].width = 14
        sheet.column_dimensions["B"].width = 32
        sheet.column_dimensions["C"].width = 12
        sheet.column_dimensions["D"].width = 12
        for col_index in range(5, len(headers) + 1):
            col_letter = sheet.cell(row=1, column=col_index).column_letter
            # 奇数列为“级排”，偶数列为“级排变化”，后者适当加宽。
            if col_index % 2 == 1:
                sheet.column_dimensions[col_letter].width = 10
            else:
                sheet.column_dimensions[col_letter].width = 13

        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=len(headers)):
            for cell in row:
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = border

        cls._apply_rank_change_conditional_format(sheet, 4)
        for col_index in range(6, len(headers) + 1, 2):
            cls._apply_rank_change_conditional_format(sheet, col_index)

        # 动态拉宽考试名称列。
        exam_name_values = [sheet.cell(row=r, column=2).value for r in range(1, sheet.max_row + 1)]
        sheet.column_dimensions["B"].width = cls._auto_width_for_column(exam_name_values, factor=1.6, min_width=18, max_width=52)

        # 图表：最新考试各科级排变化柱状图
        try:
            exams = payload.get("exams") or []
            # 图表辅助数据写到表格末尾之后，避免覆盖“科目趋势”业务列。
            base_col = len(headers) + 2
            notice_cell = sheet.cell(row=2, column=base_col).coordinate
            if len(exams) < 2:
                cls._write_chart_notice(sheet, notice_cell, "暂无可比较科目趋势")
                return

            latest_exam = exams[-1]
            previous_exam = exams[-2]
            latest_name = latest_exam.get("name") or "最新考试"

            # 生成可视化数据区（可见区域，避免隐藏计算区）。
            header_row = 2
            data_start_row = 3
            sheet.cell(row=header_row, column=base_col, value="科目")
            sheet.cell(row=header_row, column=base_col + 1, value="级排变化")

            visible_subjects = payload.get("subjects") or []
            values = []
            latest_row = len(exams) + 1

            plot_row = data_start_row
            for i, subject in enumerate(visible_subjects, start=0):
                change_col_index = 6 + i * 2
                value = sheet.cell(row=latest_row, column=change_col_index).value
                if not isinstance(value, (int, float)):
                    continue
                sheet.cell(row=plot_row, column=base_col, value=subject)
                sheet.cell(row=plot_row, column=base_col + 1, value=int(value))
                values.append(int(value))
                plot_row += 1

            if len(values) == 0:
                cls._write_chart_notice(sheet, notice_cell, "暂无可比较科目趋势")
                return

            bar = BarChart()
            previous_name = previous_exam.get("name") or "上一场考试"
            # 将“a VS b”备注内嵌到图表标题中
            bar.title = f"{latest_name}各科级排变化\n（{latest_name} VS {previous_name}）"
            bar.y_axis.title = "级排变化"
            bar.x_axis.title = "科目"
            bar.height = 12
            bar.width = 30
            bar.legend = None
            bar.y_axis.number_format = "0"
            max_abs = max(abs(v) for v in values) if values else 1
            # 纵轴间隔动态放大，避免刻度过密；仍保证整数刻度。
            major_unit = max(1, int(round(max_abs / 8)))
            bar.y_axis.majorUnit = major_unit
            # 去掉纵轴网格线，减少视觉干扰。
            bar.y_axis.majorGridlines = None

            min_data_row = data_start_row
            max_data_row = plot_row - 1
            data_ref = Reference(sheet, min_col=base_col + 1, min_row=min_data_row, max_row=max_data_row)
            cat_ref = Reference(sheet, min_col=base_col, min_row=min_data_row, max_row=max_data_row)
            bar.add_data(data_ref, titles_from_data=False)
            bar.set_categories(cat_ref)

            if len(bar.series) >= 1:
                bar.series[0].graphicalProperties.solidFill = cls.CHART_MAIN
                bar.series[0].dLbls = DataLabelList()
                bar.series[0].dLbls.showVal = True
                bar.series[0].dLbls.showCatName = False
                bar.series[0].dLbls.showSerName = False
                bar.series[0].dLbls.showLegendKey = False

                # 按数据点着色：提升(>0)=绿色，下降(<0)=低饱和红，持平=中性色。
                bar.series[0].dPt = []
                for idx, value in enumerate(values):
                    point = DataPoint(idx=idx)
                    if value > 0:
                        point.graphicalProperties.solidFill = cls.CHART_MAIN
                    elif value < 0:
                        point.graphicalProperties.solidFill = cls.CHART_NEGATIVE_SOFT
                    else:
                        point.graphicalProperties.solidFill = cls.CHART_NEUTRAL
                    bar.series[0].dPt.append(point)

            bar_chart_row = sheet.max_row + 3
            sheet.add_chart(bar, f"A{bar_chart_row}")

        except Exception:
            cls._write_chart_notice(sheet, f"A{sheet.max_row + 3}", "图表生成失败，请联系管理员", is_error=True)

    @staticmethod
    def build_filename(student_info, timestamp):
        del timestamp  # 当前命名规则不包含时间戳
        student_name = student_info.get("name") or "未知学生"
        cohort = student_info.get("cohort") or student_info.get("grade_level") or "未知届别"
        class_name = student_info.get("class_name") or "未知班级"

        raw_name = f"{cohort}{class_name}{student_name}个人成绩分析报告"
        safe_name = raw_name.replace("/", "-").replace("\\", "-").strip()
        return f"{safe_name}.xlsx"

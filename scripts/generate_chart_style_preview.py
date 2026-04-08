from pathlib import Path

from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.styles import Font, PatternFill


def _build_base_table(ws):
    headers = ["考试", "日期", "总分", "年级排名", "滚动均分", "语文级排变化", "数学级排变化", "英语级排变化"]
    ws.append(headers)
    rows = [
        ["2024-2025学年上学期期中质量检测", "2024-11-20", 592, 49, 592.0, 0, 0, 0],
        ["2024-2025学年上学期期末联合考试", "2025-01-18", 611, 36, 601.5, 16, 12, 10],
        ["2025-2026学年上学期开学学情诊断", "2025-09-05", 645, 18, 616.0, 18, 19, 13],
        ["2025-2026学年上学期期中质量检测", "2025-11-22", 638, 21, 631.3, -4, -4, -4],
    ]
    for row in rows:
        ws.append(row)

    _auto_fit_columns(ws)
    return rows


def _auto_fit_columns(ws):
    """根据表头与内容动态调宽列宽（中文场景按字符长度估算）。"""
    min_width = 10
    max_width = 48
    for col in ws.iter_cols(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        max_len = 0
        for cell in col:
            if cell.value is None:
                continue
            value_len = len(str(cell.value))
            if value_len > max_len:
                max_len = value_len

        # 中文字符在Excel中显示通常更宽，放大系数做经验补偿。
        estimated = int(max_len * 1.6) + 2
        width = max(min_width, min(max_width, estimated))
        ws.column_dimensions[col[0].column_letter].width = width


def _style_header(ws, color="00897B"):
    fill = PatternFill(fill_type="solid", fgColor=color)
    font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = fill
        cell.font = font


def _add_line_chart(ws, style_name):
    line = LineChart()
    line.title = f"{style_name} - 总分趋势（含滚动均分）"
    line.y_axis.title = "分数"
    line.x_axis.title = "考试"
    line.height = 8
    line.width = 15

    total_ref = Reference(ws, min_col=3, min_row=1, max_row=5)
    rolling_ref = Reference(ws, min_col=5, min_row=1, max_row=5)
    cats = Reference(ws, min_col=1, min_row=2, max_row=5)

    line.add_data(total_ref, titles_from_data=True)
    line.add_data(rolling_ref, titles_from_data=True)
    line.set_categories(cats)

    # 所有趋势图统一显示数值标签与三角形数据点标记。
    line.dataLabels = DataLabelList()
    line.dataLabels.showVal = True
    line.dataLabels.showCatName = False
    line.dataLabels.showSerName = False
    line.dataLabels.showLegendKey = False
    for series in line.series:
        series.marker.symbol = "triangle"
        series.marker.size = 7

    # 线条样式差异
    if style_name == "方案A-清爽教务风":
        line.series[0].graphicalProperties.line.solidFill = "00897B"
        line.series[0].graphicalProperties.line.width = 22000
        line.series[1].graphicalProperties.line.solidFill = "26A69A"
        line.series[1].graphicalProperties.line.width = 18000
    elif style_name == "方案B-数据强调风":
        line.series[0].graphicalProperties.line.solidFill = "00695C"
        line.series[0].graphicalProperties.line.width = 30000
        line.series[1].graphicalProperties.line.solidFill = "1E88E5"
        line.series[1].graphicalProperties.line.width = 26000
    else:
        line.series[0].graphicalProperties.line.solidFill = "455A64"
        line.series[0].graphicalProperties.line.width = 18000
        line.series[1].graphicalProperties.line.solidFill = "90A4AE"
        line.series[1].graphicalProperties.line.width = 16000
        line.legend = None

    ws.add_chart(line, "J2")


def _add_rank_chart(ws, style_name):
    rank = LineChart()
    rank.title = f"{style_name} - 年级排名趋势"
    rank.y_axis.title = "排名"
    rank.x_axis.title = "考试"
    rank.y_axis.scaling.orientation = "maxMin"
    rank.height = 8
    rank.width = 15

    rank_ref = Reference(ws, min_col=4, min_row=1, max_row=5)
    cats = Reference(ws, min_col=1, min_row=2, max_row=5)

    rank.add_data(rank_ref, titles_from_data=True)
    rank.set_categories(cats)

    rank.dataLabels = DataLabelList()
    rank.dataLabels.showVal = True
    rank.dataLabels.showCatName = False
    rank.dataLabels.showSerName = False
    rank.dataLabels.showLegendKey = False
    for series in rank.series:
        series.marker.symbol = "triangle"
        series.marker.size = 7

    if style_name == "方案A-清爽教务风":
        rank.series[0].graphicalProperties.line.solidFill = "1E88E5"
        rank.series[0].graphicalProperties.line.width = 22000
    elif style_name == "方案B-数据强调风":
        rank.series[0].graphicalProperties.line.solidFill = "1565C0"
        rank.series[0].graphicalProperties.line.width = 30000
    else:
        rank.series[0].graphicalProperties.line.solidFill = "607D8B"
        rank.series[0].graphicalProperties.line.width = 18000

    ws.add_chart(rank, "J20")


def _add_subject_bar(ws, style_name, exam_rows):
    # 最新一场科目级排变化使用第5行
    ws.append([])
    ws.append(["科目", "变化"])
    ws.append(["语文", ws["F5"].value])
    ws.append(["数学", ws["G5"].value])
    ws.append(["英语", ws["H5"].value])

    base_row = 8
    exam_a = exam_rows[-1][0] if len(exam_rows) >= 1 else "当前考试"

    bar = BarChart()
    bar.title = f"{style_name} - {exam_a}各科级排变化"
    bar.y_axis.title = "级排变化"
    bar.x_axis.title = "科目"
    bar.y_axis.number_format = "0"
    bar.y_axis.majorUnit = 1
    bar.legend = None
    bar.height = 8
    bar.width = 15

    data_ref = Reference(ws, min_col=2, min_row=base_row + 1, max_row=base_row + 3)
    cats = Reference(ws, min_col=1, min_row=base_row + 1, max_row=base_row + 3)

    bar.add_data(data_ref, titles_from_data=False)
    bar.set_categories(cats)
    bar.dataLabels = DataLabelList()
    bar.dataLabels.showVal = True
    bar.dataLabels.showCatName = False
    bar.dataLabels.showSerName = False
    bar.dataLabels.showLegendKey = False

    if style_name == "方案A-清爽教务风":
        bar.series[0].graphicalProperties.solidFill = "00897B"
    elif style_name == "方案B-数据强调风":
        bar.series[0].graphicalProperties.solidFill = "43A047"
        bar.dataLabels = DataLabelList()
        bar.dataLabels.showVal = True
    else:
        bar.series[0].graphicalProperties.solidFill = "78909C"

    ws.add_chart(bar, "J38")


def build_preview_workbook(output_path: Path):
    wb = Workbook()
    wb.remove(wb.active)

    styles = [
        ("方案A-清爽教务风", "00897B"),
        ("方案B-数据强调风", "00695C"),
        ("方案C-极简商务风", "546E7A"),
    ]

    for sheet_name, header_color in styles:
        ws = wb.create_sheet(sheet_name)
        exam_rows = _build_base_table(ws)
        _style_header(ws, header_color)
        _add_line_chart(ws, sheet_name)
        _add_rank_chart(ws, sheet_name)
        _add_subject_bar(ws, sheet_name, exam_rows)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


if __name__ == "__main__":
    root = Path(__file__).resolve().parents[1]
    target = root / "docs" / "个人成绩分析导出" / "图表样式预览.xlsx"
    build_preview_workbook(target)
    print(f"generated: {target}")
